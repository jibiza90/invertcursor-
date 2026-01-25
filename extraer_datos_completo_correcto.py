from openpyxl import load_workbook
import json
from datetime import datetime

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

def procesar_hoja(nombre_hoja):
    """Procesa una hoja completa con datos generales y por cliente"""
    if nombre_hoja not in wb.sheetnames:
        return None
    
    ws = wb[nombre_hoja]
    
    resultado = {
        'nombre': nombre_hoja,
        'datos_generales': [],
        'clientes': []
    }
    
    # 1. EXTRAER DATOS GENERALES (columnas E a J + fecha columna D)
    # Las filas 3-6 tienen los datos generales
    print(f"\nProcesando datos generales de {nombre_hoja}...")
    
    filas_generales = [3, 4, 5, 6]
    for fila_idx in filas_generales:
        fila_data = {
            'fila': fila_idx,
            'fecha': None,  # Columna D (4)
            'columna_e': None,  # Columna E (5)
            'columna_f': None,  # Columna F (6)
            'columna_g': None,  # Columna G (7)
            'columna_h': None,  # Columna H (8)
            'columna_i': None,  # Columna I (9)
            'columna_j': None   # Columna J (10)
        }
        
        # Leer fecha (columna D = 4)
        fecha_valor = ws.cell(row=fila_idx, column=4).value
        fila_data['fecha'] = fecha_valor
        
        # Leer columnas E a J (5 a 10)
        for col_idx in range(5, 11):
            valor = ws.cell(row=fila_idx, column=col_idx).value
            if col_idx == 5:
                fila_data['columna_e'] = valor
            elif col_idx == 6:
                fila_data['columna_f'] = valor
            elif col_idx == 7:
                fila_data['columna_g'] = valor
            elif col_idx == 8:
                fila_data['columna_h'] = valor
            elif col_idx == 9:
                fila_data['columna_i'] = valor
            elif col_idx == 10:
                fila_data['columna_j'] = valor
        
        # Solo agregar si tiene datos
        if any(v is not None for v in [fila_data['fecha'], fila_data['columna_e'], 
                                        fila_data['columna_f'], fila_data['columna_g'],
                                        fila_data['columna_h'], fila_data['columna_i'],
                                        fila_data['columna_j']]):
            resultado['datos_generales'].append(fila_data)
    
    print(f"  Datos generales: {len(resultado['datos_generales'])} filas")
    
    # 2. IDENTIFICAR CLIENTES (fila 13 tiene los números de clientes)
    print(f"\nIdentificando clientes...")
    
    fila_nombres = 13
    clientes_columnas = {}
    
    for col_idx in range(1, ws.max_column + 1):
        celda = ws.cell(row=fila_nombres, column=col_idx)
        valor = celda.value
        
        if valor is not None:
            try:
                numero_cliente = int(valor)
                if numero_cliente > 0:
                    # Este cliente empieza en esta columna
                    grupo_columnas = {
                        'numero': numero_cliente,
                        'columna_inicio': col_idx,
                        'INCREMENTO': col_idx,      # K (11 para cliente 1)
                        'DECREMENTO': col_idx + 1,  # L (12 para cliente 1)
                        'FINAL': col_idx + 3,       # N (14 para cliente 1)
                        'BENEF_€_DIARIO': col_idx + 4,  # O (15 para cliente 1)
                        'BENEF_%_DIARIO': col_idx + 5,  # P (16 para cliente 1)
                        'BENEF_€_ACUM': col_idx + 6,    # Q (17 para cliente 1)
                        'BENEF_%_ACUM': col_idx + 7     # R (18 para cliente 1)
                    }
                    
                    clientes_columnas[numero_cliente] = grupo_columnas
            except:
                pass
    
    print(f"  Clientes encontrados: {len(clientes_columnas)}")
    
    # 3. EXTRAER DATOS DIARIOS POR CLIENTE (desde fila 15)
    print(f"\nExtrayendo datos diarios por cliente...")
    
    for numero_cliente, grupo_columnas in clientes_columnas.items():
        cliente_data = {
            'numero_cliente': numero_cliente,
            'columna_inicio': grupo_columnas['columna_inicio'],
            'incrementos_total': 0,
            'decrementos_total': 0,
            'saldo_actual': None,
            'datos_diarios': []  # Lista de {fecha, saldo_diario, beneficio_diario, etc.}
        }
        
        incrementos = []
        decrementos = []
        saldos_finales = []
        
        # Leer todas las filas desde la 15 en adelante
        for fila_idx in range(15, ws.max_row + 1):
            # Leer fecha (columna D = 4)
            fecha_valor = ws.cell(row=fila_idx, column=4).value
            
            # Leer valores del cliente
            incremento = ws.cell(row=fila_idx, column=grupo_columnas['INCREMENTO']).value
            decremento = ws.cell(row=fila_idx, column=grupo_columnas['DECREMENTO']).value
            final = ws.cell(row=fila_idx, column=grupo_columnas['FINAL']).value
            benef_diario = ws.cell(row=fila_idx, column=grupo_columnas['BENEF_€_DIARIO']).value
            benef_diario_pct = ws.cell(row=fila_idx, column=grupo_columnas['BENEF_%_DIARIO']).value
            benef_acum = ws.cell(row=fila_idx, column=grupo_columnas['BENEF_€_ACUM']).value
            benef_acum_pct = ws.cell(row=fila_idx, column=grupo_columnas['BENEF_%_ACUM']).value
            
            # Si hay fecha, es una fila de datos diarios
            if fecha_valor is not None:
                dato_diario = {
                    'fila': fila_idx,
                    'fecha': fecha_valor,
                    'saldo_diario': final,
                    'beneficio_diario': benef_diario,
                    'beneficio_diario_pct': benef_diario_pct,
                    'beneficio_acumulado': benef_acum,
                    'beneficio_acumulado_pct': benef_acum_pct
                }
                cliente_data['datos_diarios'].append(dato_diario)
            
            # Acumular incrementos y decrementos
            if incremento is not None:
                try:
                    inc_val = float(incremento)
                    incrementos.append(inc_val)
                except:
                    pass
            
            if decremento is not None:
                try:
                    dec_val = float(decremento)
                    decrementos.append(dec_val)
                except:
                    pass
            
            # Acumular saldos finales (columna N)
            if final is not None:
                try:
                    final_val = float(final)
                    saldos_finales.append((fila_idx, final_val))
                except:
                    pass
        
        # Calcular totales
        cliente_data['incrementos_total'] = sum(incrementos)
        cliente_data['decrementos_total'] = sum(decrementos)
        
        # Saldo actual: último valor no cero de columna N
        if saldos_finales:
            # Buscar último valor no cero
            for fila_idx, valor in reversed(saldos_finales):
                if valor != 0:
                    cliente_data['saldo_actual'] = valor
                    cliente_data['saldo_actual_fila'] = fila_idx
                    break
            # Si no hay valores no cero, tomar el último
            if cliente_data['saldo_actual'] is None:
                cliente_data['saldo_actual'] = saldos_finales[-1][1]
                cliente_data['saldo_actual_fila'] = saldos_finales[-1][0]
        
        resultado['clientes'].append(cliente_data)
    
    # Ordenar clientes por número
    resultado['clientes'].sort(key=lambda x: x['numero_cliente'])
    
    print(f"  Clientes procesados: {len(resultado['clientes'])}")
    if resultado['clientes']:
        cliente1 = resultado['clientes'][0]
        print(f"  Cliente 1 - Datos diarios: {len(cliente1['datos_diarios'])} registros")
        print(f"  Cliente 1 - Saldo actual: {cliente1['saldo_actual']}")
    
    return resultado

# Procesar ambas hojas
resultado_completo = {
    'fecha_extraccion': datetime.now().isoformat(),
    'hojas': {}
}

print("=" * 80)
print("EXTRACCION COMPLETA DE DATOS")
print("=" * 80)

# Procesar Diario STD
if 'Diario STD' in wb.sheetnames:
    resultado_completo['hojas']['Diario STD'] = procesar_hoja('Diario STD')

# Procesar Diario VIP
if 'Diario VIP' in wb.sheetnames:
    resultado_completo['hojas']['Diario VIP'] = procesar_hoja('Diario VIP')

# Guardar resultado
with open('datos_completos.json', 'w', encoding='utf-8') as f:
    json.dump(resultado_completo, f, ensure_ascii=False, indent=2, default=str)

print("\n" + "=" * 80)
print(f"Datos guardados en: datos_completos.json")
print("=" * 80)
