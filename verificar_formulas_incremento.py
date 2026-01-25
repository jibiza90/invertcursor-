from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
ws_formulas = wb_formulas['Diario VIP']
ws_valores = wb_valores['Diario VIP']

print("=" * 80)
print("VERIFICACION FORMULAS CON INCREMENTOS - CLIENTE 1")
print("=" * 80)

# Verificar filas donde hay incrementos
filas_verificar = [15, 16, 17, 18, 19, 20]

for fila_idx in filas_verificar:
    fecha_celda = ws_valores.cell(row=fila_idx, column=4).value
    incremento_valor = ws_valores.cell(row=fila_idx, column=11).value  # K
    
    print(f"\n{'='*80}")
    print(f"FILA {fila_idx}:")
    print(f"  Fecha: {fecha_celda}")
    print(f"  K{fila_idx} (INCREMENTO valor): {incremento_valor}")
    print(f"  L{fila_idx} (DECREMENTO valor): {ws_valores.cell(row=fila_idx, column=12).value}")
    print(f"  M{fila_idx} (BASE valor): {ws_valores.cell(row=fila_idx, column=13).value}")
    print(f"  N{fila_idx} (SALDO_DIARIO valor): {ws_valores.cell(row=fila_idx, column=14).value}")
    
    # Verificar fórmulas en M (base)
    celda_m = ws_formulas.cell(row=fila_idx, column=13)  # M
    if celda_m.data_type == 'f' and celda_m.value:
        print(f"\n  FORMULA M{fila_idx} (BASE):")
        print(f"    {celda_m.value}")
        
        # Extraer referencias
        import re
        referencias = re.findall(r'([A-Z]+\d+)', str(celda_m.value))
        print(f"    Referencias encontradas: {referencias}")
        
        # Mostrar valores de las referencias
        for ref in referencias:
            col_letra = re.match(r'([A-Z]+)', ref).group(1)
            fila_num = int(re.match(r'\d+', ref).group(0))
            col_num = 0
            for i, ch in enumerate(col_letra):
                col_num = col_num * 26 + (ord(ch) - 64)
            
            valor_ref = ws_valores.cell(row=fila_num, column=col_num).value
            print(f"      {ref} = {valor_ref}")
    
    # Verificar fórmulas en N (saldo_diario)
    celda_n = ws_formulas.cell(row=fila_idx, column=14)  # N
    if celda_n.data_type == 'f' and celda_n.value:
        print(f"\n  FORMULA N{fila_idx} (SALDO_DIARIO):")
        print(f"    {celda_n.value}")
        
        import re
        referencias = re.findall(r'([A-Z]+\d+)', str(celda_n.value))
        print(f"    Referencias encontradas: {referencias}")
        
        for ref in referencias:
            col_letra = re.match(r'([A-Z]+)', ref).group(1)
            fila_num = int(re.match(r'\d+', ref).group(0))
            col_num = 0
            for i, ch in enumerate(col_letra):
                col_num = col_num * 26 + (ord(ch) - 64)
            
            valor_ref = ws_valores.cell(row=fila_num, column=col_num).value
            print(f"      {ref} = {valor_ref}")
