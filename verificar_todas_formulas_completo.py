from openpyxl import load_workbook
import json

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("VERIFICACION COMPLETA - TODAS LAS FORMULAS")
print("=" * 80)

# Cargar JSON generado
with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos_json = json.load(f)

diarios = datos_json['hojas']['Diario VIP']['datos_diarios_generales']

columnas = {
    5: 'imp_inicial',
    6: 'imp_final',
    7: 'benef_euro',
    8: 'benef_porcentaje',
    9: 'benef_euro_acum',
    10: 'benef_porcentaje_acum'
}

errores = []

for fila_idx in range(15, min(400, ws.max_row + 1)):
    fecha_celda = ws.cell(row=fila_idx, column=4)
    if fecha_celda.value is None:
        continue
    
    # Buscar en JSON
    fila_json = None
    for d in diarios:
        if d['fila'] == fila_idx:
            fila_json = d
            break
    
    if not fila_json:
        continue
    
    bloqueadas_json = fila_json.get('bloqueadas', {})
    
    # Verificar cada columna
    for col_idx, col_key in columnas.items():
        celda = ws.cell(row=fila_idx, column=col_idx)
        
        # Verificar si tiene fórmula en Excel
        tiene_formula_excel = False
        if celda.data_type == 'f':
            tiene_formula_excel = True
        elif celda.value is not None:
            valor_str = str(celda.value).strip()
            if valor_str.startswith('='):
                tiene_formula_excel = True
        
        # Verificar en JSON
        esta_bloqueada_json = bloqueadas_json.get(col_key, False)
        
        # Si hay discrepancia
        if tiene_formula_excel != esta_bloqueada_json:
            fecha_str = str(fecha_celda.value)[:10] if fecha_celda.value else 'N/A'
            errores.append({
                'fila': fila_idx,
                'fecha': fecha_str,
                'columna': col_key,
                'tiene_formula_excel': tiene_formula_excel,
                'bloqueada_json': esta_bloqueada_json,
                'formula': str(celda.value)[:60] if tiene_formula_excel else None
            })

print(f"\nTotal de discrepancias encontradas: {len(errores)}")

if errores:
    print("\n" + "=" * 80)
    print("DISCREPANCIAS ENCONTRADAS:")
    print("=" * 80)
    
    for error in errores[:30]:  # Mostrar primeras 30
        estado_excel = "[BLOQUEADA]" if error['tiene_formula_excel'] else "[EDITABLE]"
        estado_json = "[BLOQUEADA]" if error['bloqueada_json'] else "[EDITABLE]"
        print(f"\nFila {error['fila']} ({error['fecha']}), Columna {error['columna']}:")
        print(f"  Excel: {estado_excel}")
        print(f"  JSON:  {estado_json}")
        if error['formula']:
            print(f"  Formula: {error['formula']}")
    
    if len(errores) > 30:
        print(f"\n... y {len(errores) - 30} discrepancias más")
else:
    print("\n✓ No se encontraron discrepancias. Todo está correcto!")
