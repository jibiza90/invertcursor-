from openpyxl import load_workbook
import re

# Verificar cómo funciona realmente AEO en Excel
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario STD']

print("Verificando cómo funciona AEO16:")
print("=" * 80)

# Fila 15 tiene imp_inicial = AEO16
fila_15 = ws.cell(row=15, column=5).value  # Columna E
print(f"Fila 15 imp_inicial: {fila_15}")

# Verificar qué hay en la columna AEO (157) fila 16
aeo_16_formula = ws.cell(row=16, column=157).value  # Columna AEO
print(f"Fila 16 columna AEO (157): {aeo_16_formula}")

# Verificar incrementos del cliente 1 en filas 15 y 16
inc_15 = ws.cell(row=15, column=11).value  # Cliente 1 columna K
inc_16 = ws.cell(row=16, column=11).value  # Cliente 1 columna K
print(f"Incremento cliente 1 fila 15: {inc_15}")
print(f"Incremento cliente 1 fila 16: {inc_16}")

# Verificar si AEO16 tiene una fórmula que suma incrementos
if aeo_16_formula and str(aeo_16_formula).startswith('='):
    print(f"\nAEO16 tiene fórmula: {aeo_16_formula}")
    # Buscar referencias a columnas K (incrementos)
    k_refs = re.findall(r'K\d+', str(aeo_16_formula))
    l_refs = re.findall(r'L\d+', str(aeo_16_formula))
    print(f"Referencias a K (incrementos): {k_refs}")
    print(f"Referencias a L (decrementos): {l_refs}")
    
    # Verificar si usa K15 o K16
    if 'K15' in str(aeo_16_formula):
        print("✅ AEO16 usa incrementos de la fila 15")
    elif 'K16' in str(aeo_16_formula):
        print("❌ AEO16 usa incrementos de la fila 16")
    else:
        print("⚠️ AEO16 no usa directamente K15 o K16")
