#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para verificar paso a paso cómo se calculan las fórmulas del cliente 2
y comparar con lo que debería ser según Excel
"""

import json
import openpyxl
import re

# Cargar Excel
wb_formulas = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

ws_std_formulas = None
ws_std_valores = None
for sheet_name in wb_formulas.sheetnames:
    if 'STD' in sheet_name.upper() and 'VIP' not in sheet_name.upper():
        ws_std_formulas = wb_formulas[sheet_name]
        ws_std_valores = wb_valores[sheet_name]
        print(f"Hoja encontrada: {sheet_name}")
        break

# Cargar JSON
with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos = json.load(f)

hoja_std = datos.get('hojas', {}).get('Diario STD')

print("="*80)
print("VERIFICACIÓN PASO A PASO DE FÓRMULAS DEL CLIENTE 2")
print("="*80)

# Cliente 2 empieza en columna S (19)
# Cliente 1 empieza en columna K (11)
# Cliente 2: S=19 (incremento), T=20 (decremento), U=21 (base), V=22 (saldo_diario), 
#            W=23 (beneficio_diario), X=24 (beneficio_diario_pct), Y=25 (beneficio_acumulado)

clientes = hoja_std.get('clientes', [])
if len(clientes) < 2:
    print("ERROR: No hay suficientes clientes")
    exit(1)

cliente2 = clientes[1]
datos_diarios_cliente2 = cliente2.get('datos_diarios', [])

# Buscar fila 17 del cliente 2
dato17_cliente2 = [d for d in datos_diarios_cliente2 if d.get('fila') == 17]
if not dato17_cliente2:
    print("ERROR: No se encontró fila 17 del cliente 2")
    exit(1)

dato17 = dato17_cliente2[0]

print(f"\nFila 17 del Cliente 2:")
print(f"  incremento (S15): {dato17.get('incremento')}")
print(f"  decremento (T16): {dato17.get('decremento')}")

# Ver fórmulas
formulas = dato17.get('formulas', {})
print(f"\nFórmulas en JSON:")
for campo, formula in formulas.items():
    print(f"  {campo}: {formula}")

# Simular el cálculo paso a paso
print(f"\n{'='*80}")
print("SIMULACIÓN DE CÁLCULO PASO A PASO:")
print(f"{'='*80}")

# 1. Verificar valores necesarios para base = IF(N17<>0,S15-T16,0)
print(f"\n1. Cálculo de BASE (U17) = IF(N17<>0,S15-T16,0):")
print(f"   - N17 es la columna N (14), que es saldo_diario del Cliente 1 en fila 17")

# Buscar cliente 1
cliente1 = clientes[0]
datos_diarios_cliente1 = cliente1.get('datos_diarios', [])
dato17_cliente1 = [d for d in datos_diarios_cliente1 if d.get('fila') == 17]
if dato17_cliente1:
    saldo_diario_cliente1_fila17 = dato17_cliente1[0].get('saldo_diario')
    print(f"   - N17 (saldo_diario Cliente 1, fila 17) = {saldo_diario_cliente1_fila17}")
else:
    print(f"   - ERROR: No se encontró fila 17 del Cliente 1")
    saldo_diario_cliente1_fila17 = None

# Buscar incremento y decremento del cliente 2 en filas 15 y 16
dato15_cliente2 = [d for d in datos_diarios_cliente2 if d.get('fila') == 15]
dato16_cliente2 = [d for d in datos_diarios_cliente2 if d.get('fila') == 16]

incremento_fila15 = dato15_cliente2[0].get('incremento') if dato15_cliente2 else None
decremento_fila16 = dato16_cliente2[0].get('decremento') if dato16_cliente2 else None

print(f"   - S15 (incremento Cliente 2, fila 15) = {incremento_fila15}")
print(f"   - T16 (decremento Cliente 2, fila 16) = {decremento_fila16}")

# Calcular base
if saldo_diario_cliente1_fila17 and saldo_diario_cliente1_fila17 != 0:
    inc = incremento_fila15 or 0
    dec = decremento_fila16 or 0
    base_calculada = inc - dec
    print(f"   - Cálculo: IF({saldo_diario_cliente1_fila17}<>0, {inc}-{dec}, 0) = {base_calculada}")
else:
    base_calculada = 0
    print(f"   - Cálculo: IF({saldo_diario_cliente1_fila17}<>0, {incremento_fila15}-{decremento_fila16}, 0) = 0 (porque N17 es 0 o null)")

print(f"   - Valor actual en JSON: {dato17.get('base')}")
print(f"   - ¿Coinciden? {abs((dato17.get('base') or 0) - base_calculada) < 0.01}")

# 2. Verificar valores de la vista general necesarios
print(f"\n2. Valores de la vista general necesarios:")
datos_diarios_generales = hoja_std.get('datos_diarios_generales', [])
dato17_general = [d for d in datos_diarios_generales if d.get('fila') == 17]
if dato17_general:
    f17 = dato17_general[0]
    print(f"   - F17 (imp_final) = {f17.get('imp_final')}")
    print(f"   - H17 (benef_porcentaje) = {f17.get('benef_porcentaje')}")
    
    # beneficio_diario_pct = IF(U17<>0,P17,0) donde P17 es benef_porcentaje (H17)
    benef_porcentaje = f17.get('benef_porcentaje') or 0
    if base_calculada and base_calculada != 0:
        beneficio_diario_pct_calculado = benef_porcentaje
        print(f"   - beneficio_diario_pct = IF({base_calculada}<>0, {benef_porcentaje}, 0) = {beneficio_diario_pct_calculado}")
    else:
        beneficio_diario_pct_calculado = 0
        print(f"   - beneficio_diario_pct = IF({base_calculada}<>0, {benef_porcentaje}, 0) = 0")
else:
    print(f"   - ERROR: No se encontró fila 17 en datos generales")
    benef_porcentaje = 0
    beneficio_diario_pct_calculado = 0

# 3. Calcular beneficio_diario = U17*X17
print(f"\n3. Cálculo de BENEFICIO_DIARIO (W17) = U17*X17:")
beneficio_diario_calculado = base_calculada * beneficio_diario_pct_calculado
print(f"   - Cálculo: {base_calculada} * {beneficio_diario_pct_calculado} = {beneficio_diario_calculado}")
print(f"   - Valor actual en JSON: {dato17.get('beneficio_diario')}")
print(f"   - ¿Coinciden? {abs((dato17.get('beneficio_diario') or 0) - beneficio_diario_calculado) < 0.01}")

# 4. Calcular saldo_diario = IF(U17<>0,U17+W17,0)
print(f"\n4. Cálculo de SALDO_DIARIO (V17) = IF(U17<>0,U17+W17,0):")
if base_calculada and base_calculada != 0:
    saldo_diario_calculado = base_calculada + beneficio_diario_calculado
    print(f"   - Cálculo: IF({base_calculada}<>0, {base_calculada}+{beneficio_diario_calculado}, 0) = {saldo_diario_calculado}")
else:
    saldo_diario_calculado = 0
    print(f"   - Cálculo: IF({base_calculada}<>0, {base_calculada}+{beneficio_diario_calculado}, 0) = 0")
print(f"   - Valor actual en JSON: {dato17.get('saldo_diario')}")
print(f"   - ¿Coinciden? {abs((dato17.get('saldo_diario') or 0) - saldo_diario_calculado) < 0.01}")

# 5. Calcular beneficio_acumulado = IF(V17<>0,W17,0)
print(f"\n5. Cálculo de BENEFICIO_ACUMULADO (Y17) = IF(V17<>0,W17,0):")
if saldo_diario_calculado and saldo_diario_calculado != 0:
    beneficio_acumulado_calculado = beneficio_diario_calculado
    print(f"   - Cálculo: IF({saldo_diario_calculado}<>0, {beneficio_diario_calculado}, 0) = {beneficio_acumulado_calculado}")
else:
    beneficio_acumulado_calculado = 0
    print(f"   - Cálculo: IF({saldo_diario_calculado}<>0, {beneficio_diario_calculado}, 0) = 0")
print(f"   - Valor actual en JSON: {dato17.get('beneficio_acumulado')}")
print(f"   - ¿Coinciden? {abs((dato17.get('beneficio_acumulado') or 0) - beneficio_acumulado_calculado) < 0.01}")

print(f"\n{'='*80}")
print("RESUMEN:")
print(f"{'='*80}")
print(f"Si pones incremento=1000 en fila 15 del Cliente 2:")
print(f"  - base debería ser: {1000 - (decremento_fila16 or 0)} (si N17<>0)")
print(f"  - saldo_diario debería ser: base + beneficio_diario")
print(f"  - beneficio_diario debería ser: base * beneficio_diario_pct")
print(f"  - beneficio_diario_pct debería ser: H17 (benef_porcentaje de vista general)")

wb_formulas.close()
wb_valores.close()
