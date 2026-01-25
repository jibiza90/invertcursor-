from openpyxl import load_workbook
import json
import sys

# Redirigir salida a archivo
sys.stdout = open('verificacion_salida.txt', 'w', encoding='utf-8')
sys.stderr = sys.stdout

print("=" * 80)
print("VERIFICACION CELDA A CELDA: EXCEL vs JSON")
print("=" * 80)

# Cargar Excel
print("Cargando Excel...")
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
ws_formulas = wb_formulas['Diario VIP']
ws_valores = wb_valores['Diario VIP']
print("✓ Excel cargado")

# Cargar JSON
print("Cargando JSON...")
with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos_json = json.load(f)

if 'Diario VIP' not in datos_json.get('hojas', {}):
    print("ERROR: No se encuentra 'Diario VIP' en el JSON")
    print(f"Hojas disponibles: {list(datos_json.get('hojas', {}).keys())}")
    sys.exit(1)

hoja_json = datos_json['hojas']['Diario VIP']
clientes_json = hoja_json.get('clientes', [])
print(f"✓ JSON cargado - {len(clientes_json)} clientes encontrados")

# Mapeo de columnas
columna_a_campo_cliente = {
    11: 'incremento',
    12: 'decremento',
    13: 'base',
    14: 'saldo_diario',
    15: 'beneficio_diario',
    16: 'beneficio_diario_pct',
    17: 'beneficio_acumulado',
    18: 'beneficio_acumulado_pct'
}

columna_a_campo_general = {
    5: 'imp_inicial',
    6: 'imp_final',
    7: 'benef_euro',
    8: 'benef_porcentaje',
    9: 'benef_euro_acum',
    10: 'benef_porcentaje_acum'
}

errores = []
advertencias = []

# Verificar Cliente 1 - primeras filas
print("\n" + "=" * 80)
print("VERIFICANDO CLIENTE 1 (Columnas K-R)")
print("=" * 80)

cliente1_json = next((c for c in clientes_json if c.get('numero_cliente') == 1), None)

if not cliente1_json:
    print("ERROR: Cliente 1 no encontrado en JSON")
    sys.exit(1)

datos_diarios_cliente = cliente1_json.get('datos_diarios', [])
print(f"Encontradas {len(datos_diarios_cliente)} filas de datos del cliente 1")

# Verificar primeras 20 filas con datos
filas_verificar = []
for fila_excel in range(15, 35):
    fecha_excel = ws_valores.cell(row=fila_excel, column=4).value
    if fecha_excel is not None:
        filas_verificar.append(fila_excel)

print(f"\nVerificando {len(filas_verificar)} filas: {filas_verificar[:10]}...")

for fila_excel in filas_verificar:
    fecha_excel = ws_valores.cell(row=fila_excel, column=4).value
    
    fila_data_json = next((d for d in datos_diarios_cliente if d['fila'] == fila_excel), None)
    
    if not fila_data_json:
        advertencias.append(f"Cliente 1, Fila {fila_excel}: No encontrada en JSON")
        continue
    
    print(f"\n--- Fila {fila_excel} (Fecha: {fecha_excel}) ---")
    
    for col_excel, campo in columna_a_campo_cliente.items():
        celda_excel = ws_formulas.cell(row=fila_excel, column=col_excel)
        tiene_formula_excel = celda_excel.data_type == 'f' and celda_excel.value
        
        tiene_formula_json = fila_data_json.get('formulas', {}).get(campo, None)
        esta_bloqueada_json = fila_data_json.get('bloqueadas', {}).get(campo, False)
        valor_excel = ws_valores.cell(row=fila_excel, column=col_excel).value
        
        col_letra = chr(64 + col_excel) if col_excel <= 26 else '?'
        
        if tiene_formula_excel:
            formula_excel = str(celda_excel.value)
            print(f"  {col_letra}{fila_excel} ({campo}): TIENE FORMULA")
            print(f"    Excel: {formula_excel[:80]}")
            
            if tiene_formula_json:
                formula_json = tiene_formula_json
                if formula_excel != formula_json:
                    errores.append(f"Cliente 1, Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Fórmula diferente")
                    errores.append(f"  Excel: {formula_excel}")
                    errores.append(f"  JSON:  {formula_json}")
                    print(f"    ❌ DIFERENTE - JSON: {formula_json[:80]}")
                else:
                    print(f"    ✅ Coincide con JSON")
            else:
                errores.append(f"Cliente 1, Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Tiene fórmula en Excel pero NO en JSON")
                print(f"    ❌ FALTA EN JSON")
            
            if not esta_bloqueada_json:
                errores.append(f"Cliente 1, Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Tiene fórmula pero NO está bloqueada en JSON")
                print(f"    ❌ NO BLOQUEADA (debería estar bloqueada)")
            else:
                print(f"    ✅ Bloqueada correctamente")
        else:
            # No tiene fórmula
            if campo in ['incremento', 'decremento']:
                # Deben ser editables (no bloqueadas)
                if esta_bloqueada_json:
                    errores.append(f"Cliente 1, Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): NO debe estar bloqueada (es editable)")
                    print(f"  {col_letra}{fila_excel} ({campo}): ❌ BLOQUEADA (no debería)")
                else:
                    print(f"  {col_letra}{fila_excel} ({campo}): ✅ Editable (valor: {valor_excel})")
            else:
                # Otros campos pueden estar bloqueados si tienen valor
                if esta_bloqueada_json:
                    if valor_excel is None or valor_excel == 0:
                        advertencias.append(f"Cliente 1, Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Bloqueada pero sin valor")
                    print(f"  {col_letra}{fila_excel} ({campo}): Bloqueada (valor: {valor_excel})")
                else:
                    print(f"  {col_letra}{fila_excel} ({campo}): Editable (valor: {valor_excel})")

# Resumen
print("\n" + "=" * 80)
print("RESUMEN")
print("=" * 80)
print(f"Errores encontrados: {len(errores)}")
print(f"Advertencias: {len(advertencias)}")

if errores:
    print("\n❌ ERRORES:")
    for error in errores[:30]:
        print(f"  - {error}")
    if len(errores) > 30:
        print(f"  ... y {len(errores) - 30} errores más")

if advertencias:
    print("\n⚠️  ADVERTENCIAS:")
    for adv in advertencias[:20]:
        print(f"  - {adv}")
    if len(advertencias) > 20:
        print(f"  ... y {len(advertencias) - 20} advertencias más")

sys.stdout.close()
