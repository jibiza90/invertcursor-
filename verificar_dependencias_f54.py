from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("BUSCANDO FORMULAS QUE DEPENDEN DE F54")
print("=" * 80)

# Buscar todas las fórmulas que referencian F54
print("\nBuscando referencias a F54 en todas las fórmulas...")

referencias_f54 = []

# Verificar datos generales
for fila_idx in range(15, min(200, ws.max_row + 1)):
    for col_idx in range(5, 11):  # E-J
        celda = ws.cell(row=fila_idx, column=col_idx)
        if celda.data_type == 'f' and celda.value:
            formula = str(celda.value)
            if 'F54' in formula:
                col_nombre = ['E', 'F', 'G', 'H', 'I', 'J'][col_idx - 5]
                referencias_f54.append({
                    'tipo': 'general',
                    'fila': fila_idx,
                    'columna': col_nombre,
                    'formula': formula[:100]
                })

# Verificar clientes (columnas K en adelante)
for fila_idx in range(15, min(200, ws.max_row + 1)):
    for col_idx in range(11, 19):  # K-R (cliente 1)
        celda = ws.cell(row=fila_idx, column=col_idx)
        if celda.data_type == 'f' and celda.value:
            formula = str(celda.value)
            if 'F54' in formula:
                col_nombre = ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R'][col_idx - 11]
                referencias_f54.append({
                    'tipo': 'cliente',
                    'fila': fila_idx,
                    'columna': col_nombre,
                    'formula': formula[:100]
                })

if referencias_f54:
    print(f"\nEncontradas {len(referencias_f54)} fórmulas que dependen de F54:")
    for ref in referencias_f54:
        print(f"  {ref['tipo']} - Fila {ref['fila']}, Col {ref['columna']}: {ref['formula']}")
else:
    print("\nNo se encontraron fórmulas que dependan directamente de F54")
    print("Pero F54 puede afectar indirectamente a través de otras fórmulas")

# Verificar si E56 depende de F54 (patrón común: E56 = F54 + ...)
print("\nVerificando si E56 depende de F54:")
celda_e56 = ws.cell(row=56, column=5)
if celda_e56.data_type == 'f' and celda_e56.value:
    formula_e56 = str(celda_e56.value)
    print(f"  E56: {formula_e56}")
    if 'F54' in formula_e56:
        print("  -> SÍ depende de F54")
    else:
        print("  -> NO depende directamente de F54")
