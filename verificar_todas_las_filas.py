from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("VERIFICACION COMPLETA - TODAS LAS FILAS 15-1130")
print("=" * 80)

columnas_beneficios = {
    7: 'G - Benef. €',
    8: 'H - Benef. %',
    9: 'I - Benef. € Acum',
    10: 'J - Benef. % Acum'
}

filas_con_formulas = []
filas_sin_formulas = []

for fila_idx in range(15, min(1131, ws.max_row + 1)):
    fecha_celda = ws.cell(row=fila_idx, column=4)
    if fecha_celda.value is None:
        continue
    
    tiene_formulas_en_fila = False
    formulas_en_fila = {}
    
    for col_idx, nombre_col in columnas_beneficios.items():
        celda = ws.cell(row=fila_idx, column=col_idx)
        
        tiene_formula = False
        formula_texto = None
        
        if celda.data_type == 'f':
            tiene_formula = True
            formula_texto = str(celda.value)
            tiene_formulas_en_fila = True
        elif celda.value is not None:
            valor_str = str(celda.value).strip()
            if valor_str.startswith('='):
                tiene_formula = True
                formula_texto = valor_str
                tiene_formulas_en_fila = True
        
        if tiene_formula:
            formulas_en_fila[nombre_col] = formula_texto
    
    if tiene_formulas_en_fila:
        filas_con_formulas.append((fila_idx, formulas_en_fila))
    else:
        filas_sin_formulas.append(fila_idx)

print(f"\nTotal de filas con formulas en G-H-I-J: {len(filas_con_formulas)}")
print(f"Total de filas SIN formulas en G-H-I-J: {len(filas_sin_formulas)}")

if filas_sin_formulas:
    print(f"\nPrimeras 20 filas SIN formulas:")
    for fila_idx in filas_sin_formulas[:20]:
        fecha_valor = ws.cell(row=fila_idx, column=4).value
        fecha_str = str(fecha_valor).split()[0] if fecha_valor else 'N/A'
        print(f"  Fila {fila_idx} ({fecha_str})")

print(f"\nPrimeras 10 filas CON formulas:")
for fila_idx, formulas in filas_con_formulas[:10]:
    fecha_valor = ws.cell(row=fila_idx, column=4).value
    fecha_str = str(fecha_valor).split()[0] if fecha_valor else 'N/A'
    print(f"\nFila {fila_idx} ({fecha_str}):")
    for nombre_col, formula in formulas.items():
        print(f"  {nombre_col}: {formula[:60]}")

# Verificar patrón
if len(filas_con_formulas) > 0:
    print(f"\nPatrón detectado:")
    diferencias = []
    for i in range(1, min(20, len(filas_con_formulas))):
        diff = filas_con_formulas[i][0] - filas_con_formulas[i-1][0]
        diferencias.append(diff)
    print(f"Diferencias entre filas con formulas: {set(diferencias)}")
