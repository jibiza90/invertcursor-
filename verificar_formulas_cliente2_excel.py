#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificar las fórmulas del cliente 2 en Excel para entender qué columnas usa
"""

import openpyxl

wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)

ws_std = None
for sheet_name in wb.sheetnames:
    if 'STD' in sheet_name.upper() and 'VIP' not in sheet_name.upper():
        ws_std = wb[sheet_name]
        print(f"Hoja encontrada: {sheet_name}")
        break

print("="*80)
print("FÓRMULAS DEL CLIENTE 2 EN EXCEL - FILA 17")
print("="*80)

# Cliente 2 empieza en columna S (19)
# S=19, T=20, U=21, V=22, W=23, X=24, Y=25, Z=26

fila = 17

print(f"\nFila {fila}:")

# Ver todas las columnas del cliente 2
columnas_cliente2 = {
    'S': 19,  # incremento
    'T': 20,  # decremento
    'U': 21,  # base
    'V': 22,  # saldo_diario
    'W': 23,  # beneficio_diario
    'X': 24,  # beneficio_diario_pct
    'Y': 25,  # beneficio_acumulado
    'Z': 26   # beneficio_acumulado_pct
}

for letra, num_col in columnas_cliente2.items():
    celda = ws_std.cell(row=fila, column=num_col)
    if celda.data_type == 'f' and celda.value:
        print(f"  {letra}{fila} ({celda.value}): {celda.value}")
    else:
        valor = celda.value
        print(f"  {letra}{fila}: {valor}")

# Ver también las columnas generales que pueden ser referenciadas
print(f"\nColumnas generales (E-J) en fila {fila}:")
for letra, num_col in [('E', 5), ('F', 6), ('G', 7), ('H', 8), ('I', 9), ('J', 10)]:
    celda = ws_std.cell(row=fila, column=num_col)
    if celda.data_type == 'f' and celda.value:
        print(f"  {letra}{fila}: {celda.value}")
    else:
        valor = celda.value
        print(f"  {letra}{fila}: {valor}")

# Ver qué columnas referencia la fórmula de base del cliente 2
celda_base = ws_std.cell(row=fila, column=21)  # U17
if celda_base.data_type == 'f' and celda_base.value:
    formula_base = str(celda_base.value)
    print(f"\nFórmula de base (U17): {formula_base}")
    
    # Extraer referencias
    import re
    referencias = re.findall(r'([A-Z]+\d+)', formula_base)
    print(f"  Referencias encontradas: {referencias}")
    
    # Verificar qué es cada referencia
    for ref in referencias:
        col_letra = re.match(r'([A-Z]+)', ref).group(1)
        fila_num = int(re.match(r'\d+', ref).group())
        col_num = 0
        for i, c in enumerate(col_letra):
            col_num = col_num * 26 + (ord(c) - 64)
        
        celda_ref = ws_std.cell(row=fila_num, column=col_num)
        if celda_ref.data_type == 'f' and celda_ref.value:
            print(f"    {ref} (col {col_num}): fórmula = {celda_ref.value}")
        else:
            print(f"    {ref} (col {col_num}): valor = {celda_ref.value}")

wb.close()
