from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("ANALISIS DE PATRON EN COLUMNAS DE BENEFICIOS")
print("=" * 80)

# Verificar un rango más amplio para ver el patrón
print("\nAnalizando filas 15-50 para ver el patrón:")
print("-" * 80)

filas_con_formulas = []
filas_sin_formulas = []

for fila_idx in range(15, 51):
    fecha_celda = ws.cell(row=fila_idx, column=4)
    if fecha_celda.value is None:
        continue
    
    tiene_formula_benef_euro = False
    celda_benef_euro = ws.cell(row=fila_idx, column=7)
    
    if celda_benef_euro.data_type == 'f':
        tiene_formula_benef_euro = True
    elif celda_benef_euro.value is not None:
        valor_str = str(celda_benef_euro.value).strip()
        if valor_str.startswith('='):
            tiene_formula_benef_euro = True
    
    fecha_str = str(fecha_celda.value).split()[0] if fecha_celda.value else 'N/A'
    
    if tiene_formula_benef_euro:
        filas_con_formulas.append((fila_idx, fecha_str))
    else:
        filas_sin_formulas.append((fila_idx, fecha_str))

print(f"\nFilas CON formulas en benef_euro: {len(filas_con_formulas)}")
for fila_idx, fecha in filas_con_formulas[:10]:
    print(f"  Fila {fila_idx} ({fecha})")

print(f"\nFilas SIN formulas en benef_euro: {len(filas_sin_formulas)}")
for fila_idx, fecha in filas_sin_formulas[:10]:
    print(f"  Fila {fila_idx} ({fecha})")

# Verificar si hay un patrón (cada X filas)
if len(filas_con_formulas) > 0:
    diferencias = []
    for i in range(1, len(filas_con_formulas)):
        diff = filas_con_formulas[i][0] - filas_con_formulas[i-1][0]
        diferencias.append(diff)
    
    if diferencias:
        print(f"\nDiferencias entre filas con formulas: {set(diferencias)}")
        print("Patron detectado: Las formulas aparecen cada", set(diferencias), "filas")
