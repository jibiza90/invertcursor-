from openpyxl import load_workbook
import json

# Cargar Excel
print("Cargando Excel...")
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
ws_formulas = wb_formulas['Diario VIP']
ws_valores = wb_valores['Diario VIP']

# Cargar JSON
print("Cargando JSON...")
with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos_json = json.load(f)

hoja_json = datos_json['hojas']['Diario VIP']
clientes_json = hoja_json.get('clientes', [])

cliente1_json = next((c for c in clientes_json if c.get('numero_cliente') == 1), None)
datos_diarios_cliente = cliente1_json.get('datos_diarios', [])

# Mapeo
columna_a_campo_cliente = {
    11: 'incremento', 12: 'decremento', 13: 'base', 14: 'saldo_diario',
    15: 'beneficio_diario', 16: 'beneficio_diario_pct',
    17: 'beneficio_acumulado', 18: 'beneficio_acumulado_pct'
}

errores = []
advertencias = []

# Verificar primeras filas
with open('reporte_verificacion.txt', 'w', encoding='utf-8') as f:
    f.write("=" * 80 + "\n")
    f.write("VERIFICACION CELDA A CELDA: EXCEL vs JSON\n")
    f.write("=" * 80 + "\n\n")
    
    for fila_excel in range(15, 35):
        fecha_excel = ws_valores.cell(row=fila_excel, column=4).value
        if fecha_excel is None:
            continue
        
        fila_data_json = next((d for d in datos_diarios_cliente if d['fila'] == fila_excel), None)
        if not fila_data_json:
            continue
        
        f.write(f"\nFILA {fila_excel} (Fecha: {fecha_excel})\n")
        f.write("-" * 80 + "\n")
        
        for col_excel, campo in columna_a_campo_cliente.items():
            celda_excel = ws_formulas.cell(row=fila_excel, column=col_excel)
            tiene_formula_excel = celda_excel.data_type == 'f' and celda_excel.value
            
            tiene_formula_json = fila_data_json.get('formulas', {}).get(campo, None)
            esta_bloqueada_json = fila_data_json.get('bloqueadas', {}).get(campo, False)
            valor_excel = ws_valores.cell(row=fila_excel, column=col_excel).value
            
            col_letra = chr(64 + col_excel) if col_excel <= 26 else chr(64 + (col_excel-1)//26) + chr(65 + (col_excel-1)%26)
            
            if tiene_formula_excel:
                formula_excel = str(celda_excel.value)
                f.write(f"  {col_letra}{fila_excel} ({campo}): TIENE FORMULA\n")
                f.write(f"    Excel: {formula_excel}\n")
                
                if tiene_formula_json:
                    formula_json = tiene_formula_json
                    if formula_excel != formula_json:
                        errores.append(f"Fila {fila_excel}, {campo}: Fórmula diferente")
                        f.write(f"    ❌ DIFERENTE - JSON: {formula_json}\n")
                    else:
                        f.write(f"    ✅ Coincide\n")
                else:
                    errores.append(f"Fila {fila_excel}, {campo}: Fórmula en Excel pero NO en JSON")
                    f.write(f"    ❌ FALTA EN JSON\n")
                
                if not esta_bloqueada_json:
                    errores.append(f"Fila {fila_excel}, {campo}: Tiene fórmula pero NO bloqueada")
                    f.write(f"    ❌ NO BLOQUEADA\n")
                else:
                    f.write(f"    ✅ Bloqueada\n")
            else:
                if campo in ['incremento', 'decremento']:
                    if esta_bloqueada_json:
                        errores.append(f"Fila {fila_excel}, {campo}: NO debe estar bloqueada")
                        f.write(f"  {col_letra}{fila_excel} ({campo}): ❌ BLOQUEADA (no debería)\n")
                    else:
                        f.write(f"  {col_letra}{fila_excel} ({campo}): ✅ Editable (valor: {valor_excel})\n")
                else:
                    f.write(f"  {col_letra}{fila_excel} ({campo}): Valor={valor_excel}, Bloqueada={esta_bloqueada_json}\n")
    
    f.write("\n" + "=" * 80 + "\n")
    f.write("RESUMEN\n")
    f.write("=" * 80 + "\n")
    f.write(f"Errores: {len(errores)}\n")
    f.write(f"Advertencias: {len(advertencias)}\n\n")
    
    if errores:
        f.write("ERRORES:\n")
        for error in errores:
            f.write(f"  - {error}\n")

print(f"Reporte guardado en: reporte_verificacion.txt")
print(f"Errores encontrados: {len(errores)}")
