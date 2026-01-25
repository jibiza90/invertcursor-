from openpyxl import load_workbook
from datetime import datetime

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb_formulas.sheetnames:
    ws_formulas = wb_formulas['Diario VIP']
    ws_valores = wb_valores['Diario VIP']
    
    print("=" * 80)
    print("VERIFICANDO FORMULAS EN DATOS GENERALES")
    print("=" * 80)
    
    # Verificar filas con fechas alrededor del 13/01/2026
    print("\nVerificando filas alrededor del 13/01/2026:")
    
    for fila_idx in range(15, 30):
        fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
        
        if fecha_valor:
            try:
                if isinstance(fecha_valor, datetime):
                    fecha_str = fecha_valor.strftime('%d/%m/%Y')
                else:
                    fecha_str = str(fecha_valor)
                
                # Verificar fórmulas en columnas E-J
                tiene_formulas = {}
                for col_idx in range(5, 11):
                    celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                    if celda_formula.value and str(celda_formula.value).startswith('='):
                        tiene_formulas[col_idx] = True
                
                if tiene_formulas:
                    print(f"\nFila {fila_idx}, Fecha: {fecha_str}")
                    print(f"  Columnas con fórmulas: {list(tiene_formulas.keys())}")
                    
            except:
                pass
