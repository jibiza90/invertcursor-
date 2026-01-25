from openpyxl import load_workbook
import json

# Analizar la estructura para entender el problema del desfase
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario STD']

print("Analizando estructura de fórmulas para entender el problema:")
print("=" * 80)

# Verificar primeras fechas
for fila in [15, 16, 17, 18, 19, 20, 21]:
    fecha = ws.cell(row=fila, column=4).value
    formula_imp_inicial = ws.cell(row=fila, column=5).value  # Columna E
    
    print(f"\nFila {fila} (Fecha: {fecha}):")
    if formula_imp_inicial and str(formula_imp_inicial).startswith('='):
        print(f"  Formula imp_inicial: {formula_imp_inicial}")
        
        # Extraer referencias
        import re
        refs = re.findall(r'([A-Z]+\d+)', str(formula_imp_inicial))
        print(f"  Referencias: {refs}")
        
        # Verificar qué es AEO
        if 'AEO' in str(formula_imp_inicial):
            # Buscar qué fila es AEO
            aeo_refs = re.findall(r'AEO(\d+)', str(formula_imp_inicial))
            for aeo_fila_str in aeo_refs:
                aeo_fila = int(aeo_fila_str)
                fecha_aeo = ws.cell(row=aeo_fila, column=4).value
                print(f"    AEO{aeo_fila} está en fecha: {fecha_aeo}")
                print(f"    Diferencia de filas: {aeo_fila - fila}")
                
                # Verificar incrementos del cliente 1 en esas filas
                # Cliente 1 está en columna K (11)
                inc_fila_cliente = ws.cell(row=fila, column=11).value
                inc_aeo_cliente = ws.cell(row=aeo_fila, column=11).value
                print(f"    Incremento cliente fila {fila}: {inc_fila_cliente}")
                print(f"    Incremento cliente fila {aeo_fila}: {inc_aeo_cliente}")
    else:
        print(f"  Sin fórmula (valor: {ws.cell(row=fila, column=5).value})")

print("\n" + "=" * 80)
print("CONCLUSIÓN:")
print("=" * 80)
print("Si el incremento está en la fila 15 del cliente (día 1):")
print("  - La fórmula del día 1 es imp_inicial = AEO16")
print("  - AEO16 debería sumar incrementos de la fila 16 del cliente")
print("  - PERO el incremento está en la fila 15, no en la 16")
print("  - Por lo tanto, AEO16 no ve el incremento de la fila 15")
