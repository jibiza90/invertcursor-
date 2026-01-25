from openpyxl import load_workbook
from datetime import datetime

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb_formulas.sheetnames:
    ws_formulas = wb_formulas['Diario VIP']
    ws_valores = wb_valores['Diario VIP']
    
    print("=" * 80)
    print("VERIFICANDO FORMULAS EN DATOS GENERALES DIARIOS")
    print("=" * 80)
    
    # Verificar algunas fechas específicas
    fechas_verificar = [
        ('2026-01-01', 15),
        ('2026-01-13', 27),
        ('2026-01-14', 28),
        ('2026-01-15', 29),
        ('2026-06-15', 180),
        ('2026-12-31', 365)
    ]
    
    print("\nVerificando fórmulas en columnas E-J para diferentes fechas:")
    
    for fecha_str, fila_aproximada in fechas_verificar:
        # Buscar la fila con esa fecha
        for fila_idx in range(15, min(400, ws_valores.max_row + 1)):
            fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
            if fecha_valor:
                try:
                    if isinstance(fecha_valor, datetime):
                        fecha_dt_str = fecha_valor.strftime('%Y-%m-%d')
                    else:
                        fecha_dt_str = str(fecha_valor).split()[0]
                    
                    if fecha_dt_str == fecha_str:
                        print(f"\nFila {fila_idx}, Fecha: {fecha_str}")
                        columnas_info = {
                            5: 'E - Imp. Inicial',
                            6: 'F - Imp. Final',
                            7: 'G - Benef. €',
                            8: 'H - Benef. %',
                            9: 'I - Benef. € Acum',
                            10: 'J - Benef. % Acum'
                        }
                        
                        for col_idx, nombre in columnas_info.items():
                            tiene_formula = False
                            celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                            if celda_formula.value and str(celda_formula.value).startswith('='):
                                tiene_formula = True
                            
                            valor = ws_valores.cell(row=fila_idx, column=col_idx).value
                            estado = "BLOQUEADA (fórmula)" if tiene_formula else "EDITABLE"
                            print(f"  {nombre}: {valor} - {estado}")
                        
                        break
                except:
                    pass
