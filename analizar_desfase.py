from openpyxl import load_workbook
import json

# Analizar la estructura de las fórmulas para entender el problema del desfase
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("Analizando estructura de fórmulas imp_inicial para entender el desfase:")
print("=" * 80)

# Verificar primeras fechas
for fila in [15, 16, 17, 18, 19, 20, 21]:
    fecha = ws.cell(row=fila, column=4).value
    formula_imp_inicial = ws.cell(row=fila, column=5).value  # Columna E
    
    print(f"\nFila {fila} (Fecha: {fecha}):")
    if formula_imp_inicial and str(formula_imp_inicial).startswith('='):
        print(f"  Formula imp_inicial: {formula_imp_inicial}")
        
        # Extraer referencias
        import re
        refs = re.findall(r'([A-Z]+\d+)', str(formula_imp_inicial))
        print(f"  Referencias: {refs}")
        
        # Verificar qué es AEO
        if 'AEO' in str(formula_imp_inicial):
            # Buscar qué fila es AEO
            aeo_refs = re.findall(r'AEO(\d+)', str(formula_imp_inicial))
            for aeo_fila_str in aeo_refs:
                aeo_fila = int(aeo_fila_str)
                fecha_aeo = ws.cell(row=aeo_fila, column=4).value
                print(f"    AEO{aeo_fila} está en fecha: {fecha_aeo}")
                print(f"    Diferencia de filas: {aeo_fila - fila}")
                
                # Verificar si AEO corresponde al mismo día o día siguiente
                if fecha == fecha_aeo:
                    print(f"    ✅ AEO es del MISMO DÍA")
                else:
                    print(f"    ❌ AEO es de OTRO DÍA")
    else:
        print(f"  Sin fórmula (valor: {ws.cell(row=fila, column=5).value})")

print("\n" + "=" * 80)
print("CONCLUSIÓN:")
print("=" * 80)
print("Si la fórmula del día 2 es F17+AEO19:")
print("  - F17 es imp_final del día 1")
print("  - AEO19 es FA del día 2")
print("  - Entonces cuando hay un incremento en el día 1, debería afectar AEO del día 1")
print("  - Pero la fórmula del día 2 usa AEO del día 2")
print("  - Por lo tanto, el incremento del día 1 debería aparecer en el imp_inicial del día 1, no del día 2")
