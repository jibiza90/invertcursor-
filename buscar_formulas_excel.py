#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para buscar fórmulas en incremento/decremento en el Excel
"""

import openpyxl

def obtener_nombre_columna(num_col):
    """Convertir número de columna a letra"""
    result = ""
    num = num_col
    while num > 0:
        num -= 1
        result = chr(65 + (num % 26)) + result
        num //= 26
    return result

def main():
    print("="*80)
    print("BUSCANDO FÓRMULAS EN INCREMENTO/DECREMENTO")
    print("="*80)
    
    # Cargar Excel
    wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
    
    # Buscar hoja "Diario VIP"
    ws = None
    for sheet_name in wb.sheetnames:
        if 'VIP' in sheet_name.upper():
            ws = wb[sheet_name]
            print(f"\n✅ Hoja encontrada: {sheet_name}")
            break
    
    if not ws:
        ws = wb.active
        print(f"\n⚠️ Usando hoja activa: {ws.title}")
    
    # Buscar fórmulas en columnas K y L (incremento y decremento) para cliente 1
    # Cliente 1 empieza en columna K (11)
    col_inicio_cliente1 = 11  # K
    col_incremento = col_inicio_cliente1  # K
    col_decremento = col_inicio_cliente1 + 1  # L
    
    print(f"\nBuscando fórmulas en columnas {obtener_nombre_columna(col_incremento)} (incremento) y {obtener_nombre_columna(col_decremento)} (decremento)")
    print(f"Rango de filas: 15-1120\n")
    
    formulas_encontradas = []
    
    for fila in range(15, min(200, 1121)):  # Revisar hasta fila 200 para empezar
        # Verificar incremento (columna K)
        celda_inc = ws.cell(row=fila, column=col_incremento)
        if celda_inc.data_type == 'f' and celda_inc.value:
            formulas_encontradas.append({
                'fila': fila,
                'columna': obtener_nombre_columna(col_incremento),
                'campo': 'incremento',
                'formula': str(celda_inc.value)
            })
        
        # Verificar decremento (columna L)
        celda_dec = ws.cell(row=fila, column=col_decremento)
        if celda_dec.data_type == 'f' and celda_dec.value:
            formulas_encontradas.append({
                'fila': fila,
                'columna': obtener_nombre_columna(col_decremento),
                'campo': 'decremento',
                'formula': str(celda_dec.value)
            })
    
    if formulas_encontradas:
        print(f"✅ Encontradas {len(formulas_encontradas)} fórmulas:")
        for f in formulas_encontradas[:30]:
            print(f"   Fila {f['fila']}, Col {f['columna']} ({f['campo']}): {f['formula']}")
    else:
        print("⚠️ No se encontraron fórmulas en las primeras 200 filas")
        print("   Esto puede significar que no hay fórmulas o que están en otro rango")
    
    wb.close()

if __name__ == '__main__':
    main()
