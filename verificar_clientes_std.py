from openpyxl import load_workbook

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario STD' in wb.sheetnames:
    ws = wb['Diario STD']
    
    print("=" * 80)
    print("VERIFICANDO CLIENTES EN DIARIO STD")
    print("=" * 80)
    
    # Fila 13 tiene los números de clientes
    fila_nombres = 13
    clientes_encontrados = []
    
    for col_idx in range(1, min(200, ws.max_column + 1)):
        celda = ws.cell(row=fila_nombres, column=col_idx)
        valor = celda.value
        
        if valor is not None:
            try:
                numero_cliente = int(valor)
                if numero_cliente > 0:
                    clientes_encontrados.append({
                        'numero': numero_cliente,
                        'columna': col_idx
                    })
            except:
                pass
    
    print(f"\nTotal clientes encontrados: {len(clientes_encontrados)}")
    print("\nLista de clientes:")
    for cliente in clientes_encontrados:
        print(f"  Cliente {cliente['numero']}: Columna {cliente['columna']}")
