from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("VERIFICACION FORMULAS PORCENTAJES - FILAS CON FORMULAS")
print("=" * 80)

# Buscar filas con fórmulas de porcentajes (H y J)
filas_con_formulas = []
for fila_idx in range(15, min(200, ws.max_row + 1)):
    fecha_celda = ws.cell(row=fila_idx, column=4).value
    if fecha_celda is None:
        continue
    
    tiene_formula_h = ws.cell(row=fila_idx, column=8).data_type == 'f'  # H
    tiene_formula_j = ws.cell(row=fila_idx, column=10).data_type == 'f'  # J
    
    if tiene_formula_h or tiene_formula_j:
        filas_con_formulas.append(fila_idx)

print(f"\nFilas con fórmulas de porcentajes (primeras 10): {filas_con_formulas[:10]}")

# Verificar fila 56 específicamente
fila = 56
print(f"\nFila {fila} - Fórmulas:")
celda_h = ws.cell(row=fila, column=8)
celda_j = ws.cell(row=fila, column=10)

print(f"  H (Benef. %): tiene_formula={celda_h.data_type == 'f'}")
if celda_h.data_type == 'f' and celda_h.value:
    print(f"    Formula: {str(celda_h.value)[:100]}")

print(f"  J (Benef. % Acum): tiene_formula={celda_j.data_type == 'f'}")
if celda_j.data_type == 'f' and celda_j.value:
    print(f"    Formula: {str(celda_j.value)[:100]}")

# Verificar también fila 54
fila = 54
print(f"\nFila {fila} - Fórmulas:")
celda_h = ws.cell(row=fila, column=8)
celda_j = ws.cell(row=fila, column=10)

print(f"  H (Benef. %): tiene_formula={celda_h.data_type == 'f'}")
if celda_h.data_type == 'f' and celda_h.value:
    print(f"    Formula: {str(celda_h.value)[:100]}")
else:
    print(f"    Valor: {celda_h.value}")

print(f"  J (Benef. % Acum): tiene_formula={celda_j.data_type == 'f'}")
if celda_j.data_type == 'f' and celda_j.value:
    print(f"    Formula: {str(celda_j.value)[:100]}")
else:
    print(f"    Valor: {celda_j.value}")
