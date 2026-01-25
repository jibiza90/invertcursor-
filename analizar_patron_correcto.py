#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para analizar el patrón CORRECTO agrupando por fecha real
"""

import openpyxl
import sys
from datetime import datetime
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

def tiene_formula(ws, fila, columna):
    try:
        celda = ws.cell(row=fila, column=columna)
        return celda.data_type == 'f' and celda.value is not None
    except:
        return False

def parsear_fecha_excel(valor):
    """Convertir fecha de Excel a string normalizado"""
    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d')
    elif isinstance(valor, str):
        if valor.startswith('='):
            return None  # Es una fórmula, no una fecha real
        try:
            dt = datetime.strptime(valor.split()[0], '%Y-%m-%d')
            return dt.strftime('%Y-%m-%d')
        except:
            return None
    return None

wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)  # Usar data_only=True para obtener valores calculados
ws_formulas = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)  # Para fórmulas

ws = None
ws_f = None
for sheet_name in wb.sheetnames:
    if 'VIP' in sheet_name.upper():
        ws = wb[sheet_name]
        ws_f = ws_formulas[sheet_name]
        break

if not ws:
    ws = wb.active
    ws_f = ws_formulas.active

print("="*80)
print("ANÁLISIS DEL PATRÓN CORRECTO POR DÍA")
print("="*80)

# Agrupar por fecha real (usando valores calculados)
fechas_agrupadas = defaultdict(list)

for fila in range(15, 51):
    fecha_valor = ws.cell(row=fila, column=4).value  # Usar valores calculados
    fecha_str = parsear_fecha_excel(fecha_valor)
    if fecha_str:
        fechas_agrupadas[fecha_str].append(fila)

print("\nPATRÓN POR DÍA (primeros 3 días):")
print("="*80)

for fecha_str, filas in sorted(fechas_agrupadas.items())[:3]:
    filas_ordenadas = sorted(filas)
    print(f"\n{'='*80}")
    print(f"Fecha: {fecha_str}")
    print(f"Filas: {filas_ordenadas} (total: {len(filas_ordenadas)} filas)")
    print(f"{'='*80}")
    
    for idx, fila in enumerate(filas_ordenadas, 1):
        print(f"\n  Fila {fila} (posición {idx} del día):")
        
        # IMP_FINAL (columna F = 6)
        tiene_formula_imp_final = tiene_formula(ws_f, fila, 6)
        estado_imp_final = "🔒 BLOQUEADA" if tiene_formula_imp_final else "✏️ EDITABLE"
        print(f"    IMP_FINAL (F): {estado_imp_final}")
        
        # BENEF_EURO (columna G = 7)
        tiene_formula_benef_euro = tiene_formula(ws_f, fila, 7)
        estado_benef_euro = "🔒 BLOQUEADA" if tiene_formula_benef_euro else "✏️ EDITABLE"
        print(f"    BENEF_EURO (G): {estado_benef_euro}")
        
        # BENEF_PORCENTAJE (columna H = 8)
        tiene_formula_benef_pct = tiene_formula(ws_f, fila, 8)
        estado_benef_pct = "🔒 BLOQUEADA" if tiene_formula_benef_pct else "✏️ EDITABLE"
        print(f"    BENEF_PORCENTAJE (H): {estado_benef_pct}")
        
        # BENEF_EURO_ACUM (columna I = 9)
        tiene_formula_benef_acum = tiene_formula(ws_f, fila, 9)
        estado_benef_acum = "🔒 BLOQUEADA" if tiene_formula_benef_acum else "✏️ EDITABLE"
        print(f"    BENEF_EURO_ACUM (I): {estado_benef_acum}")
        
        # BENEF_PORCENTAJE_ACUM (columna J = 10)
        tiene_formula_benef_pct_acum = tiene_formula(ws_f, fila, 10)
        estado_benef_pct_acum = "🔒 BLOQUEADA" if tiene_formula_benef_pct_acum else "✏️ EDITABLE"
        print(f"    BENEF_PORCENTAJE_ACUM (J): {estado_benef_pct_acum}")

wb.close()
ws_formulas.close()
