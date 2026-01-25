from openpyxl import load_workbook
import json

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

# Buscar Diario VIP
sheet_name = None
for sheet in wb.sheetnames:
    if 'Diario VIP' in sheet or 'Diario_VIP' in sheet or 'diario_vip' in sheet.lower():
        sheet_name = sheet
        break

if sheet_name:
    ws = wb[sheet_name]
    
    # Leer todas las filas con datos
    todas_las_filas = []
    for row in ws.iter_rows(values_only=True):
        fila = [cell for cell in row]
        # Solo agregar filas que tengan al menos un valor no vacío
        if any(cell is not None and str(cell).strip() != '' for cell in fila):
            todas_las_filas.append(fila)
    
    # Buscar headers reales (buscar fila con texto que parezca headers)
    headers = None
    header_row_idx = None
    for idx, fila in enumerate(todas_las_filas[:50]):  # Buscar en las primeras 50 filas
        # Buscar fila que tenga varias columnas con texto (posibles headers)
        textos = [str(cell).strip() for cell in fila if cell is not None and str(cell).strip() != '']
        if len(textos) > 5:  # Si tiene más de 5 columnas con texto, probablemente son headers
            headers = [str(cell).strip() if cell is not None else '' for cell in fila]
            header_row_idx = idx
            break
    
    # Si no encontramos headers claros, usar la primera fila
    if headers is None:
        headers = [str(cell).strip() if cell is not None else '' for cell in todas_las_filas[0]]
        header_row_idx = 0
    
    # Leer datos después de los headers
    datos = []
    for fila in todas_las_filas[header_row_idx + 1:]:
        fila_dict = {}
        for i, valor in enumerate(fila):
            if i < len(headers):
                clave = headers[i] if headers[i] else f'Columna_{i+1}'
                fila_dict[clave] = valor
        # Solo agregar si tiene datos significativos
        valores_significativos = [v for v in fila_dict.values() if v is not None and str(v).strip() != '']
        if len(valores_significativos) > 0:
            datos.append(fila_dict)
    
    # Buscar cliente 1 de diferentes formas
    cliente_1 = None
    
    # Método 1: Buscar columna con "cliente" o "client" y valor 1
    for fila in datos:
        for key, value in fila.items():
            if key and ('cliente' in str(key).lower() or 'client' in str(key).lower()):
                try:
                    if int(value) == 1:
                        cliente_1 = fila
                        break
                except:
                    pass
        if cliente_1:
            break
    
    # Método 2: Buscar cualquier columna numérica con valor 1 en posición de ID
    if not cliente_1:
        for fila in datos:
            valores_numericos = []
            for key, value in fila.items():
                try:
                    if value is not None:
                        num = float(value)
                        if num == 1.0:
                            valores_numericos.append((key, value))
                except:
                    pass
            # Si encontramos un 1 y la fila tiene varios datos, podría ser cliente 1
            if valores_numericos and len([v for v in fila.values() if v is not None]) > 3:
                cliente_1 = fila
                break
    
    # Método 3: Buscar fila que tenga texto que indique "cliente 1" o similar
    if not cliente_1:
        for fila in datos:
            texto_completo = ' '.join([str(v) for v in fila.values() if v is not None]).lower()
            if 'cliente' in texto_completo and ('1' in texto_completo or 'uno' in texto_completo):
                cliente_1 = fila
                break
    
    # Si aún no encontramos, mostrar las primeras filas con datos significativos
    if not cliente_1:
        # Buscar la primera fila que tenga varios datos (probablemente datos reales)
        for fila in datos:
            valores_no_vacios = [v for v in fila.values() if v is not None and str(v).strip() != '']
            if len(valores_no_vacios) > 5:  # Si tiene más de 5 valores, probablemente es una fila de datos
                cliente_1 = fila
                break
    
    # Guardar resultado
    resultado = {
        "hoja": sheet_name,
        "total_filas_datos": len(datos),
        "headers_encontrados": [h for h in headers if h],
        "cliente_1": cliente_1,
        "primeras_10_filas": datos[:10] if len(datos) > 0 else []
    }
    
    # Escribir archivo de texto legible
    with open('CLIENTE1_COMPLETO.txt', 'w', encoding='utf-8') as f:
        f.write('='*80 + '\n')
        f.write(f'INFORMACIÓN DEL CLIENTE 1 - HOJA: {sheet_name}\n')
        f.write('='*80 + '\n\n')
        
        f.write(f'Total de filas con datos: {len(datos)}\n')
        f.write(f'Total de columnas: {len(headers)}\n\n')
        
        f.write('Headers encontrados (primeras 30 columnas con nombre):\n')
        headers_con_nombre = [(i, h) for i, h in enumerate(headers) if h]
        for i, (idx, header) in enumerate(headers_con_nombre[:30], 1):
            f.write(f'  {i}. Columna {idx+1}: {header}\n')
        f.write('\n')
        
        f.write('='*80 + '\n')
        f.write('DATOS DEL CLIENTE 1:\n')
        f.write('='*80 + '\n\n')
        if cliente_1:
            for key, value in cliente_1.items():
                if value is not None:
                    f.write(f'{key}: {value}\n')
        else:
            f.write('No se encontró información específica del cliente 1\n')
        
        f.write('\n' + '='*80 + '\n')
        f.write('PRIMERAS 10 FILAS CON DATOS:\n')
        f.write('='*80 + '\n\n')
        for i, fila in enumerate(datos[:10], 1):
            f.write(f'\nFila {i}:\n')
            valores = {k: v for k, v in fila.items() if v is not None and str(v).strip() != ''}
            for key, value in valores.items():
                f.write(f'  {key}: {value}\n')
    
    # También guardar JSON
    with open('CLIENTE1_COMPLETO.json', 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2, default=str)
    
    print("Archivos creados: CLIENTE1_COMPLETO.txt y CLIENTE1_COMPLETO.json")
    print(f"Cliente 1 encontrado: {cliente_1 is not None}")
