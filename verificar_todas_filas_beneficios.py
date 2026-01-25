from openpyxl import load_workbook
from datetime import datetime

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("VERIFICACION COMPLETA - COLUMNAS DE BENEFICIOS")
print("=" * 80)

columnas_beneficios = {
    7: 'benef_euro',
    8: 'benef_porcentaje',
    9: 'benef_euro_acum',
    10: 'benef_porcentaje_acum'
}

# Contar cuántas filas tienen fórmulas en cada columna de beneficios
contadores = {col_key: 0 for col_key in columnas_beneficios.values()}

# Verificar desde fila 15 hasta el final
for fila_idx in range(15, min(400, ws.max_row + 1)):
    fecha_celda = ws.cell(row=fila_idx, column=4)
    if fecha_celda.value is None:
        continue
    
    for col_idx, col_key in columnas_beneficios.items():
        celda = ws.cell(row=fila_idx, column=col_idx)
        
        tiene_formula = False
        if celda.data_type == 'f':
            tiene_formula = True
        elif celda.value is not None:
            valor_str = str(celda.value).strip()
            if valor_str.startswith('='):
                tiene_formula = True
        
        if tiene_formula:
            contadores[col_key] += 1

print("\nResumen de filas con fórmulas en columnas de beneficios:")
for col_key, count in contadores.items():
    print(f"  {col_key}: {count} filas con fórmula")

# Verificar filas específicas que sabemos que tienen fórmulas
print("\n" + "=" * 80)
print("Ejemplos de filas con fórmulas:")
print("=" * 80)

filas_con_formulas = []
for fila_idx in range(15, min(100, ws.max_row + 1)):
    fecha_celda = ws.cell(row=fila_idx, column=4)
    if fecha_celda.value is None:
        continue
    
    tiene_alguna_formula = False
    formulas_en_fila = {}
    
    for col_idx, col_key in columnas_beneficios.items():
        celda = ws.cell(row=fila_idx, column=col_idx)
        
        tiene_formula = False
        formula_texto = None
        
        if celda.data_type == 'f':
            tiene_formula = True
            formula_texto = str(celda.value)[:60]
            tiene_alguna_formula = True
        elif celda.value is not None:
            valor_str = str(celda.value).strip()
            if valor_str.startswith('='):
                tiene_formula = True
                formula_texto = valor_str[:60]
                tiene_alguna_formula = True
        
        if tiene_formula:
            formulas_en_fila[col_key] = formula_texto
    
    if tiene_alguna_formula and len(filas_con_formulas) < 10:
        fecha_str = str(fecha_celda.value).split()[0] if fecha_celda.value else 'N/A'
        filas_con_formulas.append((fila_idx, fecha_str, formulas_en_fila))

for fila_idx, fecha_str, formulas in filas_con_formulas:
    print(f"\nFila {fila_idx} (Fecha: {fecha_str}):")
    for col_key, formula in formulas.items():
        print(f"  {col_key}: {formula}")
