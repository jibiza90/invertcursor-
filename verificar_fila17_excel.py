from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

fila = 17
print(f"Fila {fila} - Verificando fórmulas:")
print("=" * 80)

columnas = {
    5: 'E - Imp. Inicial',
    6: 'F - Imp. Final',
    7: 'G - Benef. €',
    8: 'H - Benef. %',
    9: 'I - Benef. € Acum',
    10: 'J - Benef. % Acum'
}

for col_idx, nombre in columnas.items():
    celda = ws.cell(row=fila, column=col_idx)
    valor = celda.value
    tiene_formula = False
    formula_texto = None
    
    if celda.value is not None:
        valor_str = str(celda.value).strip()
        if valor_str.startswith('='):
            tiene_formula = True
            formula_texto = valor_str[:60]  # Primeros 60 caracteres
    
    estado = "[BLOQUEADA]" if tiene_formula else "[EDITABLE]"
    print(f"{nombre}: valor={valor}, tiene_formula={tiene_formula} {estado}")
    if formula_texto:
        print(f"    Formula: {formula_texto}")
