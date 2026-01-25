#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificar fórmulas y valores en Diario STD para entender la lógica correcta
"""

import json
import openpyxl

# Cargar Excel
wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws_std = None
for sheet_name in wb.sheetnames:
    if 'STD' in sheet_name.upper() and 'VIP' not in sheet_name.upper():
        ws_std = wb[sheet_name]
        print(f"Hoja encontrada: {sheet_name}")
        break

if not ws_std:
    ws_std = wb.active

# Cargar JSON
with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos = json.load(f)

hoja_std = datos.get('hojas', {}).get('Diario STD')

print("="*80)
print("ANÁLISIS DE FÓRMULAS EN DIARIO STD (Filas 15-20)")
print("="*80)

# Analizar filas 15-20
for fila in range(15, 21):
    print(f"\n{'='*80}")
    print(f"FILA {fila}:")
    print(f"{'='*80}")
    
    # Fecha
    fecha_excel = ws_std.cell(row=fila, column=4).value
    print(f"Fecha Excel: {fecha_excel}")
    
    # Valores en Excel (data_only=True)
    wb_valores = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
    ws_valores_std = None
    for sheet_name in wb_valores.sheetnames:
        if 'STD' in sheet_name.upper() and 'VIP' not in sheet_name.upper():
            ws_valores_std = wb_valores[sheet_name]
            break
    
    if ws_valores_std:
        imp_inicial_excel = ws_valores_std.cell(row=fila, column=5).value
        imp_final_excel = ws_valores_std.cell(row=fila, column=6).value
        benef_euro_excel = ws_valores_std.cell(row=fila, column=7).value
        benef_porcentaje_excel = ws_valores_std.cell(row=fila, column=8).value
        
        print(f"Excel valores:")
        print(f"  E{fila} (imp_inicial): {imp_inicial_excel}")
        print(f"  F{fila} (imp_final): {imp_final_excel}")
        print(f"  G{fila} (benef_euro): {benef_euro_excel}")
        print(f"  H{fila} (benef_porcentaje): {benef_porcentaje_excel}")
    
    # Fórmulas en Excel
    print(f"\nExcel fórmulas:")
    for col_idx, col_name in [(5, 'E (imp_inicial)'), (6, 'F (imp_final)'), (7, 'G (benef_euro)'), (8, 'H (benef_porcentaje)')]:
        celda = ws_std.cell(row=fila, column=col_idx)
        if celda.data_type == 'f' and celda.value:
            print(f"  {col_name}: {celda.value}")
    
    # Valores en JSON
    if hoja_std:
        datos_diarios = hoja_std.get('datos_diarios_generales', [])
        dato_json = [d for d in datos_diarios if d.get('fila') == fila]
        if dato_json:
            d = dato_json[0]
            print(f"\nJSON valores:")
            print(f"  imp_inicial: {d.get('imp_inicial')}")
            print(f"  imp_final: {d.get('imp_final')}")
            print(f"  benef_euro: {d.get('benef_euro')}")
            print(f"  benef_porcentaje: {d.get('benef_porcentaje')}")
            
            formulas = d.get('formulas', {})
            print(f"\nJSON fórmulas:")
            for campo in ['imp_inicial', 'imp_final', 'benef_euro', 'benef_porcentaje']:
                if campo in formulas:
                    print(f"  {campo}: {formulas[campo]}")

wb.close()
if 'wb_valores' in locals():
    wb_valores.close()
