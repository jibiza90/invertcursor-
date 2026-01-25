from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)

if 'Diario VIP' in wb_formulas.sheetnames:
    ws = wb_formulas['Diario VIP']
    
    print("=" * 80)
    print("VERIFICANDO DETECCION DE FORMULAS")
    print("=" * 80)
    
    def tiene_formula(ws_formulas, fila, columna):
        """Verifica si una celda tiene fórmula"""
        celda = ws_formulas.cell(row=fila, column=columna)
        if celda.value is not None:
            valor_str = str(celda.value)
            return valor_str.startswith('=')
        return False
    
    # Verificar filas específicas
    filas_test = [15, 17, 27, 51, 54]
    
    for fila_idx in filas_test:
        print(f"\nFila {fila_idx}:")
        for col_idx in range(5, 11):
            letra = chr(64 + col_idx)
            celda = ws.cell(row=fila_idx, column=col_idx)
            valor = celda.value
            
            tiene_form = False
            if valor is not None:
                valor_str = str(valor)
                tiene_form = valor_str.startswith('=')
            
            resultado_funcion = tiene_formula(ws, fila_idx, col_idx)
            
            estado = "OK" if tiene_form == resultado_funcion else "ERROR"
            print(f"  Col {letra}: valor={valor}, tiene_formula={tiene_form}, funcion={resultado_funcion} {estado}")
