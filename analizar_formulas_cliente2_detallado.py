#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificar las fórmulas del cliente 2 en Excel para entender qué debería calcularse
"""

import json
import openpyxl

# Cargar Excel
wb_formulas = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

ws_std_formulas = None
ws_std_valores = None
for sheet_name in wb_formulas.sheetnames:
    if 'STD' in sheet_name.upper() and 'VIP' not in sheet_name.upper():
        ws_std_formulas = wb_formulas[sheet_name]
        ws_std_valores = wb_valores[sheet_name]
        print(f"Hoja encontrada: {sheet_name}")
        break

# Cargar JSON
with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos = json.load(f)

hoja_std = datos.get('hojas', {}).get('Diario STD')

print("="*80)
print("ANÁLISIS DE FÓRMULAS DEL CLIENTE 2 - FILA 17")
print("="*80)

# Cliente 2 empieza en columna S (19)
# S=19 (incremento), T=20 (decremento), U=21 (base), V=22 (saldo_diario), W=23 (beneficio_diario), X=24 (beneficio_diario_pct), Y=25 (beneficio_acumulado)

fila = 17

print(f"\nFila {fila} (día 1, tercera fila):")

# Ver fórmulas en Excel
print("\nFórmulas en Excel:")
for col, nombre in [(21, 'U (base)'), (22, 'V (saldo_diario)'), (23, 'W (beneficio_diario)'), (24, 'X (beneficio_diario_pct)'), (25, 'Y (beneficio_acumulado)')]:
    celda = ws_std_formulas.cell(row=fila, column=col)
    if celda.data_type == 'f' and celda.value:
        print(f"  {nombre}: {celda.value}")

# Ver valores en Excel
print("\nValores en Excel:")
for col, nombre in [(19, 'S (incremento)'), (20, 'T (decremento)'), (21, 'U (base)'), (22, 'V (saldo_diario)'), (23, 'W (beneficio_diario)'), (24, 'X (beneficio_diario_pct)'), (25, 'Y (beneficio_acumulado)')]:
    celda = ws_std_valores.cell(row=fila, column=col)
    valor = celda.value
    print(f"  {nombre}: {valor}")

# Ver valores necesarios para las fórmulas (columnas generales)
print("\nValores de columnas generales necesarios:")
for col, nombre in [(5, 'E (imp_inicial)'), (6, 'F (imp_final)'), (7, 'G (benef_euro)'), (8, 'H (benef_porcentaje)')]:
    celda = ws_std_valores.cell(row=fila, column=col)
    valor = celda.value
    print(f"  {nombre}: {valor}")

# Ver incrementos/decrementos del cliente 2 en filas anteriores
print("\nIncrementos/Decrementos del cliente 2:")
for fila_check in [15, 16, 17]:
    inc = ws_std_valores.cell(row=fila_check, column=19).value
    dec = ws_std_valores.cell(row=fila_check, column=20).value
    print(f"  Fila {fila_check}: incremento={inc}, decremento={dec}")

wb_formulas.close()
wb_valores.close()
