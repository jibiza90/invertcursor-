from openpyxl import load_workbook

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)

if 'Diario VIP' in wb.sheetnames:
    ws = wb['Diario VIP']
    
    print("=" * 80)
    print("VERIFICANDO ESTRUCTURA DE COLUMNAS POR CLIENTE")
    print("=" * 80)
    
    # Fila 13 tiene los números de clientes
    fila_nombres = 13
    
    print("\nClientes y sus columnas:")
    clientes_info = {}
    
    for col_idx in range(1, min(200, ws.max_column + 1)):
        celda = ws.cell(row=fila_nombres, column=col_idx)
        valor = celda.value
        
        if valor is not None:
            try:
                numero_cliente = int(valor)
                if numero_cliente > 0:
                    # Identificar las columnas de este cliente
                    col_letras = []
                    for offset in range(8):  # K, L, M, N, O, P, Q, R
                        col_num = col_idx + offset
                        if col_num <= ws.max_column:
                            col_letra = ws.cell(row=1, column=col_num).column_letter
                            col_letras.append(f"{col_letra}({col_num})")
                    
                    clientes_info[numero_cliente] = {
                        'columna_inicio': col_idx,
                        'columnas': col_letras
                    }
                    
                    print(f"\nCliente {numero_cliente}:")
                    print(f"  Columna inicio: {col_idx}")
                    print(f"  Columnas: {' '.join(col_letras)}")
                    
                    # Verificar fórmulas en fila 17 (ejemplo)
                    if numero_cliente <= 3:  # Solo mostrar primeros 3
                        print(f"\n  Fórmulas en fila 17:")
                        for offset, nombre in enumerate(['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']):
                            col_num = col_idx + offset
                            celda_formula = ws.cell(row=17, column=col_num)
                            if celda_formula.value and str(celda_formula.value).startswith('='):
                                print(f"    {nombre} ({col_num}): {str(celda_formula.value)[:60]}")
            except:
                pass
    
    print(f"\n\nTotal clientes encontrados: {len(clientes_info)}")
