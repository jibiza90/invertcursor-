from openpyxl import load_workbook

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb.sheetnames:
    ws = wb['Diario VIP']
    
    print("=" * 80)
    print("VERIFICANDO TODAS LAS FILAS DE COLUMNA K (CLIENTE 1)")
    print("=" * 80)
    
    suma_total = 0
    valores_encontrados = []
    
    # Verificar desde fila 15 hasta el final
    for fila_idx in range(15, ws.max_row + 1):
        celda = ws.cell(row=fila_idx, column=11)  # Columna K = 11
        valor = celda.value
        
        if valor is not None:
            try:
                valor_num = float(valor)
                if valor_num != 0:
                    valores_encontrados.append((fila_idx, valor_num))
                    suma_total += valor_num
                    if len(valores_encontrados) <= 20:  # Mostrar primeros 20
                        print(f"  Fila {fila_idx}: {valor_num:,.2f}")
            except:
                pass
    
    print(f"\nTotal valores encontrados: {len(valores_encontrados)}")
    print(f"SUMA TOTAL: {suma_total:,.2f} €")
    
    # Verificar también columna L
    print("\n" + "=" * 80)
    print("VERIFICANDO COLUMNA L (DECREMENTOS)")
    print("=" * 80)
    
    suma_decrementos = 0
    for fila_idx in range(15, ws.max_row + 1):
        celda = ws.cell(row=fila_idx, column=12)  # Columna L = 12
        valor = celda.value
        
        if valor is not None:
            try:
                valor_num = float(valor)
                if valor_num != 0:
                    suma_decrementos += valor_num
            except:
                pass
    
    print(f"SUMA DECREMENTOS: {suma_decrementos:,.2f} €")
