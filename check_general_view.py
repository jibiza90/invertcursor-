#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb['Diario VIP']

print("="*80)
print("VERIFICANDO VISTA GENERAL - FILAS 15-20")
print("="*80)

columnas = {
    'E (imp_inicial)': 5,
    'F (imp_final)': 6,
    'G (benef_euro)': 7,
    'H (benef_porcentaje)': 8,
    'I (benef_euro_acum)': 9,
    'J (benef_porcentaje_acum)': 10
}

for fila in range(15, 21):
    fecha = ws.cell(row=fila, column=4).value
    print(f"\nFila {fila} (Fecha: {fecha}):")
    for col_nombre, col_num in columnas.items():
        celda = ws.cell(row=fila, column=col_num)
        tiene_formula = celda.data_type == 'f' and celda.value is not None
        valor = celda.value
        estado = "🔒 BLOQUEADA (fórmula)" if tiene_formula else "✏️ EDITABLE"
        print(f"  {col_nombre}: {estado}")
        if tiene_formula:
            print(f"    Fórmula: {valor}")

wb.close()
