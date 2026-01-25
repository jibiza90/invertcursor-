from openpyxl import load_workbook
import json
from collections import defaultdict
import sys

# Redirigir salida también a archivo para capturar todo
log_file = open('verificacion_log.txt', 'w', encoding='utf-8')

def print_and_log(*args, **kwargs):
    message = ' '.join(str(arg) for arg in args)
    print(message)
    log_file.write(message + '\n')
    log_file.flush()

print("=" * 80)
print("VERIFICACION COMPLETA: 100% DE LAS CELDAS")
print("=" * 80)
print_and_log("=" * 80)
print_and_log("VERIFICACION COMPLETA: 100% DE LAS CELDAS")
print_and_log("=" * 80)

# Cargar Excel
print("\n1. Cargando Excel...")
try:
    wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
    wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
    ws_formulas = wb_formulas['Diario VIP']
    ws_valores = wb_valores['Diario VIP']
    print(f"   ✓ Excel cargado - Max fila: {ws_valores.max_row}, Max columna: {ws_valores.max_column}")
except Exception as e:
    print(f"   ✗ Error: {e}")
    exit(1)

# Cargar JSON
print("\n2. Cargando JSON...")
try:
    with open('datos_completos.json', 'r', encoding='utf-8') as f:
        datos_json = json.load(f)
    
    if 'Diario VIP' not in datos_json.get('hojas', {}):
        print("   ✗ Error: 'Diario VIP' no encontrado en JSON")
        exit(1)
    
    hoja_json = datos_json['hojas']['Diario VIP']
    datos_generales_json = hoja_json.get('datos_generales', [])
    datos_diarios_json = hoja_json.get('datos_diarios_generales', [])
    clientes_json = hoja_json.get('clientes', [])
    
    print(f"   ✓ JSON cargado")
    print(f"     - Datos generales: {len(datos_generales_json)} filas")
    print(f"     - Datos diarios: {len(datos_diarios_json)} filas")
    print(f"     - Clientes: {len(clientes_json)}")
except Exception as e:
    print(f"   ✗ Error: {e}")
    import traceback
    traceback.print_exc()
    exit(1)

# Mapeos
columna_a_campo_general = {
    5: 'imp_inicial',      # E
    6: 'imp_final',        # F
    7: 'benef_euro',       # G
    8: 'benef_porcentaje', # H
    9: 'benef_euro_acum',  # I
    10: 'benef_porcentaje_acum' # J
}

# Función para convertir número de columna a letra
def columna_a_letra(col_num):
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result

# Contadores
errores = []
advertencias = []
total_celdas_verificadas = 0

print("\n3. Verificando DATOS GENERALES (Filas 3-6, Columnas E-J)...")
# Verificar datos generales
for fila_excel in range(3, 7):
    fila_data_json = next((d for d in datos_generales_json if d.get('fila') == fila_excel), None)
    
    if not fila_data_json:
        errores.append(f"Fila {fila_excel}: No encontrada en JSON (datos_generales)")
        continue
    
    for col_excel, campo in columna_a_campo_general.items():
        total_celdas_verificadas += 1
        col_letra = columna_a_letra(col_excel)
        
        celda_excel = ws_formulas.cell(row=fila_excel, column=col_excel)
        tiene_formula_excel = celda_excel.data_type == 'f' and celda_excel.value is not None
        
        formula_excel = str(celda_excel.value) if tiene_formula_excel else None
        formula_json = fila_data_json.get('formulas', {}).get(campo)
        bloqueada_json = fila_data_json.get('bloqueadas', {}).get(campo, False)
        
        if tiene_formula_excel:
            if not formula_json:
                errores.append(f"Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Excel tiene fórmula pero JSON NO")
            elif formula_excel != formula_json:
                errores.append(f"Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Fórmulas diferentes")
                errores.append(f"  Excel: {formula_excel[:100]}")
                errores.append(f"  JSON:  {formula_json[:100]}")
            if not bloqueada_json:
                errores.append(f"Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Tiene fórmula pero NO está bloqueada")
        else:
            if formula_json:
                advertencias.append(f"Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): JSON tiene fórmula pero Excel NO")
            if bloqueada_json and campo not in ['imp_inicial', 'imp_final']:
                advertencias.append(f"Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Bloqueada sin fórmula")

print(f"   ✓ Verificadas {total_celdas_verificadas} celdas")

print("\n4. Verificando DATOS DIARIOS GENERALES (Filas 15+, Columnas E-J)...")
celdas_diarias_verificadas = 0
for fila_excel in range(15, ws_valores.max_row + 1):
    fecha_excel = ws_valores.cell(row=fila_excel, column=4).value
    if fecha_excel is None:
        continue
    
    fila_data_json = next((d for d in datos_diarios_json if d.get('fila') == fila_excel), None)
    
    if not fila_data_json:
        advertencias.append(f"Fila {fila_excel} (datos_diarios): No encontrada en JSON")
        continue
    
    for col_excel, campo in columna_a_campo_general.items():
        total_celdas_verificadas += 1
        celdas_diarias_verificadas += 1
        col_letra = columna_a_letra(col_excel)
        
        celda_excel = ws_formulas.cell(row=fila_excel, column=col_excel)
        tiene_formula_excel = celda_excel.data_type == 'f' and celda_excel.value is not None
        
        formula_excel = str(celda_excel.value) if tiene_formula_excel else None
        formula_json = fila_data_json.get('formulas', {}).get(campo)
        bloqueada_json = fila_data_json.get('bloqueadas', {}).get(campo, False)
        
        if tiene_formula_excel:
            if not formula_json:
                errores.append(f"Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Excel tiene fórmula pero JSON NO")
            elif formula_excel != formula_json:
                errores.append(f"Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Fórmulas diferentes")
                errores.append(f"  Excel: {formula_excel[:100]}")
                errores.append(f"  JSON:  {formula_json[:100]}")
            if not bloqueada_json:
                errores.append(f"Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Tiene fórmula pero NO está bloqueada")
        else:
            if formula_json:
                advertencias.append(f"Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): JSON tiene fórmula pero Excel NO")
            if bloqueada_json and campo not in ['imp_inicial', 'imp_final']:
                valor_excel = ws_valores.cell(row=fila_excel, column=col_excel).value
                if valor_excel is None or valor_excel == 0:
                    advertencias.append(f"Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Bloqueada sin fórmula ni valor")

print(f"   ✓ Verificadas {celdas_diarias_verificadas} celdas")

print("\n5. Verificando CLIENTES (Filas 15+, Columnas K+)...")
# Identificar columnas de clientes
clientes_columnas = {}
for col_excel in range(11, min(ws_valores.max_column + 1, 200)):  # K en adelante, máximo columna 200
    celda_numero = ws_formulas.cell(row=13, column=col_excel).value  # Fila 13 tiene números de cliente
    if celda_numero is not None:
        try:
            numero_cliente = int(celda_numero)
            if numero_cliente > 0:
                if numero_cliente not in clientes_columnas:
                    clientes_columnas[numero_cliente] = []
                clientes_columnas[numero_cliente].append(col_excel)
        except:
            pass

print(f"   Encontrados {len(clientes_columnas)} clientes en Excel")

# Mapeo de offset de columna a campo
offset_a_campo = {
    0: 'incremento',      # K, S, AA...
    1: 'decremento',      # L, T, AB...
    2: 'base',            # M, U, AC...
    3: 'saldo_diario',    # N, V, AD...
    4: 'beneficio_diario', # O, W, AE...
    5: 'beneficio_diario_pct', # P, X, AF...
    6: 'beneficio_acumulado', # Q, Y, AG...
    7: 'beneficio_acumulado_pct' # R, Z, AH...
}

celdas_clientes_verificadas = 0
for numero_cliente, columnas_cliente in sorted(clientes_columnas.items()):
    cliente_json = next((c for c in clientes_json if c.get('numero_cliente') == numero_cliente), None)
    
    if not cliente_json:
        advertencias.append(f"Cliente {numero_cliente}: No encontrado en JSON")
        continue
    
    datos_diarios_cliente = cliente_json.get('datos_diarios', [])
    columna_inicio = min(columnas_cliente)
    
    print(f"   Verificando Cliente {numero_cliente} (columnas {columna_a_letra(columna_inicio)}-{columna_a_letra(columna_inicio+7)})...")
    
    for fila_excel in range(15, ws_valores.max_row + 1):
        fecha_excel = ws_valores.cell(row=fila_excel, column=4).value
        if fecha_excel is None:
            continue
        
        fila_data_json = next((d for d in datos_diarios_cliente if d.get('fila') == fila_excel), None)
        
        if not fila_data_json:
            continue
        
        # Verificar cada columna del cliente (8 columnas por cliente)
        for offset in range(8):
            col_excel = columna_inicio + offset
            if col_excel > ws_valores.max_column:
                break
            
            campo = offset_a_campo[offset]
            total_celdas_verificadas += 1
            celdas_clientes_verificadas += 1
            col_letra = columna_a_letra(col_excel)
            
            celda_excel = ws_formulas.cell(row=fila_excel, column=col_excel)
            tiene_formula_excel = celda_excel.data_type == 'f' and celda_excel.value is not None
            
            formula_excel = str(celda_excel.value) if tiene_formula_excel else None
            formula_json = fila_data_json.get('formulas', {}).get(campo)
            bloqueada_json = fila_data_json.get('bloqueadas', {}).get(campo, False)
            
            if tiene_formula_excel:
                if not formula_json:
                    errores.append(f"Cliente {numero_cliente}, Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Excel tiene fórmula pero JSON NO")
                elif formula_excel != formula_json:
                    errores.append(f"Cliente {numero_cliente}, Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Fórmulas diferentes")
                    errores.append(f"  Excel: {formula_excel[:100]}")
                    errores.append(f"  JSON:  {formula_json[:100]}")
                if not bloqueada_json:
                    errores.append(f"Cliente {numero_cliente}, Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): Tiene fórmula pero NO está bloqueada")
            else:
                # Incremento y decremento NUNCA deben estar bloqueados
                if campo in ['incremento', 'decremento']:
                    if bloqueada_json:
                        errores.append(f"Cliente {numero_cliente}, Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): NO debe estar bloqueada (es editable)")
                elif formula_json:
                    advertencias.append(f"Cliente {numero_cliente}, Fila {fila_excel}, {col_letra}{fila_excel} ({campo}): JSON tiene fórmula pero Excel NO")

print(f"   ✓ Verificadas {celdas_clientes_verificadas} celdas de clientes")

# Generar reporte
print("\n6. Generando reporte...")
with open('reporte_verificacion_completo.txt', 'w', encoding='utf-8') as f:
    f.write("=" * 80 + "\n")
    f.write("REPORTE DE VERIFICACION COMPLETA: 100% DE LAS CELDAS\n")
    f.write("=" * 80 + "\n\n")
    
    f.write(f"RESUMEN:\n")
    f.write(f"  Total de celdas verificadas: {total_celdas_verificadas:,}\n")
    f.write(f"  - Datos generales: {4 * 6}\n")
    f.write(f"  - Datos diarios generales: {celdas_diarias_verificadas:,}\n")
    f.write(f"  - Celdas de clientes: {celdas_clientes_verificadas:,}\n")
    f.write(f"  Errores encontrados: {len(errores)}\n")
    f.write(f"  Advertencias: {len(advertencias)}\n\n")
    
    if errores:
        f.write("=" * 80 + "\n")
        f.write("ERRORES CRÍTICOS\n")
        f.write("=" * 80 + "\n\n")
        for i, error in enumerate(errores, 1):
            f.write(f"{i}. {error}\n")
    
    if advertencias:
        f.write("\n" + "=" * 80 + "\n")
        f.write("ADVERTENCIAS\n")
        f.write("=" * 80 + "\n\n")
        for i, adv in enumerate(advertencias[:100], 1):  # Limitar a 100 advertencias
            f.write(f"{i}. {adv}\n")
        if len(advertencias) > 100:
            f.write(f"\n... y {len(advertencias) - 100} advertencias más\n")

print("\n" + "=" * 80)
print("VERIFICACION COMPLETADA")
print("=" * 80)
print(f"Total de celdas verificadas: {total_celdas_verificadas:,}")
print(f"Errores encontrados: {len(errores)}")
print(f"Advertencias: {len(advertencias)}")
print(f"\nReporte guardado en: reporte_verificacion_completo.txt")
print_and_log(f"\nReporte guardado en: reporte_verificacion_completo.txt")

if errores:
    print(f"\n❌ Se encontraron {len(errores)} errores críticos")
    print("   Revisa el reporte para más detalles")
    print_and_log(f"\n❌ Se encontraron {len(errores)} errores críticos")
else:
    print("\n✅ No se encontraron errores críticos")
    print_and_log("\n✅ No se encontraron errores críticos")

log_file.close()
