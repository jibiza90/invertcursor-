#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para verificar EXACTAMENTE cómo están las celdas en el Excel
"""

import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb['Diario VIP']

print("="*80)
print("VERIFICANDO EXCEL - VISTA GENERAL (FILAS 15-30)")
print("="*80)
print("\nColumnas:")
print("  E (5) = IMP. INICIAL")
print("  F (6) = IMP. FINAL")
print("  G (7) = BENEF. €")
print("  H (8) = BENEF. %")
print("  I (9) = BENEF. € ACUM")
print("  J (10) = BENEF. % ACUM")
print("\n" + "="*80)

for fila in range(15, 31):
    fecha = ws.cell(row=fila, column=4).value
    print(f"\nFila {fila} (Fecha: {fecha}):")
    
    for col_nombre, col_num in [('E (imp_inicial)', 5), ('F (imp_final)', 6), 
                                 ('G (benef_euro)', 7), ('H (benef_porcentaje)', 8),
                                 ('I (benef_euro_acum)', 9), ('J (benef_porcentaje_acum)', 10)]:
        celda = ws.cell(row=fila, column=col_num)
        tiene_formula = celda.data_type == 'f' and celda.value is not None
        
        # Verificar protección de celda
        proteccion = celda.protection
        esta_bloqueada_proteccion = proteccion.locked if proteccion else False
        
        valor = celda.value
        estado = "🔒 BLOQUEADA" if tiene_formula else "✏️ EDITABLE"
        if tiene_formula:
            estado += f" (fórmula: {valor})"
        
        print(f"  {col_nombre}: {estado}")
        if esta_bloqueada_proteccion:
            print(f"    ⚠️ Protección Excel: LOCKED")

wb.close()
