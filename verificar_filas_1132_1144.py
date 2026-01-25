#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para verificar específicamente las filas 1132-1144 que tenían problemas
"""

import json
import openpyxl
import sys

def obtener_nombre_columna(num_col):
    result = ""
    num = num_col
    while num > 0:
        num -= 1
        result = chr(65 + (num % 26)) + result
        num //= 26
    return result

def tiene_formula(ws, fila, columna):
    try:
        celda = ws.cell(row=fila, column=columna)
        return celda.data_type == 'f' and celda.value is not None
    except:
        return False

def main():
    sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None
    
    errores = []
    
    # Cargar Excel
    print("Cargando Excel...")
    wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
    ws = None
    for sheet_name in wb.sheetnames:
        if 'VIP' in sheet_name.upper():
            ws = wb[sheet_name]
            break
    if not ws:
        ws = wb.active
    
    # Cargar JSON
    print("Cargando JSON...")
    with open('datos_completos.json', 'r', encoding='utf-8') as f:
        datos_json = json.load(f)
    
    hoja_vip = datos_json.get('hojas', {}).get('Diario VIP')
    if not hoja_vip:
        print("ERROR: No se encontró hoja Diario VIP")
        return
    
    # Verificar específicamente las filas 1132-1144 del cliente 1
    print("\nVerificando filas 1132-1144 del Cliente 1...")
    clientes = hoja_vip.get('clientes', [])
    cliente1 = clientes[0] if clientes else None
    
    if not cliente1:
        print("ERROR: No se encontró Cliente 1")
        return
    
    columna_inicio = cliente1.get('columna_inicio', 11)
    datos_diarios = cliente1.get('datos_diarios', [])
    
    for fila in range(1132, 1145):
        dato = next((d for d in datos_diarios if d.get('fila') == fila), None)
        if not dato:
            continue
        
        bloqueadas = dato.get('bloqueadas', {})
        
        # Verificar INCREMENTO
        col_inc = columna_inicio
        tiene_formula_excel_inc = tiene_formula(ws, fila, col_inc)
        esta_bloqueada_json_inc = bloqueadas.get('incremento', False)
        
        try:
            celda_inc = ws.cell(row=fila, column=col_inc)
            valor_inc = celda_inc.value
            data_type_inc = celda_inc.data_type
        except:
            valor_inc = None
            data_type_inc = None
        
        print(f"\nFila {fila} - INCREMENTO (Col {obtener_nombre_columna(col_inc)}):")
        print(f"  Excel - data_type: {data_type_inc}, value: {valor_inc}")
        print(f"  JSON - bloqueada: {esta_bloqueada_json_inc}")
        print(f"  Tiene fórmula Excel: {tiene_formula_excel_inc}")
        
        if tiene_formula_excel_inc and not esta_bloqueada_json_inc:
            errores.append({
                'fila': fila,
                'campo': 'incremento',
                'columna': obtener_nombre_columna(col_inc),
                'formula': str(valor_inc) if valor_inc else None
            })
            print(f"  ❌ ERROR: Tiene fórmula pero NO está bloqueada")
        
        # Verificar DECREMENTO
        col_dec = columna_inicio + 1
        tiene_formula_excel_dec = tiene_formula(ws, fila, col_dec)
        esta_bloqueada_json_dec = bloqueadas.get('decremento', False)
        
        try:
            celda_dec = ws.cell(row=fila, column=col_dec)
            valor_dec = celda_dec.value
            data_type_dec = celda_dec.data_type
        except:
            valor_dec = None
            data_type_dec = None
        
        print(f"\nFila {fila} - DECREMENTO (Col {obtener_nombre_columna(col_dec)}):")
        print(f"  Excel - data_type: {data_type_dec}, value: {valor_dec}")
        print(f"  JSON - bloqueada: {esta_bloqueada_json_dec}")
        print(f"  Tiene fórmula Excel: {tiene_formula_excel_dec}")
        
        if tiene_formula_excel_dec and not esta_bloqueada_json_dec:
            errores.append({
                'fila': fila,
                'campo': 'decremento',
                'columna': obtener_nombre_columna(col_dec),
                'formula': str(valor_dec) if valor_dec else None
            })
            print(f"  ❌ ERROR: Tiene fórmula pero NO está bloqueada")
    
    print(f"\n{'='*80}")
    print(f"Total de errores encontrados: {len(errores)}")
    
    if errores:
        print("\nERRORES:")
        for error in errores:
            print(f"  Fila {error['fila']}, Col {error['columna']} ({error['campo']}): {error.get('formula', 'N/A')}")
    else:
        print("\n✅ No se encontraron errores en estas filas")
    
    wb.close()

if __name__ == '__main__':
    main()
