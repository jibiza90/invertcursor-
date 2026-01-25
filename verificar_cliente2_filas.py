#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificar qué fila debe tener el incremento del cliente 2 y cómo se calculan las fórmulas
"""

import json
import openpyxl

# Cargar Excel
wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)

ws_std = None
for sheet_name in wb.sheetnames:
    if 'STD' in sheet_name.upper() and 'VIP' not in sheet_name.upper():
        ws_std = wb[sheet_name]
        print(f"Hoja encontrada: {sheet_name}")
        break

# Cargar JSON
with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos = json.load(f)

hoja_std = datos.get('hojas', {}).get('Diario STD')

print("="*80)
print("ANÁLISIS DE CLIENTE 2 - FILAS Y FÓRMULAS")
print("="*80)

# Cliente 2 (índice 1)
clientes = hoja_std.get('clientes', [])
if len(clientes) > 1:
    cliente2 = clientes[1]
    numero = cliente2.get('numero_cliente')
    columna_inicio = cliente2.get('columna_inicio', 0)
    
    print(f"\nCliente {numero}:")
    print(f"  Columna inicio: {columna_inicio} ({chr(64 + columna_inicio) if columna_inicio <= 26 else '?'})")
    
    # Cliente 2 empieza en columna S (19)
    # S = incremento, T = decremento, U = base, V = saldo_diario, W = beneficio_diario, X = beneficio_diario_pct, Y = beneficio_acumulado
    
    datos_diarios_cliente = cliente2.get('datos_diarios', [])
    
    # Ver todas las filas del día 1 (filas 15, 16, 17)
    for fila in [15, 16, 17]:
        dato = [d for d in datos_diarios_cliente if d.get('fila') == fila]
        if dato:
            d = dato[0]
            print(f"\nFila {fila}:")
            print(f"  incremento (S{fila}): {d.get('incremento')}")
            print(f"  decremento (T{fila}): {d.get('decremento')}")
            print(f"  base (U{fila}): {d.get('base')}")
            print(f"  saldo_diario (V{fila}): {d.get('saldo_diario')}")
            print(f"  beneficio_diario (W{fila}): {d.get('beneficio_diario')}")
            print(f"  beneficio_diario_pct (X{fila}): {d.get('beneficio_diario_pct')}")
            
            formulas = d.get('formulas', {})
            if formulas:
                print(f"  Fórmulas:")
                for campo, formula in formulas.items():
                    print(f"    {campo}: {formula}")
            
            # Verificar en Excel qué fila tiene el incremento editable
            if ws_std:
                celda_incremento = ws_std.cell(row=fila, column=columna_inicio)
                celda_decremento = ws_std.cell(row=fila, column=columna_inicio + 1)
                
                tiene_formula_inc = celda_incremento.data_type == 'f' and celda_incremento.value
                tiene_formula_dec = celda_decremento.data_type == 'f' and celda_decremento.value
                
                print(f"  Excel - Incremento tiene fórmula: {bool(tiene_formula_inc)}")
                print(f"  Excel - Decremento tiene fórmula: {bool(tiene_formula_dec)}")
                if tiene_formula_inc:
                    print(f"    Fórmula incremento: {celda_incremento.value}")
                if tiene_formula_dec:
                    print(f"    Fórmula decremento: {celda_decremento.value}")

wb.close()
