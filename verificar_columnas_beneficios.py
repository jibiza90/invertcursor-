from openpyxl import load_workbook
from datetime import datetime

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb_formulas.sheetnames:
    ws_formulas = wb_formulas['Diario VIP']
    ws_valores = wb_valores['Diario VIP']
    
    print("=" * 80)
    print("VERIFICACION COLUMNAS DE BENEFICIOS EN DATOS GENERALES")
    print("=" * 80)
    
    # Verificar filas de datos generales (3-6)
    print("\nFilas de datos generales (3-6):")
    for fila_idx in range(3, 7):
        print(f"\nFila {fila_idx}:")
        
        columnas = {
            5: 'E - Imp. Inicial',
            6: 'F - Imp. Final',
            7: 'G - Benef. €',
            8: 'H - Benef. %',
            9: 'I - Benef. € Acum',
            10: 'J - Benef. % Acum'
        }
        
        for col_idx, nombre_col in columnas.items():
            celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
            celda_valor = ws_valores.cell(row=fila_idx, column=col_idx)
            
            tiene_formula = False
            formula_texto = None
            
            if celda_formula.data_type == 'f':
                tiene_formula = True
                formula_texto = str(celda_formula.value)[:80]
            elif celda_formula.value is not None:
                valor_str = str(celda_formula.value).strip()
                if valor_str.startswith('='):
                    tiene_formula = True
                    formula_texto = valor_str[:80]
            
            estado = "[BLOQUEADA]" if tiene_formula else "[EDITABLE]"
            valor = celda_valor.value
            print(f"  {nombre_col}: {valor} - {estado}")
            if formula_texto:
                print(f"    Formula: {formula_texto}")
    
    # Verificar algunas filas de datos diarios (15, 17, 27, etc.)
    print("\n" + "=" * 80)
    print("Filas de datos diarios (ejemplos):")
    print("=" * 80)
    
    filas_ejemplo = [15, 17, 27, 51, 54]
    for fila_idx in filas_ejemplo:
        fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
        if fecha_valor:
            try:
                if isinstance(fecha_valor, datetime):
                    fecha_str = fecha_valor.strftime('%Y-%m-%d')
                else:
                    fecha_str = str(fecha_valor).split()[0]
                
                print(f"\nFila {fila_idx} (Fecha: {fecha_str}):")
                
                columnas_beneficios = {
                    7: 'G - Benef. €',
                    8: 'H - Benef. %',
                    9: 'I - Benef. € Acum',
                    10: 'J - Benef. % Acum'
                }
                
                for col_idx, nombre_col in columnas_beneficios.items():
                    celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                    celda_valor = ws_valores.cell(row=fila_idx, column=col_idx)
                    
                    tiene_formula = False
                    formula_texto = None
                    
                    if celda_formula.data_type == 'f':
                        tiene_formula = True
                        formula_texto = str(celda_formula.value)[:80]
                    elif celda_formula.value is not None:
                        valor_str = str(celda_formula.value).strip()
                        if valor_str.startswith('='):
                            tiene_formula = True
                            formula_texto = valor_str[:80]
                    
                    estado = "[BLOQUEADA]" if tiene_formula else "[EDITABLE]"
                    valor = celda_valor.value
                    print(f"  {nombre_col}: {valor} - {estado}")
                    if formula_texto:
                        print(f"    Formula: {formula_texto}")
                        
            except Exception as e:
                pass
