from openpyxl import load_workbook
import json

# Verificar directamente en Excel qué pasa con las filas problemáticas
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("Verificando filas 1132-1144 del Cliente 1 (columna K = incremento, columna L = decremento):")
print("=" * 80)

# Cliente 1 está en columna K (columna 11)
for fila in range(1132, 1145):
    celda_k = ws.cell(row=fila, column=11)  # Columna K
    celda_l = ws.cell(row=fila, column=12)  # Columna L
    
    tiene_formula_k = celda_k.data_type == 'f' and celda_k.value
    tiene_formula_l = celda_l.data_type == 'f' and celda_l.value
    
    fecha = ws.cell(row=fila, column=4).value
    
    print(f"\nFila {fila} (Fecha: {fecha}):")
    print(f"  Columna K (incremento):")
    print(f"    Tipo: {celda_k.data_type}")
    print(f"    Tiene fórmula: {tiene_formula_k}")
    if tiene_formula_k:
        print(f"    Fórmula: {celda_k.value}")
    else:
        print(f"    Valor: {celda_k.value}")
    
    print(f"  Columna L (decremento):")
    print(f"    Tipo: {celda_l.data_type}")
    print(f"    Tiene fórmula: {tiene_formula_l}")
    if tiene_formula_l:
        print(f"    Fórmula: {celda_l.value}")
    else:
        print(f"    Valor: {celda_l.value}")

# Verificar también la función tiene_formula del script
def tiene_formula(ws_formulas, fila, columna):
    """Verifica si una celda tiene fórmula"""
    try:
        celda = ws_formulas.cell(row=fila, column=columna)
        
        # Método 1: Verificar el tipo de dato (más confiable)
        if celda.data_type == 'f':  # 'f' significa fórmula
            return True
        
        # Método 2: Verificar si el valor empieza con '='
        if celda.value is not None:
            valor_str = str(celda.value).strip()
            if valor_str.startswith('='):
                return True
        
        return False
    except Exception as e:
        return False

print("\n\nVerificando con la función tiene_formula del script:")
print("=" * 80)
for fila in range(1132, 1145):
    tiene_k = tiene_formula(ws, fila, 11)
    tiene_l = tiene_formula(ws, fila, 12)
    print(f"Fila {fila}: K tiene fórmula: {tiene_k}, L tiene fórmula: {tiene_l}")
