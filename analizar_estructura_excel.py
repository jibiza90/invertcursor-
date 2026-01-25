from openpyxl import load_workbook
import json

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

# Configurar encoding para Windows
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

print("=" * 80)
print("ANALISIS COMPLETO DEL EXCEL")
print("=" * 80)

# Mostrar todas las hojas
print("\nHojas disponibles:")
for i, sheet in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet}")

# Analizar Diario STD
print("\n" + "=" * 80)
print("ANALISIS: Diario STD")
print("=" * 80)

if 'Diario STD' in wb.sheetnames:
    ws_std = wb['Diario STD']
    filas_std = []
    for row in ws_std.iter_rows(values_only=True):
        filas_std.append(list(row))
    
    print(f"Total de filas: {len(filas_std)}")
    print(f"Total de columnas: {len(filas_std[0]) if filas_std else 0}")
    
    # Mostrar primeras 20 filas para entender estructura
    print("\nPrimeras 20 filas:")
    for i, fila in enumerate(filas_std[:20], 1):
        valores = [str(v)[:30] if v is not None else '' for v in fila[:15]]
        print(f"Fila {i}: {valores}")

# Analizar Diario VIP
print("\n" + "=" * 80)
print("ANALISIS: Diario VIP")
print("=" * 80)

if 'Diario VIP' in wb.sheetnames:
    ws_vip = wb['Diario VIP']
    filas_vip = []
    for row in ws_vip.iter_rows(values_only=True):
        filas_vip.append(list(row))
    
    print(f"Total de filas: {len(filas_vip)}")
    print(f"Total de columnas: {len(filas_vip[0]) if filas_vip else 0}")
    
    # Buscar secciones
    print("\nBuscando secciones importantes:")
    
    # Buscar filas con texto que indique secciones
    secciones = []
    for i, fila in enumerate(filas_vip, 1):
        fila_str = ' '.join([str(v) for v in fila if v is not None]).lower()
        
        # Buscar indicadores de sección
        if any(palabra in fila_str for palabra in ['total', 'general', 'resumen', 'cliente', 'mes', 'fecha']):
            valores_no_vacios = [v for v in fila[:20] if v is not None and str(v).strip() != '']
            if len(valores_no_vacios) > 3:
                secciones.append({
                    'fila': i,
                    'valores': valores_no_vacios[:10]
                })
    
    print(f"\nSecciones encontradas (primeras 10):")
    for sec in secciones[:10]:
        print(f"  Fila {sec['fila']}: {sec['valores']}")
    
    # Analizar estructura de headers
    print("\nAnalizando headers:")
    for i in range(min(20, len(filas_vip))):
        fila = filas_vip[i]
        headers_encontrados = [str(v) for v in fila[:20] if v is not None and str(v).strip() != '']
        if len(headers_encontrados) > 5:
            print(f"  Fila {i+1} (posible header): {headers_encontrados[:10]}")
    
    # Buscar donde empiezan los datos de clientes
    print("\nBuscando inicio de datos de clientes:")
    for i in range(min(50, len(filas_vip))):
        fila = filas_vip[i]
        # Buscar fila con "MES" y "FECHA" como headers
        fila_str = ' '.join([str(v) for v in fila[:10] if v is not None]).upper()
        if 'MES' in fila_str and 'FECHA' in fila_str:
            print(f"  Fila {i+1}: Posible header de clientes")
            print(f"    Valores: {[str(v) for v in fila[:15] if v is not None]}")
            # Mostrar siguientes 5 filas
            print(f"    Siguientes filas:")
            for j in range(i+1, min(i+6, len(filas_vip))):
                siguiente = filas_vip[j]
                valores = [str(v)[:20] if v is not None else '' for v in siguiente[:10]]
                print(f"      Fila {j+1}: {valores}")
            break
    
    # Buscar datos generales/totales (antes de los clientes)
    print("\nBuscando datos generales/totales:")
    for i in range(min(30, len(filas_vip))):
        fila = filas_vip[i]
        fila_str = ' '.join([str(v) for v in fila if v is not None]).lower()
        
        # Buscar indicadores de totales/general
        if any(palabra in fila_str for palabra in ['total', 'inversion', 'beneficio', 'saldo']):
            valores = [v for v in fila[:20] if v is not None]
            if len(valores) > 3:
                print(f"  Fila {i+1}: {valores[:10]}")

print("\n" + "=" * 80)
print("Analisis completado")
print("=" * 80)
