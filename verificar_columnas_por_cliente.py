from openpyxl import load_workbook

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb.sheetnames:
    ws = wb['Diario VIP']
    
    print("=" * 80)
    print("VERIFICANDO COLUMNAS POR CLIENTE")
    print("=" * 80)
    
    # Fila 13 tiene los números de clientes
    fila_nombres = 13
    
    print("\nClientes encontrados:")
    clientes = {}
    
    for col_idx in range(1, min(200, ws.max_column + 1)):
        celda = ws.cell(row=fila_nombres, column=col_idx)
        valor = celda.value
        
        if valor is not None:
            try:
                numero_cliente = int(valor)
                if numero_cliente > 0:
                    # Calcular letras de columna
                    def num_to_col(num):
                        result = ""
                        while num > 0:
                            num -= 1
                            result = chr(65 + (num % 26)) + result
                            num //= 26
                        return result
                    
                    col_inicio = num_to_col(col_idx)
                    col_k = num_to_col(col_idx)  # INCREMENTO
                    col_l = num_to_col(col_idx + 1)  # DECREMENTO
                    col_m = num_to_col(col_idx + 2)  # INICIAL
                    col_n = num_to_col(col_idx + 3)  # FINAL
                    col_o = num_to_col(col_idx + 4)  # BENEF €
                    col_p = num_to_col(col_idx + 5)  # BENEF %
                    col_q = num_to_col(col_idx + 6)  # BENEF € acum
                    col_r = num_to_col(col_idx + 7)  # BENEF % acum
                    
                    clientes[numero_cliente] = {
                        'col_inicio': col_idx,
                        'K': col_k,
                        'L': col_l,
                        'M': col_m,
                        'N': col_n,
                        'O': col_o,
                        'P': col_p,
                        'Q': col_q,
                        'R': col_r
                    }
                    
                    print(f"\nCliente {numero_cliente}:")
                    print(f"  Columna inicio: {col_inicio} ({col_idx})")
                    print(f"  K (Incremento): {col_k} ({col_idx})")
                    print(f"  L (Decremento): {col_l} ({col_idx + 1})")
                    print(f"  M (Inicial): {col_m} ({col_idx + 2})")
                    print(f"  N (Final/Saldo): {col_n} ({col_idx + 3}) - BLOQUEADA")
                    print(f"  O (Benef €): {col_o} ({col_idx + 4}) - BLOQUEADA")
                    print(f"  P (Benef %): {col_p} ({col_idx + 5}) - BLOQUEADA")
                    print(f"  Q (Benef € acum): {col_q} ({col_idx + 6}) - BLOQUEADA")
                    print(f"  R (Benef % acum): {col_r} ({col_idx + 7}) - BLOQUEADA")
                    
                    if numero_cliente <= 3:
                        # Verificar una fila de ejemplo (fila 17)
                        print(f"\n  Valores ejemplo en fila 17:")
                        print(f"    {col_k}17: {ws.cell(row=17, column=col_idx).value}")
                        print(f"    {col_l}17: {ws.cell(row=17, column=col_idx + 1).value}")
                        print(f"    {col_n}17: {ws.cell(row=17, column=col_idx + 3).value}")
                        print(f"    {col_o}17: {ws.cell(row=17, column=col_idx + 4).value}")
            except:
                pass
    
    print(f"\n\nTotal clientes: {len(clientes)}")
