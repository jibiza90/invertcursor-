from openpyxl import load_workbook
import json
import traceback

try:
    print("Iniciando verificacion...")
    
    # Cargar Excel
    print("Cargando Excel...")
    wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
    ws_formulas = wb_formulas['Diario VIP']
    print(f"Excel cargado - Max fila: {ws_formulas.max_row}")
    
    # Cargar JSON
    print("Cargando JSON...")
    with open('datos_completos.json', 'r', encoding='utf-8') as f:
        datos_json = json.load(f)
    
    hoja_json = datos_json['hojas']['Diario VIP']
    clientes_json = hoja_json.get('clientes', [])
    datos_diarios_json = hoja_json.get('datos_diarios_generales', [])
    datos_generales_json = hoja_json.get('datos_generales', [])
    
    print(f"JSON cargado - {len(clientes_json)} clientes")
    
    errores = []
    total = 0
    
    # Verificar cliente 1 primero como prueba
    cliente1 = next((c for c in clientes_json if c.get('numero_cliente') == 1), None)
    if cliente1:
        datos_cliente1 = cliente1.get('datos_diarios', [])
        print(f"\nVerificando Cliente 1 - {len(datos_cliente1)} filas...")
        
        for fila_data in datos_cliente1[:10]:  # Primeras 10 filas como prueba
            fila = fila_data.get('fila')
            if fila < 15:
                continue
            
            # Verificar columna M (base) - columna 13
            celda = ws_formulas.cell(row=fila, column=13)
            tiene_formula = celda.data_type == 'f' and celda.value is not None
            
            formula_excel = str(celda.value) if tiene_formula else None
            formula_json = fila_data.get('formulas', {}).get('base')
            bloqueada = fila_data.get('bloqueadas', {}).get('base', False)
            
            total += 1
            
            if tiene_formula:
                if not formula_json:
                    errores.append(f"Fila {fila}, M{fila} (base): Excel tiene formula pero JSON NO")
                elif formula_excel != formula_json:
                    errores.append(f"Fila {fila}, M{fila} (base): Formulas diferentes")
                if not bloqueada:
                    errores.append(f"Fila {fila}, M{fila} (base): Tiene formula pero NO bloqueada")
    
    print(f"\nVerificadas {total} celdas de prueba")
    print(f"Errores encontrados: {len(errores)}")
    
    if errores:
        print("\nErrores:")
        for error in errores[:10]:
            print(f"  - {error}")
    
    # Escribir reporte simple
    with open('reporte_simple.txt', 'w', encoding='utf-8') as f:
        f.write("REPORTE DE VERIFICACION\n")
        f.write(f"Total celdas verificadas: {total}\n")
        f.write(f"Errores: {len(errores)}\n\n")
        for error in errores:
            f.write(f"{error}\n")
    
    print("\nReporte guardado en: reporte_simple.txt")
    
except Exception as e:
    print(f"ERROR: {e}")
    traceback.print_exc()
    with open('error_verificacion.txt', 'w', encoding='utf-8') as f:
        f.write(f"ERROR: {e}\n\n")
        traceback.print_exc(file=f)
