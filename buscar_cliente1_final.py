from openpyxl import load_workbook
import json

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

# Buscar Diario VIP
sheet_name = None
for sheet in wb.sheetnames:
    if 'Diario VIP' in sheet or 'Diario_VIP' in sheet or 'diario_vip' in sheet.lower():
        sheet_name = sheet
        break

if sheet_name:
    ws = wb[sheet_name]
    
    # Leer todas las celdas para encontrar la estructura real
    todas_las_filas = []
    for row in ws.iter_rows(values_only=True):
        todas_las_filas.append(list(row))
    
    # Buscar fila con "Nombre cliente" para identificar dónde empiezan los datos de clientes
    fila_nombre_cliente = None
    for idx, fila in enumerate(todas_las_filas):
        fila_str = ' '.join([str(c) for c in fila if c is not None]).lower()
        if 'nombre cliente' in fila_str:
            fila_nombre_cliente = idx
            break
    
    # Buscar fila con número 1 que podría ser el ID del cliente 1
    cliente_1_fila_idx = None
    cliente_1_datos = {}
    
    if fila_nombre_cliente is not None:
        # Buscar después de la fila "Nombre cliente"
        for idx in range(fila_nombre_cliente + 1, min(fila_nombre_cliente + 100, len(todas_las_filas))):
            fila = todas_las_filas[idx]
            # Buscar si hay un 1 en alguna de las primeras columnas (posible ID)
            for col_idx in range(min(10, len(fila))):
                if fila[col_idx] == 1 or fila[col_idx] == '1':
                    # Esta podría ser la fila del cliente 1
                    cliente_1_fila_idx = idx
                    # Leer todas las columnas de esta fila
                    for col_idx, valor in enumerate(fila):
                        if valor is not None and str(valor).strip() != '':
                            cliente_1_datos[f'Columna_{col_idx+1}'] = valor
                    break
            if cliente_1_fila_idx:
                break
    
    # Si no encontramos por el método anterior, buscar cualquier fila con 1 y datos significativos
    if not cliente_1_datos:
        for idx, fila in enumerate(todas_las_filas):
            # Buscar fila que tenga un 1 y varios otros valores
            tiene_uno = False
            valores_significativos = 0
            datos_fila = {}
            
            for col_idx, valor in enumerate(fila):
                if valor == 1 or valor == '1':
                    tiene_uno = True
                if valor is not None and str(valor).strip() != '':
                    valores_significativos += 1
                    datos_fila[f'Columna_{col_idx+1}'] = valor
            
            # Si tiene un 1 y al menos 5 valores significativos, probablemente es una fila de datos
            if tiene_uno and valores_significativos >= 5:
                cliente_1_fila_idx = idx
                cliente_1_datos = datos_fila
                break
    
    # También buscar headers reales
    headers = {}
    if fila_nombre_cliente is not None:
        fila_header = todas_las_filas[fila_nombre_cliente]
        for col_idx, valor in enumerate(fila_header):
            if valor is not None and str(valor).strip() != '':
                headers[col_idx] = str(valor).strip()
    
    # Crear resultado
    resultado = {
        "hoja": sheet_name,
        "fila_nombre_cliente_encontrada": fila_nombre_cliente,
        "fila_cliente_1": cliente_1_fila_idx,
        "headers": headers,
        "cliente_1": cliente_1_datos,
        "total_columnas": len(todas_las_filas[0]) if todas_las_filas else 0,
        "total_filas": len(todas_las_filas)
    }
    
    # Escribir archivo de texto
    with open('CLIENTE1_FINAL.txt', 'w', encoding='utf-8') as f:
        f.write('='*80 + '\n')
        f.write(f'INFORMACIÓN COMPLETA DEL CLIENTE 1 - HOJA: {sheet_name}\n')
        f.write('='*80 + '\n\n')
        
        f.write(f'Total de filas en la hoja: {len(todas_las_filas)}\n')
        f.write(f'Total de columnas: {len(todas_las_filas[0]) if todas_las_filas else 0}\n')
        f.write(f'Fila con "Nombre cliente": {fila_nombre_cliente + 1 if fila_nombre_cliente is not None else "No encontrada"}\n')
        f.write(f'Fila del Cliente 1: {cliente_1_fila_idx + 1 if cliente_1_fila_idx is not None else "No encontrada"}\n\n')
        
        if headers:
            f.write('Headers encontrados:\n')
            for col_idx, header in sorted(headers.items()):
                f.write(f'  Columna {col_idx+1}: {header}\n')
            f.write('\n')
        
        f.write('='*80 + '\n')
        f.write('DATOS DEL CLIENTE 1:\n')
        f.write('='*80 + '\n\n')
        
        if cliente_1_datos:
            # Ordenar por número de columna
            columnas_ordenadas = sorted(cliente_1_datos.items(), 
                                      key=lambda x: int(x[0].split('_')[1]) if x[0].startswith('Columna_') else 999)
            for key, value in columnas_ordenadas:
                # Intentar usar el header si existe
                col_num = int(key.split('_')[1]) if key.startswith('Columna_') else 0
                header = headers.get(col_num - 1, key)
                f.write(f'{header}: {value}\n')
        else:
            f.write('No se encontraron datos específicos del cliente 1\n')
        
        # Mostrar también las filas alrededor del cliente 1 para contexto
        if cliente_1_fila_idx is not None:
            f.write('\n' + '='*80 + '\n')
            f.write('CONTEXTO: Filas alrededor del Cliente 1:\n')
            f.write('='*80 + '\n\n')
            inicio = max(0, cliente_1_fila_idx - 2)
            fin = min(len(todas_las_filas), cliente_1_fila_idx + 3)
            for idx in range(inicio, fin):
                f.write(f'\nFila {idx+1}:\n')
                fila = todas_las_filas[idx]
                valores = [(i+1, v) for i, v in enumerate(fila) if v is not None and str(v).strip() != '']
                for col_num, valor in valores[:30]:  # Mostrar primeras 30 columnas con datos
                    header = headers.get(col_num - 1, f'Columna_{col_num}')
                    f.write(f'  {header}: {valor}\n')
    
    # Guardar JSON
    with open('CLIENTE1_FINAL.json', 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2, default=str)
    
    print("Archivos creados: CLIENTE1_FINAL.txt y CLIENTE1_FINAL.json")
    print(f"Cliente 1 encontrado en fila: {cliente_1_fila_idx + 1 if cliente_1_fila_idx is not None else 'No encontrada'}")
