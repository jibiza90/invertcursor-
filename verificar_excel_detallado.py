from openpyxl import load_workbook

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)  # data_only=False para ver fórmulas

if 'Diario VIP' in wb.sheetnames:
    ws = wb['Diario VIP']
    
    print("=" * 80)
    print("VERIFICACION DETALLADA - CLIENTE 1 (Columna K)")
    print("=" * 80)
    
    # Verificar columna K desde fila 15
    print("\nColumna K (INCREMENTOS) desde fila 15:")
    suma_total = 0
    valores_encontrados = []
    
    for fila_idx in range(15, min(100, ws.max_row + 1)):
        celda = ws.cell(row=fila_idx, column=11)  # Columna K = 11
        
        # Verificar si tiene fórmula
        tiene_formula = celda.value is not None and str(celda.value).startswith('=')
        
        valor = celda.value
        if valor is not None:
            try:
                if tiene_formula:
                    # Calcular valor de la fórmula
                    valor_calculado = ws.cell(row=fila_idx, column=11).value
                    if isinstance(valor_calculado, (int, float)):
                        valores_encontrados.append((fila_idx, valor_calculado, True))
                        suma_total += valor_calculado
                        print(f"  Fila {fila_idx}: {valor_calculado:.2f} (FÓRMULA)")
                else:
                    valor_num = float(valor)
                    valores_encontrados.append((fila_idx, valor_num, False))
                    suma_total += valor_num
                    print(f"  Fila {fila_idx}: {valor_num:.2f}")
            except:
                pass
    
    print(f"\nSUMA TOTAL K15-K100: {suma_total:.2f}")
    print(f"Valores encontrados: {len(valores_encontrados)}")
    
    # Verificar qué celdas tienen fórmulas en las columnas del cliente 1
    print("\n" + "=" * 80)
    print("VERIFICANDO FORMULAS EN COLUMNAS DEL CLIENTE 1")
    print("=" * 80)
    
    columnas_cliente1 = {
        'K': 11,  # INCREMENTO
        'L': 12,  # DECREMENTO
        'M': 13,  # INICIAL
        'N': 14,  # FINAL
        'O': 15,  # BENEF. €
        'P': 16,  # BENEF. %
        'Q': 17,  # BENEF. € acum
        'R': 18   # BENEF. % acum
    }
    
    print("\nPrimeras 20 filas con datos (desde fila 15):")
    for fila_idx in range(15, min(35, ws.max_row + 1)):
        fecha = ws.cell(row=fila_idx, column=4).value  # Columna D
        if fecha is not None:
            print(f"\nFila {fila_idx} - Fecha: {fecha}")
            for col_nombre, col_num in columnas_cliente1.items():
                celda = ws.cell(row=fila_idx, column=col_num)
                valor = celda.value
                tiene_formula = valor is not None and str(valor).startswith('=')
                
                if valor is not None:
                    if tiene_formula:
                        print(f"  {col_nombre}: FÓRMULA ({valor[:50]})")
                    else:
                        print(f"  {col_nombre}: {valor}")
