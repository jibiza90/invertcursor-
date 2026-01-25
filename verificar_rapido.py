from openpyxl import load_workbook
import json

# Cargar datos
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws_formulas = wb_formulas['Diario VIP']

with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos_json = json.load(f)

hoja_json = datos_json['hojas']['Diario VIP']
cliente1 = next((c for c in hoja_json.get('clientes', []) if c.get('numero_cliente') == 1), None)

if not cliente1:
    print("ERROR: Cliente 1 no encontrado")
    exit(1)

datos_diarios = cliente1.get('datos_diarios', [])

# Mapeo columnas
cols = {11: 'incremento', 12: 'decremento', 13: 'base', 14: 'saldo_diario',
        15: 'beneficio_diario', 16: 'beneficio_diario_pct',
        17: 'beneficio_acumulado', 18: 'beneficio_acumulado_pct'}

# Verificar filas 15-30
resultados = []
errores = []

for fila in range(15, 31):
    fila_data = next((d for d in datos_diarios if d['fila'] == fila), None)
    if not fila_data:
        continue
    
    for col_num, campo in cols.items():
        celda = ws_formulas.cell(row=fila, column=col_num)
        tiene_formula_excel = celda.data_type == 'f' and celda.value
        formula_excel = str(celda.value) if tiene_formula_excel else None
        
        formula_json = fila_data.get('formulas', {}).get(campo)
        bloqueada_json = fila_data.get('bloqueadas', {}).get(campo, False)
        
        col_letra = chr(64 + col_num) if col_num <= 26 else '?'
        
        if tiene_formula_excel:
            if not formula_json:
                errores.append(f"Fila {fila}, {col_letra}{fila} ({campo}): Excel tiene formula pero JSON NO")
            elif formula_excel != formula_json:
                errores.append(f"Fila {fila}, {col_letra}{fila} ({campo}): Formulas diferentes")
            if not bloqueada_json:
                errores.append(f"Fila {fila}, {col_letra}{fila} ({campo}): Tiene formula pero NO bloqueada")
        elif campo in ['incremento', 'decremento'] and bloqueada_json:
            errores.append(f"Fila {fila}, {col_letra}{fila} ({campo}): NO debe estar bloqueada")

# Escribir resultados
with open('verificacion_celdas.txt', 'w', encoding='utf-8') as f:
    f.write("VERIFICACION CELDA A CELDA\n")
    f.write("=" * 80 + "\n\n")
    f.write(f"Errores encontrados: {len(errores)}\n\n")
    
    if errores:
        for error in errores:
            f.write(f"❌ {error}\n")
    else:
        f.write("✅ No se encontraron errores en las filas verificadas\n")

print(f"Verificacion completada. Errores: {len(errores)}")
print("Resultados guardados en: verificacion_celdas.txt")
