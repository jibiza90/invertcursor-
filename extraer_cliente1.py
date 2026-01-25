import traceback
import sys

try:
    from openpyxl import load_workbook
    
    print("Cargando archivo Excel...")
    wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
    
    print(f"Hojas disponibles: {wb.sheetnames}")
    
    # Buscar Diario_VIP
    sheet_name = None
    for sheet in wb.sheetnames:
        if 'Diario_VIP' in sheet or 'Diario VIP' in sheet or 'diario_vip' in sheet.lower():
            sheet_name = sheet
            break
    
    if not sheet_name:
        print("ERROR: No se encontró Diario_VIP")
        sys.exit(1)
    
    print(f"Leyendo hoja: {sheet_name}")
    ws = wb[sheet_name]
    
    # Leer headers
    headers = []
    first_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(cell.value) if cell.value is not None else '' for cell in first_row]
    
    print(f"Columnas encontradas: {len(headers)}")
    
    # Leer datos
    datos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fila = {}
        for i, value in enumerate(row):
            if i < len(headers):
                fila[headers[i]] = value
        if any(v is not None and v != '' for v in fila.values()):
            datos.append(fila)
    
    print(f"Filas de datos encontradas: {len(datos)}")
    
    # Buscar cliente 1
    cliente_1 = None
    for fila in datos:
        for key, value in fila.items():
            if value == 1 and key and ('cliente' in str(key).lower() or 'id' in str(key).lower() or 'codigo' in str(key).lower()):
                cliente_1 = fila
                print(f"Cliente 1 encontrado en columna: {key}")
                break
        if cliente_1:
            break
    
    if not cliente_1 and datos:
        cliente_1 = datos[0]
        print("Usando primera fila como cliente 1")
    
    # Escribir resultado
    with open('CLIENTE1.txt', 'w', encoding='utf-8') as f:
        f.write('='*80 + '\n')
        f.write(f'CLIENTE 1 - HOJA: {sheet_name}\n')
        f.write('='*80 + '\n\n')
        f.write(f'Total de filas: {len(datos)}\n')
        f.write(f'Total de columnas: {len(headers)}\n\n')
        f.write('Columnas disponibles:\n')
        for i, col in enumerate(headers, 1):
            f.write(f'  {i}. {col}\n')
        f.write('\n' + '='*80 + '\n')
        f.write('DATOS DEL CLIENTE 1:\n')
        f.write('='*80 + '\n\n')
        if cliente_1:
            for key, value in cliente_1.items():
                f.write(f'{key}: {value}\n')
        f.write('\n' + '='*80 + '\n')
        f.write('TODOS LOS DATOS DE LA HOJA:\n')
        f.write('='*80 + '\n\n')
        for i, fila in enumerate(datos, 1):
            f.write(f'\nFila {i}:\n')
            for key, value in fila.items():
                f.write(f'  {key}: {value}\n')
    
    print("Archivo CLIENTE1.txt creado exitosamente")
    
except Exception as e:
    error_msg = f"ERROR: {str(e)}\n\n{traceback.format_exc()}"
    print(error_msg)
    with open('ERROR.txt', 'w', encoding='utf-8') as f:
        f.write(error_msg)
    sys.exit(1)
