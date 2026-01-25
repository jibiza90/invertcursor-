"""
Extractor de datos v2 - Corregido para manejar dependencias entre clientes
"""
from openpyxl import load_workbook
import json
from datetime import datetime, date

# Cargar dos veces: una para fórmulas, otra para valores
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

def tiene_formula(ws_formulas, fila, columna):
    """Verifica si una celda tiene fórmula"""
    try:
        celda = ws_formulas.cell(row=fila, column=columna)
        if celda.data_type == 'f':
            return True
        if celda.value and str(celda.value).strip().startswith('='):
            return True
        return False
    except:
        return False

def obtener_formula(ws_formulas, fila, columna):
    """Obtiene la fórmula de una celda"""
    try:
        celda = ws_formulas.cell(row=fila, column=columna)
        if celda.value and str(celda.value).strip().startswith('='):
            return str(celda.value)
        return None
    except:
        return None

def procesar_hoja(nombre_hoja):
    """Procesa una hoja completa"""
    if nombre_hoja not in wb_formulas.sheetnames:
        return None
    
    ws_formulas = wb_formulas[nombre_hoja]
    ws_valores = wb_valores[nombre_hoja]
    
    resultado = {
        'nombre': nombre_hoja,
        'datos_generales': [],
        'datos_diarios_generales': [],
        'clientes': []
    }
    
    print(f"\n{'='*70}")
    print(f"Procesando: {nombre_hoja}")
    print('='*70)
    
    # 1. DATOS GENERALES (filas 3-6)
    print("\n1. Datos generales (filas 3-6)...")
    for fila_idx in [3, 4, 5, 6]:
        fila_data = {
            'fila': fila_idx,
            'fecha': ws_valores.cell(fila_idx, 4).value,
            'imp_inicial': ws_valores.cell(fila_idx, 5).value,
            'imp_final': ws_valores.cell(fila_idx, 6).value,
            'benef_euro': ws_valores.cell(fila_idx, 7).value,
            'benef_porcentaje': ws_valores.cell(fila_idx, 8).value,
            'benef_euro_acum': ws_valores.cell(fila_idx, 9).value,
            'benef_porcentaje_acum': ws_valores.cell(fila_idx, 10).value,
            'bloqueadas': {},
            'formulas': {}
        }
        
        # Verificar fórmulas
        cols_map = [('imp_inicial', 5), ('imp_final', 6), ('benef_euro', 7), 
                    ('benef_porcentaje', 8), ('benef_euro_acum', 9), ('benef_porcentaje_acum', 10)]
        for campo, col in cols_map:
            if tiene_formula(ws_formulas, fila_idx, col):
                fila_data['bloqueadas'][campo] = True
                formula = obtener_formula(ws_formulas, fila_idx, col)
                if formula:
                    fila_data['formulas'][campo] = formula
            else:
                fila_data['bloqueadas'][campo] = False
        
        resultado['datos_generales'].append(fila_data)
    print(f"   {len(resultado['datos_generales'])} filas")
    
    # 2. DATOS DIARIOS GENERALES (desde fila 15)
    print("\n2. Datos diarios generales...")
    for fila_idx in range(15, ws_valores.max_row + 1):
        fecha = ws_valores.cell(fila_idx, 4).value
        if fecha is None:
            continue
            
        fila_data = {
            'fila': fila_idx,
            'fecha': fecha,
            'imp_inicial': ws_valores.cell(fila_idx, 5).value,
            'imp_final': ws_valores.cell(fila_idx, 6).value,
            'benef_euro': ws_valores.cell(fila_idx, 7).value,
            'benef_porcentaje': ws_valores.cell(fila_idx, 8).value,
            'benef_euro_acum': ws_valores.cell(fila_idx, 9).value,
            'benef_porcentaje_acum': ws_valores.cell(fila_idx, 10).value,
            'bloqueadas': {},
            'formulas': {}
        }
        
        cols_map = [('imp_inicial', 5), ('imp_final', 6), ('benef_euro', 7), 
                    ('benef_porcentaje', 8), ('benef_euro_acum', 9), ('benef_porcentaje_acum', 10)]
        for campo, col in cols_map:
            if tiene_formula(ws_formulas, fila_idx, col):
                fila_data['bloqueadas'][campo] = True
                formula = obtener_formula(ws_formulas, fila_idx, col)
                if formula:
                    fila_data['formulas'][campo] = formula
            else:
                fila_data['bloqueadas'][campo] = False
        
        resultado['datos_diarios_generales'].append(fila_data)
    
    print(f"   {len(resultado['datos_diarios_generales'])} registros")
    
    # 3. IDENTIFICAR CLIENTES
    print("\n3. Identificando clientes...")
    clientes_info = []
    
    # Buscar números de cliente en fila 13
    for col in range(11, ws_valores.max_column + 1):
        valor = ws_valores.cell(13, col).value
        if valor is not None:
            try:
                num = int(valor)
                if num > 0:
                    clientes_info.append({
                        'numero': num,
                        'col_inicio': col,
                        'cols': {
                            'incremento': col,      # K, S, AA...
                            'decremento': col + 1,  # L, T, AB...
                            'base': col + 2,        # M, U, AC...
                            'saldo_diario': col + 3, # N, V, AD...
                            'beneficio_diario': col + 4, # O, W, AE...
                            'beneficio_diario_pct': col + 5, # P, X, AF...
                            'beneficio_acumulado': col + 6, # Q, Y, AG...
                            'beneficio_acumulado_pct': col + 7  # R, Z, AH...
                        }
                    })
            except:
                pass
    
    print(f"   {len(clientes_info)} clientes encontrados")
    
    # 4. EXTRAER DATOS POR CLIENTE
    print("\n4. Extrayendo datos por cliente...")
    
    for cliente_info in clientes_info:
        cliente_data = {
            'numero_cliente': cliente_info['numero'],
            'columna_inicio': cliente_info['col_inicio'],
            'incrementos_total': 0,
            'decrementos_total': 0,
            'saldo_actual': None,
            'datos_diarios': []
        }
        
        cols = cliente_info['cols']
        incrementos = []
        decrementos = []
        saldos = []
        
        for fila_idx in range(15, ws_valores.max_row + 1):
            fecha = ws_valores.cell(fila_idx, 4).value
            if fecha is None:
                continue
            
            # Leer valores
            dato = {
                'fila': fila_idx,
                'fecha': fecha,
                'incremento': ws_valores.cell(fila_idx, cols['incremento']).value,
                'decremento': ws_valores.cell(fila_idx, cols['decremento']).value,
                'base': ws_valores.cell(fila_idx, cols['base']).value,
                'saldo_diario': ws_valores.cell(fila_idx, cols['saldo_diario']).value,
                'beneficio_diario': ws_valores.cell(fila_idx, cols['beneficio_diario']).value,
                'beneficio_diario_pct': ws_valores.cell(fila_idx, cols['beneficio_diario_pct']).value,
                'beneficio_acumulado': ws_valores.cell(fila_idx, cols['beneficio_acumulado']).value,
                'beneficio_acumulado_pct': ws_valores.cell(fila_idx, cols['beneficio_acumulado_pct']).value,
                'bloqueadas': {},
                'formulas': {}
            }
            
            # ORDEN CORRECTO según dependencias:
            # 1. base, 2. beneficio_diario_pct, 3. beneficio_diario, 4. saldo_diario, 5. beneficio_acumulado, 6. beneficio_acumulado_pct
            campos_orden = [
                ('base', cols['base']),
                ('beneficio_diario_pct', cols['beneficio_diario_pct']),
                ('beneficio_diario', cols['beneficio_diario']),
                ('saldo_diario', cols['saldo_diario']),
                ('beneficio_acumulado', cols['beneficio_acumulado']),
                ('beneficio_acumulado_pct', cols['beneficio_acumulado_pct'])
            ]
            
            for campo, col in campos_orden:
                if tiene_formula(ws_formulas, fila_idx, col):
                    dato['bloqueadas'][campo] = True
                    formula = obtener_formula(ws_formulas, fila_idx, col)
                    if formula:
                        dato['formulas'][campo] = formula
                else:
                    dato['bloqueadas'][campo] = False
            
            cliente_data['datos_diarios'].append(dato)
            
            # Acumular para totales (solo 2026)
            if fecha:
                try:
                    fecha_dt = fecha if isinstance(fecha, (datetime, date)) else None
                    if fecha_dt and fecha_dt.year == 2026:
                        if dato['incremento']:
                            incrementos.append(float(dato['incremento']))
                        if dato['decremento']:
                            decrementos.append(float(dato['decremento']))
                except:
                    pass
            
            if dato['saldo_diario'] and dato['saldo_diario'] != 0:
                saldos.append((fila_idx, dato['saldo_diario']))
        
        cliente_data['incrementos_total'] = sum(incrementos)
        cliente_data['decrementos_total'] = sum(decrementos)
        if saldos:
            cliente_data['saldo_actual'] = saldos[-1][1]
            cliente_data['saldo_actual_fila'] = saldos[-1][0]
        
        resultado['clientes'].append(cliente_data)
        
        if cliente_info['numero'] <= 3:
            print(f"   Cliente {cliente_info['numero']}: {len(cliente_data['datos_diarios'])} registros, saldo={cliente_data['saldo_actual']}")
    
    # Ordenar clientes
    resultado['clientes'].sort(key=lambda x: x['numero_cliente'])
    
    return resultado

# Procesar hojas
resultado_final = {
    'fecha_extraccion': datetime.now().isoformat(),
    'hojas': {}
}

for hoja in ['Diario STD', 'Diario VIP']:
    if hoja in wb_formulas.sheetnames:
        resultado_final['hojas'][hoja] = procesar_hoja(hoja)

# Guardar
with open('datos_completos.json', 'w', encoding='utf-8') as f:
    json.dump(resultado_final, f, ensure_ascii=False, indent=2, default=str)

print(f"\n{'='*70}")
print("Datos guardados en: datos_completos.json")
print('='*70)

# Verificación
print("\nVerificación de fórmulas Cliente 1 vs Cliente 2:")
vip = resultado_final['hojas']['Diario VIP']
if len(vip['clientes']) >= 2:
    c1 = vip['clientes'][0]
    c2 = vip['clientes'][1]
    f17_c1 = [x for x in c1['datos_diarios'] if x['fila'] == 17][0]
    f17_c2 = [x for x in c2['datos_diarios'] if x['fila'] == 17][0]
    print(f"  Cliente 1 fila 17 - base formula: {f17_c1['formulas'].get('base', 'N/A')}")
    print(f"  Cliente 2 fila 17 - base formula: {f17_c2['formulas'].get('base', 'N/A')}")
