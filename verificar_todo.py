from openpyxl import load_workbook
import json
import traceback
from datetime import datetime

# Abrir archivo de salida desde el inicio
output_file = open('reporte_verificacion_100porciento.txt', 'w', encoding='utf-8')

def escribir(mensaje):
    print(mensaje)
    output_file.write(mensaje + '\n')
    output_file.flush()

escribir("=" * 80)
escribir("VERIFICACION COMPLETA: 100% DE LAS CELDAS")
escribir(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
escribir("=" * 80)

try:
    escribir("\n1. Cargando Excel...")
    wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
    ws_formulas = wb_formulas['Diario VIP']
    escribir(f"   ✓ Excel cargado - Max fila: {ws_formulas.max_row}")
    
    escribir("\n2. Cargando JSON...")
    with open('datos_completos.json', 'r', encoding='utf-8') as f:
        datos_json = json.load(f)
    
    hoja_json = datos_json['hojas']['Diario VIP']
    clientes_json = hoja_json.get('clientes', [])
    datos_diarios_json = hoja_json.get('datos_diarios_generales', [])
    datos_generales_json = hoja_json.get('datos_generales', [])
    
    escribir(f"   ✓ JSON cargado - {len(clientes_json)} clientes, {len(datos_diarios_json)} filas diarias")
    
    # Mapeos
    col_general = {5: 'imp_inicial', 6: 'imp_final', 7: 'benef_euro', 
                   8: 'benef_porcentaje', 9: 'benef_euro_acum', 10: 'benef_porcentaje_acum'}
    offset_cliente = {0: 'incremento', 1: 'decremento', 2: 'base', 3: 'saldo_diario',
                      4: 'beneficio_diario', 5: 'beneficio_diario_pct',
                      6: 'beneficio_acumulado', 7: 'beneficio_acumulado_pct'}
    
    def col_a_letra(n):
        result = ""
        while n > 0:
            n -= 1
            result = chr(65 + (n % 26)) + result
            n //= 26
        return result
    
    errores = []
    advertencias = []
    total_celdas = 0
    
    # 3. Verificar datos generales
    escribir("\n3. Verificando DATOS GENERALES (Filas 3-6)...")
    for fila in range(3, 7):
        fila_json = next((d for d in datos_generales_json if d.get('fila') == fila), None)
        if not fila_json:
            errores.append(f"Fila {fila}: No encontrada en JSON")
            continue
        
        for col, campo in col_general.items():
            total_celdas += 1
            celda = ws_formulas.cell(row=fila, column=col)
            tiene_formula = celda.data_type == 'f' and celda.value is not None
            
            if tiene_formula:
                formula_excel = str(celda.value)
                formula_json = fila_json.get('formulas', {}).get(campo)
                bloqueada = fila_json.get('bloqueadas', {}).get(campo, False)
                
                if not formula_json:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Excel tiene formula pero JSON NO")
                elif formula_excel != formula_json:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Formulas diferentes")
                if not bloqueada:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Tiene formula pero NO bloqueada")
    
    escribir(f"   ✓ Verificadas {total_celdas} celdas")
    
    # 4. Verificar datos diarios generales
    escribir("\n4. Verificando DATOS DIARIOS GENERALES...")
    celdas_diarias = 0
    for fila_json in datos_diarios_json:
        fila = fila_json.get('fila')
        if fila < 15:
            continue
        
        for col, campo in col_general.items():
            total_celdas += 1
            celdas_diarias += 1
            celda = ws_formulas.cell(row=fila, column=col)
            tiene_formula = celda.data_type == 'f' and celda.value is not None
            
            if tiene_formula:
                formula_excel = str(celda.value)
                formula_json = fila_json.get('formulas', {}).get(campo)
                bloqueada = fila_json.get('bloqueadas', {}).get(campo, False)
                
                if not formula_json:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Excel tiene formula pero JSON NO")
                elif formula_excel != formula_json:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Formulas diferentes")
                if not bloqueada:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Tiene formula pero NO bloqueada")
    
    escribir(f"   ✓ Verificadas {celdas_diarias} celdas")
    
    # 5. Verificar clientes
    escribir("\n5. Verificando CLIENTES...")
    celdas_clientes = 0
    
    # Identificar columnas de clientes desde Excel
    clientes_cols = {}
    for col in range(11, min(ws_formulas.max_column + 1, 200)):
        num_cliente = ws_formulas.cell(row=13, column=col).value
        if num_cliente and isinstance(num_cliente, (int, float)) and int(num_cliente) > 0:
            num = int(num_cliente)
            if num not in clientes_cols:
                clientes_cols[num] = []
            clientes_cols[num].append(col)
    
    escribir(f"   Encontrados {len(clientes_cols)} clientes en Excel")
    
    for num_cliente, columnas in sorted(clientes_cols.items()):
        cliente_json = next((c for c in clientes_json if c.get('numero_cliente') == num_cliente), None)
        if not cliente_json:
            advertencias.append(f"Cliente {num_cliente}: No encontrado en JSON")
            continue
        
        datos_cliente = cliente_json.get('datos_diarios', [])
        col_inicio = min(columnas)
        
        escribir(f"   Verificando Cliente {num_cliente}...")
        
        for fila_data in datos_cliente:
            fila = fila_data.get('fila')
            if fila < 15:
                continue
            
            for offset, campo in offset_cliente.items():
                col = col_inicio + offset
                if col > ws_formulas.max_column:
                    break
                
                total_celdas += 1
                celdas_clientes += 1
                
                celda = ws_formulas.cell(row=fila, column=col)
                tiene_formula = celda.data_type == 'f' and celda.value is not None
                
                if tiene_formula:
                    formula_excel = str(celda.value)
                    formula_json = fila_data.get('formulas', {}).get(campo)
                    bloqueada = fila_data.get('bloqueadas', {}).get(campo, False)
                    
                    if not formula_json:
                        errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Excel tiene formula pero JSON NO")
                    elif formula_excel != formula_json:
                        errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Formulas diferentes")
                    if not bloqueada:
                        errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Tiene formula pero NO bloqueada")
                else:
                    if campo in ['incremento', 'decremento']:
                        bloqueada = fila_data.get('bloqueadas', {}).get(campo, False)
                        if bloqueada:
                            errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): NO debe estar bloqueada")
    
    escribir(f"   ✓ Verificadas {celdas_clientes} celdas de clientes")
    
    # Resumen final
    escribir("\n" + "=" * 80)
    escribir("RESUMEN FINAL")
    escribir("=" * 80)
    escribir(f"Total de celdas verificadas: {total_celdas:,}")
    escribir(f"  - Datos generales: {4 * 6}")
    escribir(f"  - Datos diarios generales: {celdas_diarias:,}")
    escribir(f"  - Celdas de clientes: {celdas_clientes:,}")
    escribir(f"\nErrores críticos: {len(errores)}")
    escribir(f"Advertencias: {len(advertencias)}")
    
    if errores:
        escribir("\n" + "=" * 80)
        escribir("ERRORES CRÍTICOS")
        escribir("=" * 80)
        for i, error in enumerate(errores, 1):
            escribir(f"{i}. {error}")
    
    if advertencias:
        escribir("\n" + "=" * 80)
        escribir("ADVERTENCIAS")
        escribir("=" * 80)
        for i, adv in enumerate(advertencias[:100], 1):
            escribir(f"{i}. {adv}")
        if len(advertencias) > 100:
            escribir(f"\n... y {len(advertencias) - 100} advertencias más")
    
    escribir("\n" + "=" * 80)
    escribir("VERIFICACION COMPLETADA")
    escribir("=" * 80)
    
except Exception as e:
    escribir(f"\nERROR CRÍTICO: {e}")
    escribir(traceback.format_exc())

finally:
    output_file.close()
    print("\nReporte guardado en: reporte_verificacion_100porciento.txt")
