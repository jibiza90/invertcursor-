from openpyxl import load_workbook
import json
import re

print("=" * 80)
print("VERIFICACION CELDA A CELDA: EXCEL vs JSON")
print("=" * 80)

try:
    # Cargar Excel
    print("Cargando Excel...")
    wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
    wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
    ws_formulas = wb_formulas['Diario VIP']
    ws_valores = wb_valores['Diario VIP']
    print("✓ Excel cargado")
except Exception as e:
    print(f"ERROR al cargar Excel: {e}")
    exit(1)

try:
    # Cargar JSON
    print("Cargando JSON...")
    with open('datos_completos.json', 'r', encoding='utf-8') as f:
        datos_json = json.load(f)
    
    if 'Diario VIP' not in datos_json.get('hojas', {}):
        print("ERROR: No se encuentra 'Diario VIP' en el JSON")
        print(f"Hojas disponibles: {list(datos_json.get('hojas', {}).keys())}")
        exit(1)
    
    hoja_json = datos_json['hojas']['Diario VIP']
    clientes_json = hoja_json.get('clientes', [])
    print(f"✓ JSON cargado - {len(clientes_json)} clientes encontrados")
except Exception as e:
    print(f"ERROR al cargar JSON: {e}")
    import traceback
    traceback.print_exc()
    exit(1)

# Mapeo de columnas a campos
columna_a_campo_cliente = {
    11: 'incremento',      # K
    12: 'decremento',      # L
    13: 'base',            # M
    14: 'saldo_diario',    # N
    15: 'beneficio_diario', # O
    16: 'beneficio_diario_pct', # P
    17: 'beneficio_acumulado',  # Q
    18: 'beneficio_acumulado_pct' # R
}

columna_a_campo_general = {
    5: 'imp_inicial',      # E
    6: 'imp_final',        # F
    7: 'benef_euro',       # G
    8: 'benef_porcentaje', # H
    9: 'benef_euro_acum',  # I
    10: 'benef_porcentaje_acum' # J
}

errores = []
advertencias = []

# Función para convertir referencia Excel a número de columna
def letra_a_columna(letra):
    col = 0
    for ch in letra:
        col = col * 26 + (ord(ch) - 64)
    return col

# Verificar datos generales (filas 3-6 y 15+)
print("\n1. VERIFICANDO DATOS GENERALES (Columnas E-J)...")
datos_generales_json = hoja_json.get('datos_generales', [])
datos_diarios_json = hoja_json.get('datos_diarios_generales', [])

# Verificar filas 3-6
for fila_excel in range(3, 7):
    fila_data_json = next((d for d in datos_generales_json if d['fila'] == fila_excel), None)
    
    if not fila_data_json:
        advertencias.append(f"Fila {fila_excel}: No encontrada en JSON")
        continue
    
    for col_excel, campo in columna_a_campo_general.items():
        celda_excel = ws_formulas.cell(row=fila_excel, column=col_excel)
        tiene_formula_excel = celda_excel.data_type == 'f' and celda_excel.value
        
        # Verificar en JSON
        tiene_formula_json = fila_data_json.get('formulas', {}).get(campo, None)
        esta_bloqueada_json = fila_data_json.get('bloqueadas', {}).get(campo, False)
        
        if tiene_formula_excel:
            formula_excel = str(celda_excel.value)
            if tiene_formula_json:
                formula_json = tiene_formula_json
                if formula_excel != formula_json:
                    errores.append(f"Fila {fila_excel}, Col {col_excel} ({campo}): Fórmula diferente")
                    errores.append(f"  Excel: {formula_excel[:100]}")
                    errores.append(f"  JSON:  {formula_json[:100]}")
            else:
                errores.append(f"Fila {fila_excel}, Col {col_excel} ({campo}): Tiene fórmula en Excel pero NO en JSON")
            
            # Debe estar bloqueada
            if not esta_bloqueada_json:
                errores.append(f"Fila {fila_excel}, Col {col_excel} ({campo}): Tiene fórmula pero NO está bloqueada en JSON")
        else:
            # No tiene fórmula, no debe estar bloqueada (a menos que sea campo especial)
            if esta_bloqueada_json and campo not in ['imp_inicial', 'imp_final']:
                advertencias.append(f"Fila {fila_excel}, Col {col_excel} ({campo}): Sin fórmula pero bloqueada en JSON")

# Verificar datos diarios generales (filas 15+)
print("\n2. VERIFICANDO DATOS DIARIOS GENERALES (Columnas E-J, Filas 15+)...")
for fila_excel in range(15, min(100, ws_valores.max_row + 1)):  # Revisar primeras 85 filas
    fecha_excel = ws_valores.cell(row=fila_excel, column=4).value
    if fecha_excel is None:
        continue
    
    fila_data_json = next((d for d in datos_diarios_json if d['fila'] == fila_excel), None)
    
    if not fila_data_json:
        advertencias.append(f"Fila {fila_excel}: No encontrada en JSON")
        continue
    
    for col_excel, campo in columna_a_campo_general.items():
        celda_excel = ws_formulas.cell(row=fila_excel, column=col_excel)
        tiene_formula_excel = celda_excel.data_type == 'f' and celda_excel.value
        
        tiene_formula_json = fila_data_json.get('formulas', {}).get(campo, None)
        esta_bloqueada_json = fila_data_json.get('bloqueadas', {}).get(campo, False)
        
        if tiene_formula_excel:
            formula_excel = str(celda_excel.value)
            if tiene_formula_json:
                formula_json = tiene_formula_json
                if formula_excel != formula_json:
                    errores.append(f"Fila {fila_excel}, Col {col_excel} ({campo}): Fórmula diferente")
                    errores.append(f"  Excel: {formula_excel[:100]}")
                    errores.append(f"  JSON:  {formula_json[:100]}")
            else:
                errores.append(f"Fila {fila_excel}, Col {col_excel} ({campo}): Tiene fórmula en Excel pero NO en JSON")
            
            if not esta_bloqueada_json:
                errores.append(f"Fila {fila_excel}, Col {col_excel} ({campo}): Tiene fórmula pero NO está bloqueada en JSON")
        else:
            # Verificar si debería estar bloqueada por otras razones
            valor_excel = ws_valores.cell(row=fila_excel, column=col_excel).value
            if esta_bloqueada_json and not tiene_formula_excel and campo in ['imp_inicial', 'imp_final']:
                # imp_inicial e imp_final pueden estar bloqueadas sin fórmula
                pass
            elif esta_bloqueada_json and not tiene_formula_excel:
                advertencias.append(f"Fila {fila_excel}, Col {col_excel} ({campo}): Sin fórmula pero bloqueada en JSON")

# Verificar clientes (Cliente 1 como ejemplo, columnas K-R)
print("\n3. VERIFICANDO CLIENTE 1 (Columnas K-R)...")
cliente1_json = next((c for c in clientes_json if c.get('numero_cliente') == 1), None)

if cliente1_json:
    datos_diarios_cliente = cliente1_json.get('datos_diarios', [])
    
    for fila_excel in range(15, min(100, ws_valores.max_row + 1)):  # Revisar primeras 85 filas
        fecha_excel = ws_valores.cell(row=fila_excel, column=4).value
        if fecha_excel is None:
            continue
        
        fila_data_json = next((d for d in datos_diarios_cliente if d['fila'] == fila_excel), None)
        
        if not fila_data_json:
            advertencias.append(f"Cliente 1, Fila {fila_excel}: No encontrada en JSON")
            continue
        
        for col_excel, campo in columna_a_campo_cliente.items():
            celda_excel = ws_formulas.cell(row=fila_excel, column=col_excel)
            tiene_formula_excel = celda_excel.data_type == 'f' and celda_excel.value
            
            tiene_formula_json = fila_data_json.get('formulas', {}).get(campo, None)
            esta_bloqueada_json = fila_data_json.get('bloqueadas', {}).get(campo, False)
            
            if tiene_formula_excel:
                formula_excel = str(celda_excel.value)
                if tiene_formula_json:
                    formula_json = tiene_formula_json
                    if formula_excel != formula_json:
                        errores.append(f"Cliente 1, Fila {fila_excel}, Col {col_excel} ({campo}): Fórmula diferente")
                        errores.append(f"  Excel: {formula_excel[:100]}")
                        errores.append(f"  JSON:  {formula_json[:100]}")
                else:
                    errores.append(f"Cliente 1, Fila {fila_excel}, Col {col_excel} ({campo}): Tiene fórmula en Excel pero NO en JSON")
                
                if not esta_bloqueada_json:
                    errores.append(f"Cliente 1, Fila {fila_excel}, Col {col_excel} ({campo}): Tiene fórmula pero NO está bloqueada en JSON")
            else:
                # Incremento y decremento NO deben estar bloqueados (son editables)
                if esta_bloqueada_json and campo in ['incremento', 'decremento']:
                    errores.append(f"Cliente 1, Fila {fila_excel}, Col {col_excel} ({campo}): NO debe estar bloqueada (es editable)")
                # Base, saldo_diario, etc. pueden estar bloqueadas si tienen valor calculado
                elif esta_bloqueada_json and campo not in ['incremento', 'decremento']:
                    valor_excel = ws_valores.cell(row=fila_excel, column=col_excel).value
                    if valor_excel is None or valor_excel == 0:
                        advertencias.append(f"Cliente 1, Fila {fila_excel}, Col {col_excel} ({campo}): Bloqueada pero sin valor en Excel")

# Mostrar resultados
print("\n" + "=" * 80)
print("RESUMEN DE VERIFICACION")
print("=" * 80)

if errores:
    print(f"\n❌ ERRORES ENCONTRADOS: {len(errores)}")
    for i, error in enumerate(errores[:50], 1):  # Mostrar primeros 50
        print(f"{i}. {error}")
    if len(errores) > 50:
        print(f"... y {len(errores) - 50} errores más")
else:
    print("\n✅ No se encontraron errores críticos")

if advertencias:
    print(f"\n⚠️  ADVERTENCIAS: {len(advertencias)}")
    for i, adv in enumerate(advertencias[:30], 1):  # Mostrar primeras 30
        print(f"{i}. {adv}")
    if len(advertencias) > 30:
        print(f"... y {len(advertencias) - 30} advertencias más")

print("\n" + "=" * 80)
print("VERIFICACION COMPLETADA")
print("=" * 80)

# Guardar reporte detallado
with open('reporte_verificacion_celdas.txt', 'w', encoding='utf-8') as f:
    f.write("REPORTE DE VERIFICACION CELDA A CELDA\n")
    f.write("=" * 80 + "\n\n")
    f.write(f"ERRORES: {len(errores)}\n")
    for error in errores:
        f.write(f"  - {error}\n")
    f.write(f"\nADVERTENCIAS: {len(advertencias)}\n")
    for adv in advertencias:
        f.write(f"  - {adv}\n")

print(f"\n📄 Reporte guardado en: reporte_verificacion_celdas.txt")
