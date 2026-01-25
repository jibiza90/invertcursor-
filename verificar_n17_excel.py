#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificar qué es realmente N17 en Excel y cómo debería calcularse
"""

import openpyxl

wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)

ws_std = None
for sheet_name in wb.sheetnames:
    if 'STD' in sheet_name.upper() and 'VIP' not in sheet_name.upper():
        ws_std = wb[sheet_name]
        print(f"Hoja encontrada: {sheet_name}")
        break

print("="*80)
print("VERIFICACIÓN DE COLUMNAS EN FILA 17")
print("="*80)

fila = 17

# Ver todas las columnas de la fila 17
print(f"\nFila {fila}:")

# Columnas del Cliente 1 (K=11 hasta R=18)
print("\nCliente 1 (columnas K-R, 11-18):")
for col_num in range(11, 19):
    col_letra = openpyxl.utils.get_column_letter(col_num)
    celda = ws_std.cell(row=fila, column=col_num)
    if celda.data_type == 'f' and celda.value:
        print(f"  {col_letra}{fila}: {celda.value}")
    else:
        valor = celda.value
        print(f"  {col_letra}{fila}: {valor}")

# Columnas del Cliente 2 (S=19 hasta Z=26)
print("\nCliente 2 (columnas S-Z, 19-26):")
for col_num in range(19, 27):
    col_letra = openpyxl.utils.get_column_letter(col_num)
    celda = ws_std.cell(row=fila, column=col_num)
    if celda.data_type == 'f' and celda.value:
        print(f"  {col_letra}{fila}: {celda.value}")
    else:
        valor = celda.value
        print(f"  {col_letra}{fila}: {valor}")

# Verificar específicamente N17 (columna 14)
print(f"\nN17 (columna 14) específicamente:")
celda_n17 = ws_std.cell(row=fila, column=14)
if celda_n17.data_type == 'f' and celda_n17.value:
    print(f"  Fórmula: {celda_n17.value}")
else:
    print(f"  Valor: {celda_n17.value}")

# Verificar la fórmula de base del Cliente 2 (U17, columna 21)
print(f"\nU17 (base Cliente 2, columna 21):")
celda_u17 = ws_std.cell(row=fila, column=21)
if celda_u17.data_type == 'f' and celda_u17.value:
    formula_u17 = str(celda_u17.value)
    print(f"  Fórmula: {formula_u17}")
    
    # Extraer referencias
    import re
    referencias = re.findall(r'([A-Z]+\d+)', formula_u17)
    print(f"  Referencias encontradas: {referencias}")
    
    # Verificar cada referencia
    for ref in referencias:
        col_letra = re.match(r'([A-Z]+)', ref).group(1)
        fila_num = int(re.match(r'\d+', ref).group())
        col_num = openpyxl.utils.column_index_from_string(col_letra)
        
        celda_ref = ws_std.cell(row=fila_num, column=col_num)
        if celda_ref.data_type == 'f' and celda_ref.value:
            print(f"    {ref} (col {col_num}): fórmula = {celda_ref.value}")
        else:
            print(f"    {ref} (col {col_num}): valor = {celda_ref.value}")

wb.close()
