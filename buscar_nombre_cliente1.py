from openpyxl import load_workbook

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

# Buscar Diario VIP
sheet_name = None
for sheet in wb.sheetnames:
    if 'Diario VIP' in sheet or 'Diario_VIP' in sheet or 'diario_vip' in sheet.lower():
        sheet_name = sheet
        break

if sheet_name:
    ws = wb[sheet_name]
    
    # Leer todas las filas
    todas_las_filas = []
    for row in ws.iter_rows(values_only=True):
        todas_las_filas.append(list(row))
    
    # La fila 15 (índice 14) es el cliente 1 según el análisis anterior
    fila_cliente_1 = 14  # índice 0-based
    
    # Buscar en un rango más amplio alrededor del cliente 1
    inicio_busqueda = max(0, fila_cliente_1 - 10)
    fin_busqueda = min(len(todas_las_filas), fila_cliente_1 + 50)
    
    # Buscar nombre del cliente (probablemente está en alguna fila anterior o en otra columna)
    nombre_cliente = None
    
    # Buscar en las filas anteriores al cliente 1
    for idx in range(fila_cliente_1 - 1, max(0, fila_cliente_1 - 20), -1):
        fila = todas_las_filas[idx]
        # Buscar texto que parezca un nombre (más de 3 caracteres, no numérico)
        for valor in fila:
            if valor is not None:
                valor_str = str(valor).strip()
                # Si es texto largo y no parece ser un header o número
                if len(valor_str) > 3 and not valor_str.replace('.', '').replace(',', '').isdigit():
                    # Excluir palabras comunes de headers
                    if valor_str.upper() not in ['NOMBRE CLIENTE', 'SALDO DIARIO', 'SALDO ACUMULADO', 
                                                 'MES', 'FECHA', 'IMP. INICIAL', 'IMP. FINAL', 
                                                 'BENEF. €', 'BENEF. %', 'INCREMENTO', 'DECREMENTO',
                                                 'INICIAL', 'FINAL', 'PERIODO', 'SELECCIÓN', 'ANUAL']:
                        # Podría ser el nombre
                        nombre_cliente = valor_str
                        break
        if nombre_cliente:
            break
    
    # Leer la fila completa del cliente 1 con contexto de headers
    fila_cliente = todas_las_filas[fila_cliente_1]
    
    # Buscar headers en la fila 14 (índice 13)
    headers = {}
    if fila_cliente_1 > 0:
        fila_headers = todas_las_filas[fila_cliente_1 - 1]
        for col_idx, valor in enumerate(fila_headers):
            if valor is not None and str(valor).strip() != '':
                headers[col_idx] = str(valor).strip()
    
    # Crear diccionario con datos del cliente usando headers cuando sea posible
    datos_cliente_1 = {}
    for col_idx, valor in enumerate(fila_cliente):
        if valor is not None:
            header = headers.get(col_idx, f'Columna_{col_idx+1}')
            datos_cliente_1[header] = valor
    
    # Escribir resultado completo
    with open('CLIENTE1_NOMBRE.txt', 'w', encoding='utf-8') as f:
        f.write('='*80 + '\n')
        f.write('INFORMACIÓN COMPLETA DEL CLIENTE 1\n')
        f.write('='*80 + '\n\n')
        
        f.write(f'Hoja: {sheet_name}\n')
        f.write(f'Fila del cliente 1: {fila_cliente_1 + 1}\n')
        if nombre_cliente:
            f.write(f'Nombre del cliente (posible): {nombre_cliente}\n')
        f.write('\n')
        
        f.write('='*80 + '\n')
        f.write('DATOS DEL CLIENTE 1:\n')
        f.write('='*80 + '\n\n')
        
        # Ordenar por número de columna
        items_ordenados = []
        for key, value in datos_cliente_1.items():
            if key.startswith('Columna_'):
                try:
                    num = int(key.split('_')[1])
                    items_ordenados.append((num, key, value))
                except:
                    items_ordenados.append((999, key, value))
            else:
                items_ordenados.append((0, key, value))
        
        items_ordenados.sort(key=lambda x: x[0])
        
        for num, key, value in items_ordenados:
            f.write(f'{key}: {value}\n')
        
        # Mostrar también filas anteriores para contexto
        f.write('\n' + '='*80 + '\n')
        f.write('FILAS ANTERIORES (para contexto):\n')
        f.write('='*80 + '\n\n')
        
        for idx in range(max(0, fila_cliente_1 - 5), fila_cliente_1):
            f.write(f'\nFila {idx+1}:\n')
            fila = todas_las_filas[idx]
            valores = [(i, v) for i, v in enumerate(fila) if v is not None and str(v).strip() != '']
            for col_idx, valor in valores[:40]:
                header = headers.get(col_idx, f'Columna_{col_idx+1}')
                f.write(f'  {header}: {valor}\n')
    
    print("Archivo CLIENTE1_NOMBRE.txt creado")
    if nombre_cliente:
        print(f"Posible nombre del cliente: {nombre_cliente}")
