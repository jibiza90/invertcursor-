from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("VERIFICACION FORMULAS TOTALES (FILAS 3-6)")
print("=" * 80)

filas_totales = [3, 4, 5, 6]
columnas_verificar = {
    5: 'E - Imp. Inicial',
    6: 'F - Imp. Final',
    7: 'G - Benef. €',
    8: 'H - Benef. %',
    9: 'I - Benef. € Acum',
    10: 'J - Benef. % Acum'
}

for fila in filas_totales:
    print(f"\nFila {fila}:")
    for col_idx, nombre in columnas_verificar.items():
        celda = ws.cell(row=fila, column=col_idx)
        tiene_formula = celda.data_type == 'f'
        valor = celda.value
        
        if tiene_formula and valor:
            print(f"  {nombre}: {str(valor)[:100]}")
        elif valor is not None:
            print(f"  {nombre}: valor={valor} (sin formula)")
