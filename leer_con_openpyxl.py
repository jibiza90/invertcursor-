from openpyxl import load_workbook
import json

try:
    wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
    
    print("Hojas disponibles:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
    
    # Buscar Diario_VIP
    sheet_name = None
    for sheet in wb.sheetnames:
        if 'Diario_VIP' in sheet or 'Diario VIP' in sheet or 'diario_vip' in sheet.lower():
            sheet_name = sheet
            break
    
    if not sheet_name:
        print("ERROR: No se encontró Diario_VIP")
        exit(1)
    
    ws = wb[sheet_name]
    
    # Leer todas las filas
    datos = []
    headers = []
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx == 1:
            headers = [str(cell) if cell is not None else '' for cell in row]
        else:
            fila = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    fila[headers[col_idx]] = cell
            if any(fila.values()):  # Solo agregar si tiene datos
                datos.append(fila)
    
    # Guardar en archivo
    resultado = {
        "hoja": sheet_name,
        "total_filas": len(datos),
        "columnas": headers,
        "datos": datos
    }
    
    # Buscar cliente 1
    cliente_1 = None
    for fila in datos:
        for key, value in fila.items():
            if value == 1 and ('cliente' in key.lower() or 'id' in key.lower() or 'codigo' in key.lower()):
                cliente_1 = fila
                break
        if cliente_1:
            break
    
    if not cliente_1 and len(datos) > 0:
        cliente_1 = datos[0]
        resultado["nota"] = "Se muestra la primera fila como cliente 1"
    
    resultado["cliente_1"] = cliente_1
    
    # Guardar JSON
    with open('datos_cliente1.json', 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2, default=str)
    
    # También guardar texto legible
    with open('datos_cliente1.txt', 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write(f"INFORMACIÓN DEL CLIENTE 1 - HOJA: {sheet_name}\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Total de filas: {len(datos)}\n")
        f.write(f"Columnas: {len(headers)}\n\n")
        f.write("Columnas disponibles:\n")
        for i, col in enumerate(headers, 1):
            f.write(f"  {i}. {col}\n")
        f.write("\n" + "=" * 80 + "\n")
        f.write("DATOS DEL CLIENTE 1:\n")
        f.write("=" * 80 + "\n\n")
        if cliente_1:
            for key, value in cliente_1.items():
                f.write(f"{key}: {value}\n")
        f.write("\n" + "=" * 80 + "\n")
        f.write("TODOS LOS DATOS:\n")
        f.write("=" * 80 + "\n\n")
        for i, fila in enumerate(datos, 1):
            f.write(f"\nFila {i}:\n")
            for key, value in fila.items():
                f.write(f"  {key}: {value}\n")
    
    print("Archivos creados: datos_cliente1.json y datos_cliente1.txt")
    
except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
