from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

fila = 54
print(f"Fila {fila} - Verificando fórmulas del cliente 1 (columnas K-R):")

columnas_cliente = {
    11: 'K - Incremento',
    12: 'L - Decremento',
    13: 'M - (intermedio)',
    14: 'N - Saldo Diario',
    15: 'O - Benef. € Diario',
    16: 'P - Benef. % Diario',
    17: 'Q - Benef. € Acum',
    18: 'R - Benef. % Acum'
}

for col_idx, nombre in columnas_cliente.items():
    celda = ws.cell(row=fila, column=col_idx)
    tiene_formula = celda.data_type == 'f'
    valor = celda.value
    
    if tiene_formula and valor:
        formula = str(valor)
        print(f"  {nombre}: {formula[:100]}")
        
        # Verificar si referencia columnas generales
        if any(col in formula for col in ['E', 'F', 'G', 'H', 'I', 'J']):
            print(f"    -> DEPENDE DE DATOS GENERALES")
    elif valor is not None:
        print(f"  {nombre}: valor={valor}, sin formula")
