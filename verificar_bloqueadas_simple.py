#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script mejorado para verificar el 100% de las celdas bloqueadas/desbloqueadas
"""

import json
import openpyxl
import sys

def obtener_nombre_columna(num_col):
    result = ""
    num = num_col
    while num > 0:
        num -= 1
        result = chr(65 + (num % 26)) + result
        num //= 26
    return result

def tiene_formula(ws, fila, columna):
    try:
        celda = ws.cell(row=fila, column=columna)
        return celda.data_type == 'f' and celda.value is not None
    except:
        return False

def main():
    sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None
    
    errores = []
    
    # Cargar Excel
    print("Cargando Excel...")
    wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
    ws = None
    for sheet_name in wb.sheetnames:
        if 'VIP' in sheet_name.upper():
            ws = wb[sheet_name]
            break
    if not ws:
        ws = wb.active
    
    # Cargar JSON
    print("Cargando JSON...")
    with open('datos_completos.json', 'r', encoding='utf-8') as f:
        datos_json = json.load(f)
    
    hoja_vip = datos_json.get('hojas', {}).get('Diario VIP')
    if not hoja_vip:
        print("ERROR: No se encontró hoja Diario VIP")
        return
    
    # Verificar clientes - INCREMENTO y DECREMENTO
    print("\nVerificando incremento/decremento de clientes...")
    clientes = hoja_vip.get('clientes', [])
    
    estadisticas = {
        'total_celdas': 0,
        'con_formula_excel': 0,
        'bloqueadas_json': 0,
        'con_formula_y_bloqueada': 0,
        'con_formula_y_no_bloqueada': 0
    }
    
    for cliente_idx, cliente in enumerate(clientes, 1):
        numero_cliente = cliente.get('numero_cliente', cliente_idx)
        columna_inicio = cliente.get('columna_inicio', 11)
        datos_diarios = cliente.get('datos_diarios', [])
        
        for dato in datos_diarios:
            fila = dato.get('fila')
            if not fila or fila < 15 or fila > 1120:
                continue
            
            bloqueadas = dato.get('bloqueadas', {})
            
            # Verificar INCREMENTO (columna K = columna_inicio)
            col_inc = columna_inicio
            tiene_formula_excel_inc = tiene_formula(ws, fila, col_inc)
            esta_bloqueada_json_inc = bloqueadas.get('incremento', False)
            
            estadisticas['total_celdas'] += 1
            if tiene_formula_excel_inc:
                estadisticas['con_formula_excel'] += 1
            if esta_bloqueada_json_inc:
                estadisticas['bloqueadas_json'] += 1
            if tiene_formula_excel_inc and esta_bloqueada_json_inc:
                estadisticas['con_formula_y_bloqueada'] += 1
            if tiene_formula_excel_inc and not esta_bloqueada_json_inc:
                estadisticas['con_formula_y_no_bloqueada'] += 1
            
            if tiene_formula_excel_inc and not esta_bloqueada_json_inc:
                try:
                    celda = ws.cell(row=fila, column=col_inc)
                    formula = str(celda.value) if celda.value else None
                except:
                    formula = None
                
                errores.append({
                    'tipo': 'cliente',
                    'cliente': numero_cliente,
                    'fila': fila,
                    'columna': obtener_nombre_columna(col_inc),
                    'campo': 'incremento',
                    'fecha': dato.get('fecha'),
                    'problema': 'Tiene fórmula en Excel pero NO está bloqueada',
                    'formula': formula
                })
            
            # Verificar DECREMENTO (columna L = columna_inicio + 1)
            col_dec = columna_inicio + 1
            tiene_formula_excel_dec = tiene_formula(ws, fila, col_dec)
            esta_bloqueada_json_dec = bloqueadas.get('decremento', False)
            
            estadisticas['total_celdas'] += 1
            if tiene_formula_excel_dec:
                estadisticas['con_formula_excel'] += 1
            if esta_bloqueada_json_dec:
                estadisticas['bloqueadas_json'] += 1
            if tiene_formula_excel_dec and esta_bloqueada_json_dec:
                estadisticas['con_formula_y_bloqueada'] += 1
            if tiene_formula_excel_dec and not esta_bloqueada_json_dec:
                estadisticas['con_formula_y_no_bloqueada'] += 1
            
            if tiene_formula_excel_dec and not esta_bloqueada_json_dec:
                try:
                    celda = ws.cell(row=fila, column=col_dec)
                    formula = str(celda.value) if celda.value else None
                except:
                    formula = None
                
                errores.append({
                    'tipo': 'cliente',
                    'cliente': numero_cliente,
                    'fila': fila,
                    'columna': obtener_nombre_columna(col_dec),
                    'campo': 'decremento',
                    'fecha': dato.get('fecha'),
                    'problema': 'Tiene fórmula en Excel pero NO está bloqueada',
                    'formula': formula
                })
    
    # Mostrar resultados
    print(f"\n{'='*80}")
    print(f"ESTADÍSTICAS")
    print(f"{'='*80}")
    print(f"Total de celdas verificadas (incremento + decremento): {estadisticas['total_celdas']}")
    print(f"Celdas con fórmula en Excel: {estadisticas['con_formula_excel']}")
    print(f"Celdas bloqueadas en JSON: {estadisticas['bloqueadas_json']}")
    print(f"Celdas con fórmula Y bloqueadas: {estadisticas['con_formula_y_bloqueada']}")
    print(f"Celdas con fórmula Y NO bloqueadas: {estadisticas['con_formula_y_no_bloqueada']}")
    
    print(f"\n{'='*80}")
    print(f"RESULTADOS")
    print(f"{'='*80}")
    print(f"Total de errores encontrados: {len(errores)}")
    
    if errores:
        print(f"\nERRORES ENCONTRADOS:")
        for error in errores[:50]:
            print(f"  ❌ Cliente {error['cliente']}, Fila {error['fila']}, Col {error['columna']} ({error['campo']}): {error['problema']}")
            if error.get('formula'):
                print(f"     Fórmula: {error['formula']}")
        
        # Guardar en archivo
        with open('errores_bloqueadas.json', 'w', encoding='utf-8') as f:
            json.dump(errores, f, indent=2, ensure_ascii=False)
        print(f"\n📄 Errores guardados en 'errores_bloqueadas.json'")
    else:
        print("\n✅ No se encontraron errores")
    
    wb.close()

if __name__ == '__main__':
    main()
