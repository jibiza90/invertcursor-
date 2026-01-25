from openpyxl import load_workbook
from datetime import datetime

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb_formulas.sheetnames:
    ws_formulas = wb_formulas['Diario VIP']
    ws_valores = wb_valores['Diario VIP']
    
    print("=" * 80)
    print("ANALISIS COMPLETO DE FORMULAS EN DIARIO VIP")
    print("=" * 80)
    
    # Analizar datos diarios generales (filas 15 en adelante)
    print("\n" + "=" * 80)
    print("DATOS DIARIOS GENERALES - Columnas E-J (5-10)")
    print("=" * 80)
    
    columnas_nombres = {
        5: 'E - Imp. Inicial',
        6: 'F - Imp. Final',
        7: 'G - Benef. €',
        8: 'H - Benef. %',
        9: 'I - Benef. € Acum',
        10: 'J - Benef. % Acum'
    }
    
    # Analizar varias filas representativas
    filas_analizar = [15, 16, 17, 18, 27, 28, 29, 30, 50, 51, 100, 200]
    
    print("\nAnalizando filas representativas:")
    for fila_idx in filas_analizar:
        if fila_idx > ws_valores.max_row:
            continue
            
        fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
        
        if fecha_valor:
            try:
                if isinstance(fecha_valor, datetime):
                    fecha_str = fecha_valor.strftime('%d/%m/%Y')
                else:
                    fecha_str = str(fecha_valor).split()[0]
                
                print(f"\nFila {fila_idx}, Fecha: {fecha_str}")
                
                for col_idx in range(5, 11):
                    nombre_col = columnas_nombres[col_idx]
                    
                    # Verificar fórmula
                    celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                    tiene_formula = False
                    formula_texto = None
                    
                    if celda_formula.value:
                        valor_str = str(celda_formula.value)
                        if valor_str.startswith('='):
                            tiene_formula = True
                            formula_texto = valor_str[:60]  # Primeros 60 caracteres
                    
                    # Valor calculado
                    valor_calculado = ws_valores.cell(row=fila_idx, column=col_idx).value
                    
                    estado = "🔒 BLOQUEADA" if tiene_formula else "✏️ EDITABLE"
                    print(f"  {nombre_col}: {valor_calculado} - {estado}")
                    if tiene_formula and formula_texto:
                        print(f"    Fórmula: {formula_texto}")
                        
            except Exception as e:
                pass
    
    # Analizar patrones
    print("\n" + "=" * 80)
    print("PATRONES ENCONTRADOS:")
    print("=" * 80)
    
    # Contar cuántas filas tienen fórmulas en cada columna
    contadores = {col: {'con_formula': 0, 'sin_formula': 0} for col in range(5, 11)}
    
    for fila_idx in range(15, min(400, ws_valores.max_row + 1)):
        fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
        if fecha_valor:
            for col_idx in range(5, 11):
                celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                if celda_formula.value and str(celda_formula.value).startswith('='):
                    contadores[col_idx]['con_formula'] += 1
                else:
                    contadores[col_idx]['sin_formula'] += 1
    
    print("\nResumen por columna (primeras 400 filas):")
    for col_idx in range(5, 11):
        nombre_col = columnas_nombres[col_idx]
        total = contadores[col_idx]['con_formula'] + contadores[col_idx]['sin_formula']
        pct_formula = (contadores[col_idx]['con_formula'] / total * 100) if total > 0 else 0
        print(f"  {nombre_col}: {contadores[col_idx]['con_formula']} con fórmula ({pct_formula:.1f}%), {contadores[col_idx]['sin_formula']} sin fórmula")
