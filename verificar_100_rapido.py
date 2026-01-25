from openpyxl import load_workbook
import json
from datetime import datetime
from collections import defaultdict

# Abrir archivo con buffer más grande para escritura más rápida
output_file = open('reporte_verificacion_100porciento.txt', 'w', encoding='utf-8', buffering=8192)

def escribir(mensaje):
    output_file.write(mensaje + '\n')

escribir("=" * 80)
escribir("VERIFICACION COMPLETA: 100% DE LAS CELDAS (VERSION OPTIMIZADA)")
escribir(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
escribir("=" * 80)

print("Cargando datos...")

# Cargar Excel
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws_formulas = wb_formulas['Diario VIP']
max_fila = ws_formulas.max_row
max_col = ws_formulas.max_column

# Cargar JSON
with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos_json = json.load(f)

hoja_json = datos_json['hojas']['Diario VIP']
clientes_json = hoja_json.get('clientes', [])
datos_diarios_json = hoja_json.get('datos_diarios_generales', [])
datos_generales_json = hoja_json.get('datos_generales', [])

# Crear mapas para acceso rápido O(1) en lugar de O(n)
mapa_filas_diarios = {d.get('fila'): d for d in datos_diarios_json}
mapa_clientes_json = {c.get('numero_cliente'): c for c in clientes_json}

# Mapeos
col_general = {5: 'imp_inicial', 6: 'imp_final', 7: 'benef_euro', 
               8: 'benef_porcentaje', 9: 'benef_euro_acum', 10: 'benef_porcentaje_acum'}
offset_cliente = {0: 'incremento', 1: 'decremento', 2: 'base', 3: 'saldo_diario',
                  4: 'beneficio_diario', 5: 'beneficio_diario_pct',
                  6: 'beneficio_acumulado', 7: 'beneficio_acumulado_pct'}

def col_a_letra(n):
    result = ""
    while n > 0:
        n -= 1
        result = chr(65 + (n % 26)) + result
        n //= 26
    return result

errores = []
advertencias = []
total_celdas = 0

print("Verificando datos generales...")
# 1. Datos generales
celdas_generales = 0
for fila in range(3, 7):
    fila_json = next((d for d in datos_generales_json if d.get('fila') == fila), None)
    if not fila_json:
        errores.append(f"Fila {fila}: No encontrada en JSON")
        continue
    
    formulas_json = fila_json.get('formulas', {})
    bloqueadas_json = fila_json.get('bloqueadas', {})
    
    for col, campo in col_general.items():
        total_celdas += 1
        celdas_generales += 1
        celda = ws_formulas.cell(row=fila, column=col)
        
        if celda.data_type == 'f' and celda.value is not None:
            formula_excel = str(celda.value)
            formula_json = formulas_json.get(campo)
            bloqueada = bloqueadas_json.get(campo, False)
            
            if not formula_json:
                errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Excel tiene formula pero JSON NO")
            elif formula_excel != formula_json:
                errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Formulas diferentes")
            if not bloqueada:
                errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Tiene formula pero NO bloqueada")

print("Verificando datos diarios generales...")
# 2. Datos diarios generales - VERIFICAR TODAS LAS FILAS DEL EXCEL
celdas_diarias = 0
mapa_filas_diarios = {d.get('fila'): d for d in datos_diarios_json}

# Verificar TODAS las filas desde 15 hasta max_fila del Excel
for fila in range(15, max_fila + 1):
    # Verificar si esta fila tiene fecha (es una fila de datos)
    fecha_excel = ws_formulas.cell(row=fila, column=4).value
    if fecha_excel is None:
        continue
    
    fila_json = mapa_filas_diarios.get(fila)
    if not fila_json:
        advertencias.append(f"Fila {fila} (datos_diarios): Existe en Excel pero NO en JSON")
        continue
    
    formulas_json = fila_json.get('formulas', {})
    bloqueadas_json = fila_json.get('bloqueadas', {})
    
    for col, campo in col_general.items():
        total_celdas += 1
        celdas_diarias += 1
        celda = ws_formulas.cell(row=fila, column=col)
        
        if celda.data_type == 'f' and celda.value is not None:
            formula_excel = str(celda.value)
            formula_json = formulas_json.get(campo)
            bloqueada = bloqueadas_json.get(campo, False)
            
            if not formula_json:
                errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Excel tiene formula pero JSON NO")
            elif formula_excel != formula_json:
                errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Formulas diferentes")
            if not bloqueada:
                errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Tiene formula pero NO bloqueada")

print("Identificando clientes en Excel...")
# 3. Identificar clientes - más rápido usando solo fila 13
clientes_cols = {}
for col in range(11, min(max_col + 1, 500)):
    try:
        num_cliente = ws_formulas.cell(row=13, column=col).value
        if num_cliente is not None:
            try:
                num = int(float(num_cliente))
                if num > 0:
                    if num not in clientes_cols:
                        clientes_cols[num] = []
                    clientes_cols[num].append(col)
            except:
                pass
    except:
        pass

print(f"Verificando {len(clientes_cols)} clientes...")
# 4. Verificar clientes - VERIFICAR TODAS LAS FILAS DEL EXCEL PARA CADA CLIENTE
celdas_clientes = 0
for num_cliente, columnas in sorted(clientes_cols.items()):
    cliente_json = mapa_clientes_json.get(num_cliente)
    if not cliente_json:
        advertencias.append(f"Cliente {num_cliente}: Encontrado en Excel pero NO en JSON")
        continue
    
    datos_cliente = cliente_json.get('datos_diarios', [])
    col_inicio = min(columnas)
    
    # Crear mapa de filas del cliente para acceso rápido
    mapa_filas_cliente = {d.get('fila'): d for d in datos_cliente}
    
    # Verificar TODAS las filas desde 15 hasta max_fila del Excel
    for fila in range(15, max_fila + 1):
        # Verificar si esta fila tiene fecha
        fecha_excel = ws_formulas.cell(row=fila, column=4).value
        if fecha_excel is None:
            continue
        
        fila_data = mapa_filas_cliente.get(fila)
        if not fila_data:
            # Fila existe en Excel pero no en JSON del cliente
            advertencias.append(f"Cliente {num_cliente}, Fila {fila}: Existe en Excel pero NO en JSON")
            continue
        
        formulas_json = fila_data.get('formulas', {})
        bloqueadas_json = fila_data.get('bloqueadas', {})
        
        for offset, campo in offset_cliente.items():
            col = col_inicio + offset
            if col > max_col:
                break
            
            total_celdas += 1
            celdas_clientes += 1
            
            celda = ws_formulas.cell(row=fila, column=col)
            
            if celda.data_type == 'f' and celda.value is not None:
                formula_excel = str(celda.value)
                formula_json = formulas_json.get(campo)
                bloqueada = bloqueadas_json.get(campo, False)
                
                if not formula_json:
                    errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Excel tiene formula pero JSON NO")
                elif formula_excel != formula_json:
                    errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Formulas diferentes")
                if not bloqueada:
                    errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Tiene formula pero NO bloqueada")
            else:
                if campo in ['incremento', 'decremento']:
                    bloqueada = bloqueadas_json.get(campo, False)
                    if bloqueada:
                        errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): NO debe estar bloqueada")

print("Generando reporte...")
# Generar reporte
escribir("\n" + "=" * 80)
escribir("RESUMEN FINAL - VERIFICACION 100% COMPLETA")
escribir("=" * 80)
escribir(f"Total de celdas verificadas: {total_celdas:,}")
escribir(f"  - Datos generales (Filas 3-6): {celdas_generales:,} celdas")
escribir(f"  - Datos diarios generales: {celdas_diarias:,} celdas")
escribir(f"  - Celdas de clientes: {celdas_clientes:,} celdas")
escribir(f"\nErrores críticos encontrados: {len(errores)}")
escribir(f"Advertencias: {len(advertencias)}")

if errores:
    escribir("\n" + "=" * 80)
    escribir("ERRORES CRÍTICOS")
    escribir("=" * 80)
    for i, error in enumerate(errores, 1):
        escribir(f"{i}. {error}")
else:
    escribir("\n✅ NO SE ENCONTRARON ERRORES CRÍTICOS")

if advertencias:
    escribir("\n" + "=" * 80)
    escribir("ADVERTENCIAS")
    escribir("=" * 80)
    for i, adv in enumerate(advertencias[:200], 1):
        escribir(f"{i}. {adv}")
    if len(advertencias) > 200:
        escribir(f"\n... y {len(advertencias) - 200} advertencias más")

escribir("\n" + "=" * 80)
escribir("VERIFICACION COMPLETADA AL 100%")
escribir("=" * 80)
escribir(f"Fecha finalización: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

output_file.close()
print(f"\n✓ Verificación completada: {total_celdas:,} celdas verificadas")
print(f"✓ Errores encontrados: {len(errores)}")
print(f"✓ Reporte guardado en: reporte_verificacion_100porciento.txt")
