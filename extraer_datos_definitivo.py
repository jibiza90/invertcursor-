from openpyxl import load_workbook
import json
from datetime import datetime, date, time

# Cargar dos veces: una para fórmulas, otra para valores
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

def tiene_formula(ws_formulas, fila, columna):
    """Verifica si una celda tiene fórmula"""
    try:
        celda = ws_formulas.cell(row=fila, column=columna)
        
        # Método 1: Verificar el tipo de dato (más confiable)
        if celda.data_type == 'f':  # 'f' significa fórmula
            return True
        
        # Método 2: Verificar si el valor empieza con '='
        if celda.value is not None:
            valor_str = str(celda.value).strip()
            if valor_str.startswith('='):
                return True
        
        return False
    except Exception as e:
        # Si hay error, asumir que no tiene fórmula para no bloquear incorrectamente
        return False

def procesar_hoja(nombre_hoja):
    """Procesa una hoja completa con datos generales y por cliente"""
    if nombre_hoja not in wb_formulas.sheetnames:
        return None
    
    ws_formulas = wb_formulas[nombre_hoja]
    ws_valores = wb_valores[nombre_hoja]
    
    resultado = {
        'nombre': nombre_hoja,
        'datos_generales': [],
        'clientes': []
    }
    
    # 1. EXTRAER DATOS GENERALES (filas 3-6 con resumen del portafolio)
    print(f"\nProcesando datos generales de {nombre_hoja}...")
    
    # Leer headers de la fila 14 para entender las columnas
    headers = {}
    for col_idx in range(1, 15):
        header = ws_valores.cell(row=14, column=col_idx).value
        if header:
            headers[col_idx] = str(header).strip()
    
    filas_generales = [3, 4, 5, 6]
    for fila_idx in filas_generales:
        fila_data = {
            'fila': fila_idx,
            'fecha': None,
            'imp_inicial': None,      # Columna E (5) - IMP. INICIAL
            'imp_final': None,         # Columna F (6) - IMP. FINAL
            'benef_euro': None,        # Columna G (7) - BENEF. €
            'benef_porcentaje': None,  # Columna H (8) - BENEF. %
            'benef_euro_acum': None,   # Columna I (9) - BENEF. €
            'benef_porcentaje_acum': None,  # Columna J (10) - BENEF. %
            'bloqueadas': {}
        }
        
        # Leer fecha (columna D = 4)
        fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
        fila_data['fecha'] = fecha_valor
        
        # Leer valores con nombres reales
        fila_data['imp_inicial'] = ws_valores.cell(row=fila_idx, column=5).value  # E
        fila_data['imp_final'] = ws_valores.cell(row=fila_idx, column=6).value   # F
        fila_data['benef_euro'] = ws_valores.cell(row=fila_idx, column=7).value  # G
        fila_data['benef_porcentaje'] = ws_valores.cell(row=fila_idx, column=8).value  # H
        fila_data['benef_euro_acum'] = ws_valores.cell(row=fila_idx, column=9).value   # I
        fila_data['benef_porcentaje_acum'] = ws_valores.cell(row=fila_idx, column=10).value  # J
        
        # Verificar fórmulas y extraerlas
        fila_data['formulas'] = {}
        for col_key, col_idx in [('imp_inicial', 5), ('imp_final', 6), ('benef_euro', 7), 
                                  ('benef_porcentaje', 8), ('benef_euro_acum', 9), 
                                  ('benef_porcentaje_acum', 10)]:
            if tiene_formula(ws_formulas, fila_idx, col_idx):
                fila_data['bloqueadas'][col_key] = True
                # Extraer la fórmula
                celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                if celda_formula.value:
                    fila_data['formulas'][col_key] = str(celda_formula.value)
            else:
                # Si no tiene fórmula, NO está bloqueada (será editable)
                fila_data['bloqueadas'][col_key] = False
        
        # Solo agregar si tiene datos
        if any(v is not None for v in [fila_data['fecha'], fila_data['imp_inicial'], 
                                        fila_data['imp_final'], fila_data['benef_euro'],
                                        fila_data['benef_porcentaje'], fila_data['benef_euro_acum'],
                                        fila_data['benef_porcentaje_acum']]):
            resultado['datos_generales'].append(fila_data)
    
    print(f"  Datos generales: {len(resultado['datos_generales'])} filas")
    
    # 1b. EXTRAER DATOS DIARIOS GENERALES (desde fila 15)
    print(f"\nExtrayendo datos diarios generales...")
    datos_diarios_generales = []
    
    # Primero, agrupar por fecha para identificar posición dentro del día
    # IMPORTANTE: Usar ws_valores (data_only=True) para obtener fechas calculadas
    fechas_agrupadas = {}
    for fila_idx in range(15, ws_valores.max_row + 1):
        fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
        if fecha_valor is not None:
            # Normalizar fecha para agrupar
            fecha_str = None
            if isinstance(fecha_valor, datetime):
                fecha_str = fecha_valor.strftime('%Y-%m-%d')
            elif isinstance(fecha_valor, date):
                fecha_str = datetime.combine(fecha_valor, time.min).strftime('%Y-%m-%d')
            elif isinstance(fecha_valor, str):
                try:
                    # Intentar parsear fecha
                    if ' ' in fecha_valor:
                        fecha_str = fecha_valor.split()[0]
                    else:
                        fecha_str = fecha_valor
                except:
                    pass
            
            if fecha_str:
                if fecha_str not in fechas_agrupadas:
                    fechas_agrupadas[fecha_str] = []
                fechas_agrupadas[fecha_str].append(fila_idx)
    
    # Ordenar filas dentro de cada fecha
    for fecha_str in fechas_agrupadas:
        fechas_agrupadas[fecha_str].sort()
    
    # Debug: mostrar agrupación
    print(f"  Fechas agrupadas: {len(fechas_agrupadas)} días únicos")
    for fecha_str, filas in sorted(list(fechas_agrupadas.items()))[:3]:
        print(f"    {fecha_str}: {filas}")
    
    def obtener_posicion_dia(fila_idx, fecha_valor):
        """Obtener posición dentro del día (1, 2, o 3)"""
        # Normalizar fecha
        fecha_str = None
        if isinstance(fecha_valor, datetime):
            fecha_str = fecha_valor.strftime('%Y-%m-%d')
        elif isinstance(fecha_valor, date):
            fecha_str = datetime.combine(fecha_valor, time.min).strftime('%Y-%m-%d')
        elif isinstance(fecha_valor, str):
            try:
                if ' ' in fecha_valor:
                    fecha_str = fecha_valor.split()[0]
                else:
                    fecha_str = fecha_valor
            except:
                pass
        
        if fecha_str and fecha_str in fechas_agrupadas:
            filas_del_dia = fechas_agrupadas[fecha_str]
            try:
                return filas_del_dia.index(fila_idx) + 1  # 1-indexed
            except:
                return None
        return None
    
    for fila_idx in range(15, ws_valores.max_row + 1):
        fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
        
        if fecha_valor is not None:
            # Determinar posición dentro del día (1, 2, o 3)
            posicion_dia = obtener_posicion_dia(fila_idx, fecha_valor)
            
            # Debug: mostrar posiciones para primeras filas
            if fila_idx <= 20:
                print(f"DEBUG: Fila {fila_idx}, fecha={fecha_valor}, posicion_dia={posicion_dia}")
            
            dato_diario = {
                'fila': fila_idx,
                'fecha': fecha_valor,
                'imp_inicial': ws_valores.cell(row=fila_idx, column=5).value,
                'imp_final': ws_valores.cell(row=fila_idx, column=6).value,
                'benef_euro': ws_valores.cell(row=fila_idx, column=7).value,
                'benef_porcentaje': ws_valores.cell(row=fila_idx, column=8).value,
                'benef_euro_acum': ws_valores.cell(row=fila_idx, column=9).value,
                'benef_porcentaje_acum': ws_valores.cell(row=fila_idx, column=10).value,
                'bloqueadas': {},
                'formulas': {}
            }
            
            # Verificar fórmulas y extraerlas para datos diarios generales
            bloqueadas_diario = {}
            formulas_diario = {}
            
            for col_key, col_idx in [('imp_inicial', 5), ('imp_final', 6), ('benef_euro', 7), 
                                      ('benef_porcentaje', 8), ('benef_euro_acum', 9), 
                                      ('benef_porcentaje_acum', 10)]:
                tiene_formula_excel = tiene_formula(ws_formulas, fila_idx, col_idx)
                
                # REGLAS ESPECIALES según el usuario:
                # 1. Beneficio €, Beneficio %, Beneficio € Acumulado, Beneficio % Acumulado: SIEMPRE BLOQUEADAS
                if col_key in ['benef_euro', 'benef_porcentaje', 'benef_euro_acum', 'benef_porcentaje_acum']:
                    bloqueadas_diario[col_key] = True
                    if tiene_formula_excel:
                        celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                        if celda_formula.value:
                            formulas_diario[col_key] = str(celda_formula.value)
                
                # 2. IMP_FINAL: Solo editable en la posición 3 del día, bloqueada en posiciones 1 y 2
                elif col_key == 'imp_final':
                    if posicion_dia == 3:
                        # Posición 3: EDITABLE (a menos que tenga fórmula)
                        bloqueadas_diario[col_key] = tiene_formula_excel
                        if tiene_formula_excel:
                            celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                            if celda_formula.value:
                                formulas_diario[col_key] = str(celda_formula.value)
                    elif posicion_dia in [1, 2]:
                        # Posiciones 1 y 2: BLOQUEADA
                        bloqueadas_diario[col_key] = True
                        if tiene_formula_excel:
                            celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                            if celda_formula.value:
                                formulas_diario[col_key] = str(celda_formula.value)
                    else:
                        # Si no se pudo determinar posición, usar lógica por defecto (bloquear si tiene fórmula)
                        bloqueadas_diario[col_key] = tiene_formula_excel
                        if tiene_formula_excel:
                            celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                            if celda_formula.value:
                                formulas_diario[col_key] = str(celda_formula.value)
                
                # 3. IMP_INICIAL: Bloquear solo si tiene fórmula
                else:  # imp_inicial
                    bloqueadas_diario[col_key] = tiene_formula_excel
                    if tiene_formula_excel:
                        celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
                        if celda_formula.value:
                            formulas_diario[col_key] = str(celda_formula.value)
            
            dato_diario['bloqueadas'] = bloqueadas_diario
            dato_diario['formulas'] = formulas_diario
            
            datos_diarios_generales.append(dato_diario)
    
    resultado['datos_diarios_generales'] = datos_diarios_generales
    print(f"  Datos diarios generales: {len(datos_diarios_generales)} registros")
    
    # 2. IDENTIFICAR CLIENTES (fila 13 tiene los números de clientes)
    print(f"\nIdentificando clientes...")
    
    fila_nombres = 13
    clientes_columnas = {}
    
    # Buscar en TODAS las columnas, no limitar a 200
    for col_idx in range(1, ws_valores.max_column + 1):
        celda = ws_valores.cell(row=fila_nombres, column=col_idx)
        valor = celda.value
        
        if valor is not None:
            try:
                numero_cliente = int(valor)
                if numero_cliente > 0 and numero_cliente not in clientes_columnas:
                    grupo_columnas = {
                        'numero': numero_cliente,
                        'columna_inicio': col_idx,
                        'INCREMENTO': col_idx,
                        'DECREMENTO': col_idx + 1,
                        'FINAL': col_idx + 3,
                        'BENEF_€_DIARIO': col_idx + 4,
                        'BENEF_%_DIARIO': col_idx + 5,
                        'BENEF_€_ACUM': col_idx + 6,
                        'BENEF_%_ACUM': col_idx + 7
                    }
                    
                    clientes_columnas[numero_cliente] = grupo_columnas
            except:
                pass
    
    print(f"  Clientes encontrados: {len(clientes_columnas)}")
    
    # 3. EXTRAER DATOS DIARIOS POR CLIENTE (desde fila 15)
    print(f"\nExtrayendo datos diarios por cliente...")
    
    for numero_cliente, grupo_columnas in clientes_columnas.items():
        cliente_data = {
            'numero_cliente': numero_cliente,
            'columna_inicio': grupo_columnas['columna_inicio'],
            'incrementos_total': 0,
            'decrementos_total': 0,
            'saldo_actual': None,
            'datos_diarios': []
        }
        
        incrementos = []
        decrementos = []
        saldos_finales = []
        
        # Leer TODAS las filas desde la 15 hasta el final para datos diarios
        for fila_idx in range(15, ws_valores.max_row + 1):
            # Leer fecha (columna D = 4) - usar valores
            fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
            
            # Leer valores del cliente - usar valores calculados
            incremento = ws_valores.cell(row=fila_idx, column=grupo_columnas['INCREMENTO']).value
            decremento = ws_valores.cell(row=fila_idx, column=grupo_columnas['DECREMENTO']).value
            # Columna M (intermedio/base) - clave para que N/O/P/Q/R se calculen
            base = ws_valores.cell(row=fila_idx, column=grupo_columnas['INCREMENTO'] + 2).value
            final = ws_valores.cell(row=fila_idx, column=grupo_columnas['FINAL']).value
            benef_diario = ws_valores.cell(row=fila_idx, column=grupo_columnas['BENEF_€_DIARIO']).value
            benef_diario_pct = ws_valores.cell(row=fila_idx, column=grupo_columnas['BENEF_%_DIARIO']).value
            benef_acum = ws_valores.cell(row=fila_idx, column=grupo_columnas['BENEF_€_ACUM']).value
            benef_acum_pct = ws_valores.cell(row=fila_idx, column=grupo_columnas['BENEF_%_ACUM']).value
            
            # IMPORTANTE: Las columnas M-R (BASE, FINAL, BENEF_€_DIARIO, BENEF_%_DIARIO, BENEF_€_ACUM, BENEF_%_ACUM)
            # están SIEMPRE bloqueadas porque se calculan automáticamente con fórmulas basadas en la columna F
            bloqueadas = {
                'incremento': False,  # Columna K - por defecto editable, pero se verifica si tiene fórmula
                'decremento': False,  # Columna L - por defecto editable, pero se verifica si tiene fórmula
                'base': True,  # Columna M - siempre bloqueada (fórmula)
                'saldo_diario': True,  # Columna N - siempre bloqueada
                'beneficio_diario': True,  # Columna O - siempre bloqueada
                'beneficio_diario_pct': True,  # Columna P - siempre bloqueada
                'beneficio_acumulado': True,  # Columna Q - siempre bloqueada
                'beneficio_acumulado_pct': True  # Columna R - siempre bloqueada
            }
            
            # Extraer fórmulas de las columnas bloqueadas del cliente
            # IMPORTANTE: Orden según dependencias para evaluación correcta
            formulas_cliente = {}
            
            # Verificar si incremento y decremento tienen fórmulas (deben estar bloqueadas si tienen fórmula)
            if tiene_formula(ws_formulas, fila_idx, grupo_columnas['INCREMENTO']):
                bloqueadas['incremento'] = True
                celda_formula = ws_formulas.cell(row=fila_idx, column=grupo_columnas['INCREMENTO'])
                if celda_formula.value:
                    formulas_cliente['incremento'] = str(celda_formula.value)
            
            if tiene_formula(ws_formulas, fila_idx, grupo_columnas['DECREMENTO']):
                bloqueadas['decremento'] = True
                celda_formula = ws_formulas.cell(row=fila_idx, column=grupo_columnas['DECREMENTO'])
                if celda_formula.value:
                    formulas_cliente['decremento'] = str(celda_formula.value)
            
            columna_map_cliente = [
                ('base', grupo_columnas['INCREMENTO'] + 2),  # Columna M (K+2) - depende de F, K, L, N anterior
                ('beneficio_diario', grupo_columnas['BENEF_€_DIARIO']),  # Columna O - depende de base y H
                ('beneficio_diario_pct', grupo_columnas['BENEF_%_DIARIO']),  # Columna P - depende de base y H
                ('saldo_diario', grupo_columnas['FINAL']),  # Columna N - depende de base y beneficio_diario
                ('beneficio_acumulado', grupo_columnas['BENEF_€_ACUM']),  # Columna Q - depende de saldo_diario
                ('beneficio_acumulado_pct', grupo_columnas['BENEF_%_ACUM'])  # Columna R - depende de beneficio_acumulado
            ]
            
            for campo_cliente, col_idx_cliente in columna_map_cliente:
                if tiene_formula(ws_formulas, fila_idx, col_idx_cliente):
                    celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx_cliente)
                    if celda_formula.value:
                        formulas_cliente[campo_cliente] = str(celda_formula.value)
            
            # Si hay fecha, es una fila de datos diarios
            # O si tiene fórmulas en incremento/decremento (filas de resumen mensual)
            tiene_formula_inc = tiene_formula(ws_formulas, fila_idx, grupo_columnas['INCREMENTO'])
            tiene_formula_dec = tiene_formula(ws_formulas, fila_idx, grupo_columnas['DECREMENTO'])
            
            if fecha_valor is not None or tiene_formula_inc or tiene_formula_dec:
                dato_diario = {
                    'fila': fila_idx,
                    'fecha': fecha_valor,
                    'incremento': incremento,  # Columna K - EDITABLE o BLOQUEADA si tiene fórmula
                    'decremento': decremento,  # Columna L - EDITABLE o BLOQUEADA si tiene fórmula
                    'base': base,  # Columna M - BLOQUEADA (fórmula)
                    'saldo_diario': final,
                    'beneficio_diario': benef_diario,
                    'beneficio_diario_pct': benef_diario_pct,
                    'beneficio_acumulado': benef_acum,
                    'beneficio_acumulado_pct': benef_acum_pct,
                    'bloqueadas': bloqueadas,
                    'formulas': formulas_cliente
                }
                cliente_data['datos_diarios'].append(dato_diario)
            
            # Acumular saldos finales (columna N) - todas las filas
            if final is not None:
                try:
                    final_val = float(final)
                    if final_val != 0:
                        saldos_finales.append((fila_idx, final_val))
                except:
                    pass
        
        # IMPORTANTE: Acumular incrementos y decrementos SOLO desde fila 15 hasta 1130
        # Y SOLO para fechas del año 2026 (01/01/2026 a 31/12/2026)
        fecha_inicio_2026 = datetime(2026, 1, 1)
        fecha_fin_2026 = datetime(2026, 12, 31, 23, 59, 59)
        
        for fila_idx in range(15, min(1131, ws_valores.max_row + 1)):  # 1131 porque range es exclusivo
            # Leer fecha para verificar si está en 2026
            fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
            
            # Verificar si la fecha está en el rango 2026
            fecha_valida = False
            if fecha_valor is not None:
                try:
                    fecha_dt = None
                    if isinstance(fecha_valor, datetime):
                        fecha_dt = fecha_valor
                    elif isinstance(fecha_valor, date):
                        fecha_dt = datetime.combine(fecha_valor, time.min)
                    elif isinstance(fecha_valor, str):
                        # Intentar diferentes formatos
                        try:
                            fecha_dt = datetime.strptime(fecha_valor.split()[0], '%Y-%m-%d')
                        except:
                            try:
                                fecha_dt = datetime.strptime(fecha_valor.split()[0], '%d/%m/%Y')
                            except:
                                pass
                    
                    if fecha_dt and isinstance(fecha_dt, datetime):
                        # Comparar solo año, mes y día
                        if fecha_dt.year == 2026:
                            fecha_valida = True
                except Exception as e:
                    pass
            
            # Solo contar incrementos/decrementos si la fecha está en 2026
            if fecha_valida:
                incremento = ws_valores.cell(row=fila_idx, column=grupo_columnas['INCREMENTO']).value
                decremento = ws_valores.cell(row=fila_idx, column=grupo_columnas['DECREMENTO']).value
                
                # Acumular incrementos (columna K) - solo fechas 2026
                if incremento is not None:
                    try:
                        inc_val = float(incremento)
                        if inc_val != 0:
                            incrementos.append(inc_val)
                    except:
                        pass
                
                # Acumular decrementos (columna L) - solo fechas 2026
                if decremento is not None:
                    try:
                        dec_val = float(decremento)
                        if dec_val != 0:
                            decrementos.append(dec_val)
                    except:
                        pass
        
        # Calcular totales
        cliente_data['incrementos_total'] = sum(incrementos)
        cliente_data['decrementos_total'] = sum(decrementos)
        
        # Saldo actual: último valor no cero de columna N
        if saldos_finales:
            cliente_data['saldo_actual'] = saldos_finales[-1][1]
            cliente_data['saldo_actual_fila'] = saldos_finales[-1][0]
        
        resultado['clientes'].append(cliente_data)
    
    # Ordenar clientes por número
    resultado['clientes'].sort(key=lambda x: x['numero_cliente'])
    
    print(f"  Clientes procesados: {len(resultado['clientes'])}")
    if resultado['clientes']:
        cliente1 = resultado['clientes'][0]
        print(f"  Cliente 1 - Datos diarios: {len(cliente1['datos_diarios'])} registros")
        print(f"  Cliente 1 - Incrementos (desde K15): {cliente1['incrementos_total']:,.2f}")
        print(f"  Cliente 1 - Decrementos (desde L15): {cliente1['decrementos_total']:,.2f}")
        if cliente1['saldo_actual']:
            print(f"  Cliente 1 - Saldo actual: {cliente1['saldo_actual']:,.2f} (Fila {cliente1['saldo_actual_fila']})")
    
    return resultado

# Procesar ambas hojas
resultado_completo = {
    'fecha_extraccion': datetime.now().isoformat(),
    'hojas': {}
}

print("=" * 80)
print("EXTRACCION DEFINITIVA CON DETECCION DE FORMULAS")
print("=" * 80)

# Procesar Diario STD
if 'Diario STD' in wb_formulas.sheetnames:
    resultado_completo['hojas']['Diario STD'] = procesar_hoja('Diario STD')

# Procesar Diario VIP
if 'Diario VIP' in wb_formulas.sheetnames:
    resultado_completo['hojas']['Diario VIP'] = procesar_hoja('Diario VIP')

# Guardar resultado
with open('datos_completos.json', 'w', encoding='utf-8') as f:
    json.dump(resultado_completo, f, ensure_ascii=False, indent=2, default=str)

print("\n" + "=" * 80)
print(f"Datos guardados en: datos_completos.json")
print("=" * 80)
