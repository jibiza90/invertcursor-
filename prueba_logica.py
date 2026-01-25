#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de prueba para verificar la lógica de bloqueo
"""

from openpyxl import load_workbook
from datetime import datetime, date, time

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

ws_formulas = None
ws_valores_obj = None
for sheet_name in wb_formulas.sheetnames:
    if 'VIP' in sheet_name.upper():
        ws_formulas = wb_formulas[sheet_name]
        ws_valores_obj = wb_valores[sheet_name]
        break

if not ws_formulas:
    ws_formulas = wb_formulas.active
    ws_valores_obj = wb_valores.active

def tiene_formula(ws, fila, columna):
    try:
        celda = ws.cell(row=fila, column=columna)
        return celda.data_type == 'f' and celda.value is not None
    except:
        return False

# Agrupar por fecha
fechas_agrupadas = {}
for fila_idx in range(15, 51):
    fecha_valor = ws_valores_obj.cell(row=fila_idx, column=4).value
    if fecha_valor is not None:
        fecha_str = None
        if isinstance(fecha_valor, datetime):
            fecha_str = fecha_valor.strftime('%Y-%m-%d')
        elif isinstance(fecha_valor, date):
            fecha_str = datetime.combine(fecha_valor, time.min).strftime('%Y-%m-%d')
        elif isinstance(fecha_valor, str):
            if ' ' in fecha_valor:
                fecha_str = fecha_valor.split()[0]
            else:
                fecha_str = fecha_valor
        
        if fecha_str:
            if fecha_str not in fechas_agrupadas:
                fechas_agrupadas[fecha_str] = []
            fechas_agrupadas[fecha_str].append(fila_idx)

# Ordenar
for fecha_str in fechas_agrupadas:
    fechas_agrupadas[fecha_str].sort()

print("="*80)
print("PRUEBA DE LÓGICA - Primeras 3 fechas")
print("="*80)

for fecha_str, filas in sorted(fechas_agrupadas.items())[:3]:
    print(f"\nFecha: {fecha_str}, Filas: {filas}")
    for idx, fila in enumerate(filas, 1):
        posicion = idx
        print(f"  Fila {fila} (posición {posicion}):")
        
        # IMP_FINAL
        tiene_formula_imp_final = tiene_formula(ws_formulas, fila, 6)
        if posicion == 3:
            deberia_estar = "EDITABLE" if not tiene_formula_imp_final else "BLOQUEADA"
        else:
            deberia_estar = "BLOQUEADA"
        print(f"    IMP_FINAL: {deberia_estar} (tiene fórmula: {tiene_formula_imp_final})")
        
        # BENEF_EURO
        tiene_formula_benef = tiene_formula(ws_formulas, fila, 7)
        deberia_estar_benef = "BLOQUEADA"  # Siempre bloqueada
        print(f"    BENEF_EURO: {deberia_estar_benef} (tiene fórmula: {tiene_formula_benef})")

wb_formulas.close()
wb_valores.close()
