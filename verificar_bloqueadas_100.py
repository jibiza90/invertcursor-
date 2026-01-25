#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para verificar el 100% de las celdas si deberían estar bloqueadas o no.
Compara el JSON con el Excel original para verificar que las celdas con fórmulas
estén bloqueadas y las celdas sin fórmulas estén desbloqueadas.
"""

import json
import openpyxl
from openpyxl.formula.translate import Translator
from collections import defaultdict

def tiene_formula(ws, fila, columna):
    """Verificar si una celda tiene fórmula"""
    try:
        celda = ws.cell(row=fila, column=columna)
        # Verificar si tiene fórmula (data_type == 'f' significa fórmula)
        if celda.data_type == 'f':
            return True
        # También verificar si el valor empieza con '=' (por si acaso)
        if celda.value and isinstance(celda.value, str) and celda.value.startswith('='):
            return True
        return False
    except Exception as e:
        # Si hay error, asumir que no tiene fórmula
        return False

def obtener_nombre_columna(num_col):
    """Convertir número de columna a letra (1=A, 27=AA, etc.)"""
    result = ""
    num = num_col
    while num > 0:
        num -= 1
        result = chr(65 + (num % 26)) + result
        num //= 26
    return result

def verificar_datos_generales(ws, datos_json, errores):
    """Verificar datos generales (filas 1-14 aproximadamente)"""
    print("\n" + "="*80)
    print("VERIFICANDO DATOS GENERALES")
    print("="*80)
    
    datos_generales = datos_json.get('datos_generales', [])
    print(f"Total de filas en datos generales: {len(datos_generales)}")
    
    # Mapeo de columnas a campos
    columna_map = {
        5: 'imp_inicial',   # E
        6: 'imp_final',     # F
        7: 'benef_euro',    # G
        8: 'benef_porcentaje', # H
        9: 'benef_euro_acum',  # I
        10: 'benef_porcentaje_acum' # J
    }
    
    total_verificadas = 0
    errores_encontrados = 0
    
    for dato in datos_generales:
        fila = dato.get('fila')
        if not fila:
            continue
        
        bloqueadas_json = dato.get('bloqueadas', {})
        formulas_json = dato.get('formulas', {})
        
        for col_num, campo in columna_map.items():
            total_verificadas += 1
            col_letra = obtener_nombre_columna(col_num)
            
            # Verificar si tiene fórmula en Excel
            tiene_formula_excel = tiene_formula(ws, fila, col_num)
            
            # Verificar estado en JSON
            esta_bloqueada_json = bloqueadas_json.get(campo, False)
            tiene_formula_json = campo in formulas_json and formulas_json[campo]
            
            # La celda DEBE estar bloqueada si tiene fórmula en Excel
            if tiene_formula_excel and not esta_bloqueada_json:
                errores_encontrados += 1
                error = {
                    'tipo': 'datos_generales',
                    'fila': fila,
                    'columna': col_letra,
                    'campo': campo,
                    'problema': f'Tiene fórmula en Excel pero NO está bloqueada en JSON',
                    'excel_tiene_formula': True,
                    'json_bloqueada': False,
                    'json_tiene_formula': tiene_formula_json
                }
                errores.append(error)
                if errores_encontrados <= 20:
                    print(f"❌ Fila {fila}, Col {col_letra} ({campo}): Tiene fórmula en Excel pero NO está bloqueada")
            
            # La celda NO debe estar bloqueada si NO tiene fórmula en Excel (a menos que sea por otra razón)
            if not tiene_formula_excel and esta_bloqueada_json:
                # Verificar si tiene fórmula en JSON (puede ser que se haya agregado manualmente)
                if not tiene_formula_json:
                    errores_encontrados += 1
                    error = {
                        'tipo': 'datos_generales',
                        'fila': fila,
                        'columna': col_letra,
                        'campo': campo,
                        'problema': f'NO tiene fórmula en Excel pero está bloqueada en JSON',
                        'excel_tiene_formula': False,
                        'json_bloqueada': True,
                        'json_tiene_formula': tiene_formula_json
                    }
                    errores.append(error)
                    if errores_encontrados <= 20:
                        print(f"⚠️ Fila {fila}, Col {col_letra} ({campo}): NO tiene fórmula en Excel pero está bloqueada")
    
    print(f"\n✅ Verificadas {total_verificadas} celdas en datos generales")
    print(f"❌ Encontrados {errores_encontrados} errores")

def verificar_datos_diarios_generales(ws, datos_json, errores):
    """Verificar datos diarios generales (filas 15+)"""
    print("\n" + "="*80)
    print("VERIFICANDO DATOS DIARIOS GENERALES")
    print("="*80)
    
    datos_diarios = datos_json.get('datos_diarios_generales', [])
    
    # Mapeo de columnas a campos
    columna_map = {
        5: 'imp_inicial',   # E
        6: 'imp_final',     # F
        7: 'benef_euro',    # G
        8: 'benef_porcentaje', # H
        9: 'benef_euro_acum',  # I
        10: 'benef_porcentaje_acum' # J
    }
    
    total_verificadas = 0
    errores_encontrados = 0
    
    for dato in datos_diarios:
        fila = dato.get('fila')
        if not fila or fila < 15 or fila > 1120:
            continue
        
        bloqueadas_json = dato.get('bloqueadas', {})
        formulas_json = dato.get('formulas', {})
        
        for col_num, campo in columna_map.items():
            total_verificadas += 1
            col_letra = obtener_nombre_columna(col_num)
            
            # Verificar si tiene fórmula en Excel
            tiene_formula_excel = tiene_formula(ws, fila, col_num)
            
            # Verificar estado en JSON
            esta_bloqueada_json = bloqueadas_json.get(campo, False)
            tiene_formula_json = campo in formulas_json and formulas_json[campo]
            
            # La celda DEBE estar bloqueada si tiene fórmula en Excel
            if tiene_formula_excel and not esta_bloqueada_json:
                errores_encontrados += 1
                error = {
                    'tipo': 'datos_diarios_generales',
                    'fila': fila,
                    'columna': col_letra,
                    'campo': campo,
                    'fecha': dato.get('fecha'),
                    'problema': f'Tiene fórmula en Excel pero NO está bloqueada en JSON',
                    'excel_tiene_formula': True,
                    'json_bloqueada': False,
                    'json_tiene_formula': tiene_formula_json
                }
                errores.append(error)
                if errores_encontrados <= 20:  # Mostrar solo los primeros 20
                    print(f"❌ Fila {fila}, Col {col_letra} ({campo}): Tiene fórmula en Excel pero NO está bloqueada")
            
            # La celda NO debe estar bloqueada si NO tiene fórmula en Excel
            if not tiene_formula_excel and esta_bloqueada_json and not tiene_formula_json:
                errores_encontrados += 1
                error = {
                    'tipo': 'datos_diarios_generales',
                    'fila': fila,
                    'columna': col_letra,
                    'campo': campo,
                    'fecha': dato.get('fecha'),
                    'problema': f'NO tiene fórmula en Excel pero está bloqueada en JSON',
                    'excel_tiene_formula': False,
                    'json_bloqueada': True,
                    'json_tiene_formula': False
                }
                errores.append(error)
                if errores_encontrados <= 20:
                    print(f"⚠️ Fila {fila}, Col {col_letra} ({campo}): NO tiene fórmula en Excel pero está bloqueada")
    
    print(f"\n✅ Verificadas {total_verificadas} celdas en datos diarios generales")
    print(f"❌ Encontrados {errores_encontrados} errores")

def verificar_clientes(ws, datos_json, errores):
    """Verificar datos de clientes"""
    print("\n" + "="*80)
    print("VERIFICANDO DATOS DE CLIENTES")
    print("="*80)
    
    clientes = datos_json.get('clientes', [])
    
    # Mapeo de campos a columnas relativas (K=0, L=1, M=2, etc.)
    campo_map = {
        'incremento': 0,      # K (columna 11)
        'decremento': 1,      # L (columna 12)
        'base': 2,            # M (columna 13)
        'saldo_diario': 3,    # N (columna 14)
        'beneficio_diario': 4, # O (columna 15)
        'beneficio_diario_pct': 5, # P (columna 16)
        'beneficio_acumulado': 6, # Q (columna 17)
        'beneficio_acumulado_pct': 7 # R (columna 18)
    }
    
    total_verificadas = 0
    errores_encontrados = 0
    ejemplos_formulas = []  # Guardar ejemplos de celdas con fórmulas
    
    for cliente_idx, cliente in enumerate(clientes, 1):
        numero_cliente = cliente.get('numero_cliente', cliente_idx)
        columna_inicio = cliente.get('columna_inicio', 11)  # Por defecto K (11)
        datos_diarios = cliente.get('datos_diarios', [])
        
        for dato in datos_diarios:
            fila = dato.get('fila')
            if not fila or fila < 15 or fila > 1120:
                continue
            
            bloqueadas_json = dato.get('bloqueadas', {})
            formulas_json = dato.get('formulas', {})
            
            for campo, offset_col in campo_map.items():
                total_verificadas += 1
                col_num_excel = columna_inicio + offset_col
                col_letra = obtener_nombre_columna(col_num_excel)
                
                # Verificar si tiene fórmula en Excel
                tiene_formula_excel = tiene_formula(ws, fila, col_num_excel)
                
                # Guardar ejemplos de fórmulas encontradas (solo para incremento/decremento)
                if tiene_formula_excel and campo in ['incremento', 'decremento'] and len(ejemplos_formulas) < 20:
                    try:
                        celda_excel = ws.cell(row=fila, column=col_num_excel)
                        formula_excel = str(celda_excel.value) if celda_excel.value else None
                        ejemplos_formulas.append({
                            'cliente': numero_cliente,
                            'fila': fila,
                            'columna': col_letra,
                            'campo': campo,
                            'formula': formula_excel,
                            'bloqueada_json': bloqueadas_json.get(campo, False),
                            'data_type': celda_excel.data_type
                        })
                    except Exception as e:
                        pass
                
                # Verificar estado en JSON
                esta_bloqueada_json = bloqueadas_json.get(campo, False)
                tiene_formula_json = campo in formulas_json and formulas_json[campo]
                
                # La celda DEBE estar bloqueada si tiene fórmula en Excel
                if tiene_formula_excel and not esta_bloqueada_json:
                    errores_encontrados += 1
                    # Obtener la fórmula de Excel para el reporte
                    try:
                        celda_excel = ws.cell(row=fila, column=col_num_excel)
                        formula_excel = str(celda_excel.value) if celda_excel.value else None
                    except:
                        formula_excel = None
                    
                    error = {
                        'tipo': 'cliente',
                        'cliente': numero_cliente,
                        'fila': fila,
                        'columna': col_letra,
                        'campo': campo,
                        'fecha': dato.get('fecha'),
                        'problema': f'Tiene fórmula en Excel pero NO está bloqueada en JSON',
                        'excel_tiene_formula': True,
                        'excel_formula': formula_excel,
                        'json_bloqueada': False,
                        'json_tiene_formula': tiene_formula_json
                    }
                    errores.append(error)
                    if errores_encontrados <= 50:  # Mostrar más errores
                        print(f"❌ Cliente {numero_cliente}, Fila {fila}, Col {col_letra} ({campo}): Tiene fórmula '{formula_excel}' pero NO bloqueada")
                
                # La celda NO debe estar bloqueada si NO tiene fórmula en Excel
                # EXCEPCIÓN: base, saldo_diario, beneficio_* siempre deben estar bloqueadas (son calculadas)
                campos_siempre_bloqueados = ['base', 'saldo_diario', 'beneficio_diario', 
                                            'beneficio_diario_pct', 'beneficio_acumulado', 
                                            'beneficio_acumulado_pct']
                
                if not tiene_formula_excel and esta_bloqueada_json and not tiene_formula_json:
                    # Si es incremento o decremento y no tiene fórmula, NO debería estar bloqueada
                    if campo in ['incremento', 'decremento']:
                        errores_encontrados += 1
                        error = {
                            'tipo': 'cliente',
                            'cliente': numero_cliente,
                            'fila': fila,
                            'columna': col_letra,
                            'campo': campo,
                            'fecha': dato.get('fecha'),
                            'problema': f'NO tiene fórmula en Excel pero está bloqueada en JSON',
                            'excel_tiene_formula': False,
                            'json_bloqueada': True,
                            'json_tiene_formula': False
                        }
                        errores.append(error)
                        if errores_encontrados <= 20:
                            print(f"⚠️ Cliente {numero_cliente}, Fila {fila}, Col {col_letra} ({campo}): NO tiene fórmula pero está bloqueada")
    
    print(f"\n✅ Verificadas {total_verificadas} celdas en datos de clientes")
    print(f"❌ Encontrados {errores_encontrados} errores")
    
    # Mostrar ejemplos de fórmulas encontradas
    if ejemplos_formulas:
        print(f"\n📋 Ejemplos de fórmulas encontradas en Excel (incremento/decremento):")
        for ejemplo in ejemplos_formulas[:20]:
            estado = "✅ BLOQUEADA" if ejemplo['bloqueada_json'] else "❌ NO BLOQUEADA"
            print(f"   Cliente {ejemplo['cliente']}, Fila {ejemplo['fila']}, Col {ejemplo['columna']} ({ejemplo['campo']}): {ejemplo['formula']} - {estado} (data_type: {ejemplo['data_type']})")
    else:
        print(f"\n⚠️ No se encontraron fórmulas en incremento/decremento (puede ser normal si no hay fórmulas)")

def main():
    import sys
    sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None
    
    print("="*80)
    print("VERIFICACIÓN 100% DE CELDAS BLOQUEADAS/DESBLOQUEADAS")
    print("="*80)
    
    # Cargar Excel
    print("\n📂 Cargando archivo Excel...")
    try:
        wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
        ws = wb.active
        print("✅ Excel cargado correctamente")
    except Exception as e:
        print(f"❌ Error al cargar Excel: {e}")
        return
    
    # Cargar JSON
    print("\n📂 Cargando archivo JSON...")
    try:
        with open('datos_completos.json', 'r', encoding='utf-8') as f:
            datos_json = json.load(f)
        print("✅ JSON cargado correctamente")
    except Exception as e:
        print(f"❌ Error al cargar JSON: {e}")
        return
    
    # Buscar hoja "Diario VIP"
    hoja_vip = None
    for nombre_hoja, hoja_data in datos_json.get('hojas', {}).items():
        if nombre_hoja == 'Diario VIP':
            hoja_vip = hoja_data
            break
    
    if not hoja_vip:
        print("❌ No se encontró la hoja 'Diario VIP'")
        return
    
    errores = []
    
    # Verificar datos generales
    verificar_datos_generales(ws, hoja_vip, errores)
    
    # Verificar datos diarios generales
    verificar_datos_diarios_generales(ws, hoja_vip, errores)
    
    # Verificar datos de clientes
    verificar_clientes(ws, hoja_vip, errores)
    
    # Resumen final
    print("\n" + "="*80)
    print("RESUMEN FINAL")
    print("="*80)
    print(f"Total de errores encontrados: {len(errores)}")
    
    if errores:
        print("\n❌ ERRORES ENCONTRADOS:")
        
        # Agrupar por tipo
        por_tipo = defaultdict(list)
        for error in errores:
            por_tipo[error['tipo']].append(error)
        
        for tipo, lista_errores in por_tipo.items():
            print(f"\n  {tipo.upper()}: {len(lista_errores)} errores")
            for error in lista_errores[:10]:  # Mostrar solo los primeros 10
                if tipo == 'cliente':
                    print(f"    - Cliente {error['cliente']}, Fila {error['fila']}, Col {error['columna']} ({error['campo']}): {error['problema']}")
                else:
                    print(f"    - Fila {error['fila']}, Col {error['columna']} ({error['campo']}): {error['problema']}")
        
        # Guardar reporte detallado
        with open('reporte_bloqueadas_errores.json', 'w', encoding='utf-8') as f:
            json.dump(errores, f, indent=2, ensure_ascii=False)
        print(f"\n📄 Reporte detallado guardado en 'reporte_bloqueadas_errores.json'")
    else:
        print("\n✅ PERFECTO: Todas las celdas están correctamente bloqueadas/desbloqueadas")
    
    wb.close()

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        import traceback
        print(f"❌ Error: {e}")
        traceback.print_exc()
