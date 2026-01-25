#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificar la formula base del Cliente 2 directamente en Excel
"""

import openpyxl

wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
hoja_std = wb['Diario STD']

print("="*80)
print("VERIFICAR FORMULA BASE CLIENTE 2 EN EXCEL")
print("="*80)

# Cliente 2 empieza en columna S (columna 19)
# base está en columna U (columna 21)
# fila 17

col_base_cliente2 = 21  # U
fila = 17

celda_base = hoja_std.cell(row=fila, column=col_base_cliente2)
print(f"\nCelda U{fila} (base Cliente 2):")
print(f"  Formula: {celda_base.value}")
print(f"  Valor calculado: {celda_base.value if celda_base.value else 'None'}")

# Verificar N17 (saldo_diario Cliente 1, columna N = 14)
col_saldo_cliente1 = 14  # N
celda_n17 = hoja_std.cell(row=fila, column=col_saldo_cliente1)
print(f"\nCelda N{fila} (saldo_diario Cliente 1):")
print(f"  Formula: {celda_n17.value}")
print(f"  Valor: {celda_n17.value if celda_n17.value else 'None'}")

# Verificar S15 (incremento Cliente 2, columna S = 19)
col_incremento_cliente2 = 19  # S
celda_s15 = hoja_std.cell(row=15, column=col_incremento_cliente2)
print(f"\nCelda S15 (incremento Cliente 2):")
print(f"  Formula: {celda_s15.value}")
print(f"  Valor: {celda_s15.value if celda_s15.value else 'None'}")

# Verificar T16 (decremento Cliente 2, columna T = 20)
col_decremento_cliente2 = 20  # T
celda_t16 = hoja_std.cell(row=16, column=col_decremento_cliente2)
print(f"\nCelda T16 (decremento Cliente 2):")
print(f"  Formula: {celda_t16.value}")
print(f"  Valor: {celda_t16.value if celda_t16.value else 'None'}")

# Verificar si hay alguna otra interpretacion
print(f"\nINTERPRETACION:")
if celda_base.value and isinstance(celda_base.value, str) and celda_base.value.startswith('='):
    formula = celda_base.value
    print(f"  Formula encontrada: {formula}")
    if 'N17' in formula:
        print(f"  La formula SI requiere N17")
        print(f"  Si N17 es 0, la formula siempre devolvera 0")
    else:
        print(f"  La formula NO requiere N17")
else:
    print(f"  No hay formula, es un valor directo")

wb.close()
