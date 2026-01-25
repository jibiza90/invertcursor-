from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("VERIFICACION FORMULAS CLIENTES - DEPENDENCIAS CON DATOS GENERALES")
print("=" * 80)

# Cliente 1 está en columnas K-R (11-18)
# Verificar algunas filas para ver las fórmulas
print("\nCliente 1 - Columnas K-R (11-18):")
print("-" * 80)

filas_ejemplo = [17, 20, 23, 26, 29]

for fila_idx in filas_ejemplo:
    fecha_celda = ws.cell(row=fila_idx, column=4)
    if fecha_celda.value is None:
        continue
    
    print(f"\nFila {fila_idx}:")
    
    # Columnas del cliente 1
    columnas_cliente = {
        11: 'K - Incremento',
        12: 'L - Decremento',
        13: 'M - (no usado)',
        14: 'N - Saldo Diario',
        15: 'O - Benef. € Diario',
        16: 'P - Benef. % Diario',
        17: 'Q - Benef. € Acum',
        18: 'R - Benef. % Acum'
    }
    
    tiene_formulas = False
    for col_idx, nombre_col in columnas_cliente.items():
        celda = ws.cell(row=fila_idx, column=col_idx)
        
        if celda.data_type == 'f' and celda.value:
            tiene_formulas = True
            formula = str(celda.value)
            print(f"  {nombre_col}: {formula[:80]}")
    
    if not tiene_formulas:
        print("  (Sin fórmulas en esta fila)")

# Verificar si hay referencias a columnas generales (E, F, G, H, I, J) en las fórmulas de clientes
print("\n" + "=" * 80)
print("BUSCANDO REFERENCIAS A COLUMNAS GENERALES EN FORMULAS DE CLIENTES")
print("=" * 80)

columnas_generales = ['E', 'F', 'G', 'H', 'I', 'J']
referencias_encontradas = {}

for fila_idx in range(15, min(100, ws.max_row + 1)):
    fecha_celda = ws.cell(row=fila_idx, column=4)
    if fecha_celda.value is None:
        continue
    
    # Verificar columnas del cliente 1 (K-R)
    for col_idx in range(11, 19):
        celda = ws.cell(row=fila_idx, column=col_idx)
        if celda.data_type == 'f' and celda.value:
            formula = str(celda.value)
            
            # Buscar referencias a columnas generales
            for col_gen in columnas_generales:
                if col_gen in formula:
                    if col_gen not in referencias_encontradas:
                        referencias_encontradas[col_gen] = []
                    referencias_encontradas[col_gen].append({
                        'fila': fila_idx,
                        'columna': col_idx,
                        'formula': formula[:100]
                    })

print("\nReferencias encontradas:")
for col_gen, refs in referencias_encontradas.items():
    print(f"\nColumna {col_gen} referenciada en {len(refs)} fórmulas de clientes:")
    for ref in refs[:5]:  # Mostrar primeras 5
        print(f"  Fila {ref['fila']}, Col {ref['columna']}: {ref['formula']}")
