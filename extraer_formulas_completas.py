from openpyxl import load_workbook
import json

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb_formulas.sheetnames:
    ws_formulas = wb_formulas['Diario VIP']
    ws_valores = wb_valores['Diario VIP']
    
    print("=" * 80)
    print("EXTRACCION DE FORMULAS PARA CALCULO AUTOMATICO")
    print("=" * 80)
    
    formulas_data = {
        'datos_generales': {},
        'datos_diarios_generales': {},
        'clientes': {}
    }
    
    # 1. Extraer fórmulas de datos generales (filas 3-6)
    print("\n1. Extrayendo formulas de datos generales...")
    for fila_idx in range(3, 7):
        formulas_fila = {}
        for col_idx in range(5, 11):
            celda = ws_formulas.cell(row=fila_idx, column=col_idx)
            if celda.data_type == 'f' and celda.value:
                formula = str(celda.value)
                col_key = ['imp_inicial', 'imp_final', 'benef_euro', 'benef_porcentaje', 
                          'benef_euro_acum', 'benef_porcentaje_acum'][col_idx - 5]
                formulas_fila[col_key] = formula
                print(f"  Fila {fila_idx}, {col_key}: {formula[:60]}")
        if formulas_fila:
            formulas_data['datos_generales'][fila_idx] = formulas_fila
    
    # 2. Extraer fórmulas de datos diarios generales
    print("\n2. Extrayendo formulas de datos diarios generales...")
    contador = 0
    for fila_idx in range(15, min(200, ws_valores.max_row + 1)):
        fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
        if fecha_valor is None:
            continue
        
        formulas_fila = {}
        for col_idx in range(5, 11):
            celda = ws_formulas.cell(row=fila_idx, column=col_idx)
            if celda.data_type == 'f' and celda.value:
                formula = str(celda.value)
                col_key = ['imp_inicial', 'imp_final', 'benef_euro', 'benef_porcentaje', 
                          'benef_euro_acum', 'benef_porcentaje_acum'][col_idx - 5]
                formulas_fila[col_key] = formula
                if contador < 5:
                    print(f"  Fila {fila_idx}, {col_key}: {formula[:60]}")
        if formulas_fila:
            formulas_data['datos_diarios_generales'][fila_idx] = formulas_fila
            contador += 1
    
    # 3. Extraer fórmulas de clientes (ejemplo para cliente 1)
    print("\n3. Extrayendo formulas de cliente 1 (ejemplo)...")
    # Cliente 1 está en columnas K-R (11-18)
    for fila_idx in range(15, min(100, ws_valores.max_row + 1)):
        fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
        if fecha_valor is None:
            continue
        
        formulas_fila = {}
        # Columnas del cliente 1: K(11), L(12), M(13), N(14), O(15), P(16), Q(17), R(18)
        columnas_cliente = {
            13: 'saldo_diario',  # M (pero en realidad N es saldo diario)
            14: 'saldo_diario',  # N
            15: 'benef_euro_diario',  # O
            16: 'benef_porcentaje_diario',  # P
            17: 'benef_euro_acum',  # Q
            18: 'benef_porcentaje_acum'  # R
        }
        
        for col_idx, col_key in columnas_cliente.items():
            celda = ws_formulas.cell(row=fila_idx, column=col_idx)
            if celda.data_type == 'f' and celda.value:
                formula = str(celda.value)
                formulas_fila[col_key] = formula
                if len(formulas_fila) == 1:  # Solo mostrar primera vez
                    print(f"  Fila {fila_idx}, {col_key}: {formula[:60]}")
        
        if formulas_fila:
            if 'cliente_1' not in formulas_data['clientes']:
                formulas_data['clientes']['cliente_1'] = {}
            formulas_data['clientes']['cliente_1'][fila_idx] = formulas_fila
    
    # Guardar fórmulas en JSON
    with open('formulas.json', 'w', encoding='utf-8') as f:
        json.dump(formulas_data, f, indent=2, ensure_ascii=False)
    
    print("\n" + "=" * 80)
    print("Formulas guardadas en: formulas.json")
    print("=" * 80)
