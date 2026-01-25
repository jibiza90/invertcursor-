from openpyxl import load_workbook
from datetime import datetime

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("VERIFICACION DEPENDENCIAS CLIENTE - 14 ENERO 2026")
print("=" * 80)

# Buscar la fila del 14 de enero 2026
fila_14_enero = None
for fila_idx in range(15, min(200, ws.max_row + 1)):
    fecha_celda = ws.cell(row=fila_idx, column=4).value
    if fecha_celda:
        try:
            if isinstance(fecha_celda, datetime):
                fecha_dt = fecha_celda
            else:
                fecha_dt = datetime.strptime(str(fecha_celda).split()[0], '%Y-%m-%d')
            
            if fecha_dt.year == 2026 and fecha_dt.month == 1 and fecha_dt.day == 14:
                fila_14_enero = fila_idx
                break
        except:
            pass

if fila_14_enero:
    print(f"\nFila encontrada: {fila_14_enero}")
    
    # Verificar datos generales en esa fila
    print("\nDatos generales (columnas E-J):")
    columnas_generales = {
        5: 'E - Imp. Inicial',
        6: 'F - Imp. Final',
        7: 'G - Benef. €',
        8: 'H - Benef. %',
        9: 'I - Benef. € Acum',
        10: 'J - Benef. % Acum'
    }
    
    for col_idx, nombre in columnas_generales.items():
        celda = ws.cell(row=fila_14_enero, column=col_idx)
        tiene_formula = celda.data_type == 'f'
        valor = celda.value
        formula_texto = str(valor)[:80] if tiene_formula else None
        print(f"  {nombre}: valor={ws.cell(row=fila_14_enero, column=col_idx).value}, formula={tiene_formula}")
        if formula_texto:
            print(f"    Formula: {formula_texto}")
    
    # Verificar fórmulas del cliente 1 (columnas K-R)
    print("\nCliente 1 - Fórmulas (columnas K-R):")
    columnas_cliente = {
        11: 'K - Incremento',
        12: 'L - Decremento',
        13: 'M - (intermedio)',
        14: 'N - Saldo Diario',
        15: 'O - Benef. € Diario',
        16: 'P - Benef. % Diario',
        17: 'Q - Benef. € Acum',
        18: 'R - Benef. % Acum'
    }
    
    for col_idx, nombre in columnas_cliente.items():
        celda = ws.cell(row=fila_14_enero, column=col_idx)
        tiene_formula = celda.data_type == 'f'
        valor = celda.value
        if tiene_formula:
            formula_texto = str(valor)
            print(f"  {nombre}: {formula_texto[:100]}")
            
            # Verificar si referencia columnas generales (E-J)
            if any(col in formula_texto for col in ['E', 'F', 'G', 'H', 'I', 'J']):
                print(f"    -> DEPENDE DE DATOS GENERALES")
else:
    print("No se encontró la fila del 14 de enero 2026")
