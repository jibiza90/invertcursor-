#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para verificar qué celdas están bloqueadas/desbloqueadas en la vista general del Excel
"""

import openpyxl

def obtener_nombre_columna(num_col):
    result = ""
    num = num_col
    while num > 0:
        num -= 1
        result = chr(65 + (num % 26)) + result
        num //= 26
    return result

def main():
    print("="*80)
    print("VERIFICANDO VISTA GENERAL EN EXCEL")
    print("="*80)
    
    # Cargar Excel
    wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
    ws = None
    for sheet_name in wb.sheetnames:
        if 'VIP' in sheet_name.upper():
            ws = wb[sheet_name]
            print(f"\n✅ Hoja encontrada: {sheet_name}")
            break
    
    if not ws:
        ws = wb.active
        print(f"\n⚠️ Usando hoja activa: {ws.title}")
    
    # Columnas de la vista general
    columnas_generales = {
        'E': 5,   # IMP. INICIAL
        'F': 6,   # IMP. FINAL
        'G': 7,   # BENEF. €
        'H': 8,   # BENEF. %
        'I': 9,   # BENEF. € (acumulado)
        'J': 10   # BENEF. % (acumulado)
    }
    
    print("\n" + "="*80)
    print("VERIFICANDO FILAS DE DATOS GENERALES (3-6)")
    print("="*80)
    
    for fila in [3, 4, 5, 6]:
        print(f"\nFila {fila}:")
        for col_nombre, col_num in columnas_generales.items():
            celda = ws.cell(row=fila, column=col_num)
            tiene_formula = celda.data_type == 'f' and celda.value is not None
            valor = celda.value
            print(f"  Col {col_nombre} ({col_num}): formula={tiene_formula}, valor={valor}")
    
    print("\n" + "="*80)
    print("VERIFICANDO FILAS DE DATOS DIARIOS (15-30 como muestra)")
    print("="*80)
    
    for fila in range(15, 31):  # Primeras 16 filas de datos diarios
        fecha = ws.cell(row=fila, column=4).value
        print(f"\nFila {fila} (Fecha: {fecha}):")
        for col_nombre, col_num in columnas_generales.items():
            celda = ws.cell(row=fila, column=col_num)
            tiene_formula = celda.data_type == 'f' and celda.value is not None
            valor = celda.value
            estado = "🔒 BLOQUEADA (fórmula)" if tiene_formula else "✏️ EDITABLE"
            print(f"  Col {col_nombre} ({col_num}): {estado}, valor={valor}")
            if tiene_formula and valor:
                print(f"    Fórmula: {valor}")
    
    print("\n" + "="*80)
    print("RESUMEN")
    print("="*80)
    print("Según el usuario:")
    print("- Solo IMP. FINAL (Col F) debería ser EDITABLE")
    print("- Las demás columnas (E, G, H, I, J) deberían estar BLOQUEADAS (tener fórmulas)")
    
    wb.close()

if __name__ == '__main__':
    main()
