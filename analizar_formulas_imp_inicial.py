from openpyxl import load_workbook
import json

# Cargar Excel
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

# Verificar estructura de fórmulas para entender el patrón
print("Analizando estructura de fórmulas imp_inicial:")
print("=" * 80)

# Verificar primeras fechas
for fila in [15, 18, 21, 24]:
    fecha = ws.cell(row=fila, column=4).value
    formula_imp_inicial = ws.cell(row=fila, column=5).value  # Columna E
    
    if formula_imp_inicial and str(formula_imp_inicial).startswith('='):
        print(f"\nFila {fila} (Fecha: {fecha}):")
        print(f"  Formula imp_inicial: {formula_imp_inicial}")
        
        # Extraer referencias
        import re
        refs = re.findall(r'([A-Z]+\d+)', str(formula_imp_inicial))
        print(f"  Referencias: {refs}")
        
        # Verificar qué es AEO
        if 'AEO' in str(formula_imp_inicial):
            # Buscar qué fila es AEO
            aeo_refs = re.findall(r'AEO(\d+)', str(formula_imp_inicial))
            for aeo_fila in aeo_refs:
                aeo_valor = ws.cell(row=int(aeo_fila), column=157).value  # AEO es columna 157
                print(f"  AEO{aeo_fila} (columna 157, fila {aeo_fila}) = {aeo_valor}")
                
                # Verificar si esta fila corresponde a incrementos/decrementos
                # Las filas de incrementos están en las mismas filas que las fechas
                print(f"    Esta fila corresponde a fecha: {ws.cell(row=int(aeo_fila), column=4).value}")

print("\n" + "=" * 80)
print("Verificando relación entre filas de imp_inicial y AEO:")
print("=" * 80)

# Verificar patrón: si imp_inicial está en fila 18, qué AEO usa?
for fila_imp_inicial in [15, 18, 21, 24]:
    fecha = ws.cell(row=fila_imp_inicial, column=4).value
    formula = ws.cell(row=fila_imp_inicial, column=5).value
    
    if formula and 'AEO' in str(formula):
        import re
        aeo_refs = re.findall(r'AEO(\d+)', str(formula))
        for aeo_fila_str in aeo_refs:
            aeo_fila = int(aeo_fila_str)
            fecha_aeo = ws.cell(row=aeo_fila, column=4).value
            
            print(f"\nFila {fila_imp_inicial} (Fecha: {fecha}) usa AEO{aeo_fila}")
            print(f"  AEO{aeo_fila} está en fecha: {fecha_aeo}")
            print(f"  Diferencia de filas: {aeo_fila - fila_imp_inicial}")
            
            # Verificar si AEO corresponde al mismo día o día siguiente
            if fecha == fecha_aeo:
                print(f"  ✅ AEO es del MISMO DÍA")
            else:
                print(f"  ❌ AEO es de OTRO DÍA (diferencia)")
