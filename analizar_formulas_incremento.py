from openpyxl import load_workbook
import re

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
ws_formulas = wb_formulas['Diario VIP']
ws_valores = wb_valores['Diario VIP']

print("=" * 80)
print("ANALISIS DE FORMULAS CON INCREMENTOS - CLIENTE 1")
print("=" * 80)

# Analizar filas 15-20
for fila_idx in range(15, 21):
    fecha = ws_valores.cell(row=fila_idx, column=4).value
    k_valor = ws_valores.cell(row=fila_idx, column=11).value  # K = incremento
    l_valor = ws_valores.cell(row=fila_idx, column=12).value  # L = decremento
    m_valor = ws_valores.cell(row=fila_idx, column=13).value  # M = base
    n_valor = ws_valores.cell(row=fila_idx, column=14).value  # N = saldo_diario
    
    print(f"\n{'='*80}")
    print(f"FILA {fila_idx} - Fecha: {fecha}")
    print(f"  Valores:")
    print(f"    K{fila_idx} (INCREMENTO): {k_valor}")
    print(f"    L{fila_idx} (DECREMENTO): {l_valor}")
    print(f"    M{fila_idx} (BASE): {m_valor}")
    print(f"    N{fila_idx} (SALDO_DIARIO): {n_valor}")
    
    # Verificar fórmulas
    celda_m = ws_formulas.cell(row=fila_idx, column=13)  # M
    celda_n = ws_formulas.cell(row=fila_idx, column=14)  # N
    
    if celda_m.data_type == 'f' and celda_m.value:
        formula_m = str(celda_m.value)
        print(f"\n  FORMULA M{fila_idx} (BASE): {formula_m}")
        referencias = re.findall(r'([A-Z]+\d+)', formula_m)
        print(f"    Referencias: {referencias}")
        for ref in referencias:
            col_letra = re.match(r'([A-Z]+)', ref).group(1)
            fila_num = int(re.match(r'\d+', ref).group(0))
            col_num = 0
            for ch in col_letra:
                col_num = col_num * 26 + (ord(ch) - 64)
            valor_ref = ws_valores.cell(row=fila_num, column=col_num).value
            print(f"      {ref} = {valor_ref}")
    
    if celda_n.data_type == 'f' and celda_n.value:
        formula_n = str(celda_n.value)
        print(f"\n  FORMULA N{fila_idx} (SALDO_DIARIO): {formula_n}")
        referencias = re.findall(r'([A-Z]+\d+)', formula_n)
        print(f"    Referencias: {referencias}")
        for ref in referencias:
            col_letra = re.match(r'([A-Z]+)', ref).group(1)
            fila_num = int(re.match(r'\d+', ref).group(0))
            col_num = 0
            for ch in col_letra:
                col_num = col_num * 26 + (ord(ch) - 64)
            valor_ref = ws_valores.cell(row=fila_num, column=col_num).value
            print(f"      {ref} = {valor_ref}")
