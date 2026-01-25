from openpyxl import load_workbook
import json

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

print("=" * 80)
print("ANALISIS DETALLADO DE COLUMNAS - DIARIO VIP")
print("=" * 80)

if 'Diario VIP' in wb.sheetnames:
    ws = wb['Diario VIP']
    
    # Analizar las primeras filas para entender la estructura
    print("\nFILAS 1-20:")
    for fila_idx in range(1, min(21, ws.max_row + 1)):
        fila = []
        for col_idx in range(1, min(30, ws.max_column + 1)):
            celda = ws.cell(row=fila_idx, column=col_idx)
            valor = celda.value
            if valor is not None:
                fila.append(f"{col_idx}:{str(valor)[:20]}")
        if len(fila) > 0:
            print(f"Fila {fila_idx}: {fila[:15]}")
    
    print("\n" + "=" * 80)
    print("BUSCANDO ESTRUCTURA DE DATOS GENERALES")
    print("=" * 80)
    
    # Buscar fila con "TOTAL" que indica datos generales
    print("\nFilas con TOTAL (datos generales):")
    for fila_idx in range(1, min(15, ws.max_row + 1)):
        fila_str = ''
        valores = []
        for col_idx in range(1, min(30, ws.max_column + 1)):
            celda = ws.cell(row=fila_idx, column=col_idx)
            valor = celda.value
            if valor is not None:
                fila_str += str(valor) + ' '
                valores.append((col_idx, valor))
        
        if 'TOTAL' in fila_str.upper():
            print(f"\nFila {fila_idx}:")
            for col_idx, valor in valores[:15]:
                print(f"  Columna {col_idx}: {valor}")
    
    print("\n" + "=" * 80)
    print("BUSCANDO HEADERS DE CLIENTES")
    print("=" * 80)
    
    # Buscar fila con headers MES, FECHA, etc.
    for fila_idx in range(1, min(20, ws.max_row + 1)):
        fila_str = ''
        headers_encontrados = []
        for col_idx in range(1, min(30, ws.max_column + 1)):
            celda = ws.cell(row=fila_idx, column=col_idx)
            valor = celda.value
            if valor is not None:
                fila_str += str(valor).upper() + ' '
                if str(valor).strip() in ['MES', 'FECHA', 'IMP. INICIAL', 'IMP. FINAL']:
                    headers_encontrados.append((col_idx, str(valor)))
        
        if 'MES' in fila_str and 'FECHA' in fila_str:
            print(f"\nFila {fila_idx} - HEADERS ENCONTRADOS:")
            for col_idx, header in headers_encontrados:
                print(f"  Columna {col_idx}: {header}")
            
            # Mostrar todas las columnas con headers en esta fila
            print(f"\nTodos los headers de la fila {fila_idx}:")
            for col_idx in range(1, min(30, ws.max_column + 1)):
                celda = ws.cell(row=fila_idx, column=col_idx)
                valor = celda.value
                if valor is not None and str(valor).strip() != '':
                    print(f"  Columna {col_idx}: {valor}")
            
            # Mostrar siguientes 5 filas de datos
            print(f"\nPrimeras 5 filas de datos (desde fila {fila_idx + 1}):")
            for i in range(fila_idx + 1, min(fila_idx + 6, ws.max_row + 1)):
                fila_datos = []
                for col_idx in range(1, min(30, ws.max_column + 1)):
                    celda = ws.cell(row=i, column=col_idx)
                    valor = celda.value
                    if valor is not None:
                        fila_datos.append(f"{col_idx}:{str(valor)[:15]}")
                if len(fila_datos) > 0:
                    print(f"  Fila {i}: {fila_datos[:10]}")
            break
    
    print("\n" + "=" * 80)
    print("ANALISIS DE COLUMNAS POR SECCIONES")
    print("=" * 80)
    
    # Analizar columnas agrupadas (parece que hay múltiples grupos de columnas)
    print("\nAnalizando estructura de columnas múltiples...")
    for fila_idx in [13, 14]:
        print(f"\nFila {fila_idx}:")
        grupos_columnas = {}
        for col_idx in range(1, min(50, ws.max_column + 1)):
            celda = ws.cell(row=fila_idx, column=col_idx)
            valor = celda.value
            if valor is not None and str(valor).strip() != '':
                # Identificar grupo (parece que cada X columnas hay un nuevo grupo)
                grupo = (col_idx - 1) // 8  # Aproximadamente cada 8 columnas
                if grupo not in grupos_columnas:
                    grupos_columnas[grupo] = []
                grupos_columnas[grupo].append((col_idx, str(valor)))
        
        for grupo, columnas in sorted(grupos_columnas.items()):
            print(f"  Grupo {grupo}: {columnas[:5]}")
