from openpyxl import load_workbook

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb.sheetnames:
    ws = wb['Diario VIP']
    
    print("=" * 80)
    print("VERIFICACION: LECTURA CORRECTA PARA CLIENTE 1")
    print("=" * 80)
    
    # Para Cliente 1, las columnas son:
    # K = 11 (INCREMENTO)
    # L = 12 (DECREMENTO)
    # N = 14 (FINAL)
    # O = 15 (BENEF. €)
    # P = 16 (BENEF. %)
    # Q = 17 (BENEF. € acumulado)
    # R = 18 (BENEF. % acumulado)
    
    print("\nColumnas para Cliente 1:")
    print("  K (11): INCREMENTO")
    print("  L (12): DECREMENTO")
    print("  N (14): FINAL")
    print("  O (15): BENEF. €")
    print("  P (16): BENEF. %")
    print("  Q (17): BENEF. € acumulado")
    print("  R (18): BENEF. % acumulado")
    
    # Leer todas las filas desde la 15 en adelante
    print("\n" + "=" * 80)
    print("LEYENDO TODAS LAS FILAS (desde fila 15)")
    print("=" * 80)
    
    incrementos = []
    decrementos = []
    saldos_finales = []
    beneficios_diarios = []
    beneficios_diarios_pct = []
    beneficios_acumulados = []
    beneficios_acumulados_pct = []
    
    for fila_idx in range(15, ws.max_row + 1):
        # Leer columna K (11) - INCREMENTO
        inc = ws.cell(row=fila_idx, column=11).value
        if inc is not None:
            try:
                incrementos.append((fila_idx, float(inc)))
            except:
                pass
        
        # Leer columna L (12) - DECREMENTO
        dec = ws.cell(row=fila_idx, column=12).value
        if dec is not None:
            try:
                decrementos.append((fila_idx, float(dec)))
            except:
                pass
        
        # Leer columna N (14) - FINAL
        final = ws.cell(row=fila_idx, column=14).value
        if final is not None:
            try:
                saldos_finales.append((fila_idx, float(final)))
            except:
                pass
        
        # Leer columna O (15) - BENEF. €
        benef_d = ws.cell(row=fila_idx, column=15).value
        if benef_d is not None:
            try:
                beneficios_diarios.append((fila_idx, float(benef_d)))
            except:
                pass
        
        # Leer columna P (16) - BENEF. %
        benef_d_pct = ws.cell(row=fila_idx, column=16).value
        if benef_d_pct is not None:
            try:
                beneficios_diarios_pct.append((fila_idx, float(benef_d_pct)))
            except:
                pass
        
        # Leer columna Q (17) - BENEF. € acumulado
        benef_a = ws.cell(row=fila_idx, column=17).value
        if benef_a is not None:
            try:
                beneficios_acumulados.append((fila_idx, float(benef_a)))
            except:
                pass
        
        # Leer columna R (18) - BENEF. % acumulado
        benef_a_pct = ws.cell(row=fila_idx, column=18).value
        if benef_a_pct is not None:
            try:
                beneficios_acumulados_pct.append((fila_idx, float(benef_a_pct)))
            except:
                pass
    
    # Mostrar resultados
    print("\n" + "=" * 80)
    print("RESULTADOS PARA CLIENTE 1")
    print("=" * 80)
    
    print(f"\nINCREMENTOS (Columna K - suma):")
    suma_inc = sum([v for _, v in incrementos])
    print(f"  Total valores encontrados: {len(incrementos)}")
    print(f"  Valores: {[v for _, v in incrementos[:10]]}...")
    print(f"  SUMA TOTAL: {suma_inc:.2f}")
    
    print(f"\nDECREMENTOS (Columna L - suma):")
    suma_dec = sum([v for _, v in decrementos])
    print(f"  Total valores encontrados: {len(decrementos)}")
    print(f"  Valores: {[v for _, v in decrementos[:10]]}...")
    print(f"  SUMA TOTAL: {suma_dec:.2f}")
    
    print(f"\nSALDO FINAL (Columna N - ultimo valor no cero):")
    print(f"  Total valores encontrados: {len(saldos_finales)}")
    valores_finales = [v for _, v in saldos_finales]
    print(f"  Ultimos 5 valores: {valores_finales[-5:]}")
    # Buscar último no cero
    ultimo_final = None
    for valor in reversed(valores_finales):
        if valor != 0:
            ultimo_final = valor
            break
    if ultimo_final is None and valores_finales:
        ultimo_final = valores_finales[-1]
    print(f"  ULTIMO VALOR: {ultimo_final}")
    
    print(f"\nBENEFICIO DIARIO (Columna O - ultimo valor):")
    print(f"  Total valores encontrados: {len(beneficios_diarios)}")
    if beneficios_diarios:
        print(f"  ULTIMO VALOR: {beneficios_diarios[-1][1]}")
    else:
        print(f"  ULTIMO VALOR: None")
    
    print(f"\nBENEFICIO DIARIO % (Columna P - ultimo valor):")
    print(f"  Total valores encontrados: {len(beneficios_diarios_pct)}")
    if beneficios_diarios_pct:
        print(f"  ULTIMO VALOR: {beneficios_diarios_pct[-1][1]}")
    else:
        print(f"  ULTIMO VALOR: None")
    
    print(f"\nBENEFICIO ACUMULADO (Columna Q - ultimo valor no cero):")
    print(f"  Total valores encontrados: {len(beneficios_acumulados)}")
    valores_acum = [v for _, v in beneficios_acumulados]
    print(f"  Ultimos 5 valores: {valores_acum[-5:]}")
    # Buscar último no cero
    ultimo_acum = None
    for valor in reversed(valores_acum):
        if valor != 0:
            ultimo_acum = valor
            break
    if ultimo_acum is None and valores_acum:
        ultimo_acum = valores_acum[-1]
    print(f"  ULTIMO VALOR: {ultimo_acum}")
    
    print(f"\nBENEFICIO ACUMULADO % (Columna R - ultimo valor no cero):")
    print(f"  Total valores encontrados: {len(beneficios_acumulados_pct)}")
    valores_acum_pct = [v for _, v in beneficios_acumulados_pct]
    print(f"  Ultimos 5 valores: {valores_acum_pct[-5:]}")
    # Buscar último no cero
    ultimo_acum_pct = None
    for valor in reversed(valores_acum_pct):
        if valor != 0:
            ultimo_acum_pct = valor
            break
    if ultimo_acum_pct is None and valores_acum_pct:
        ultimo_acum_pct = valores_acum_pct[-1]
    print(f"  ULTIMO VALOR: {ultimo_acum_pct}")
    
    print("\n" + "=" * 80)
    print("RESUMEN FINAL CLIENTE 1:")
    print("=" * 80)
    print(f"Incrementos: {suma_inc:.2f}")
    print(f"Decrementos: {suma_dec:.2f}")
    print(f"Saldo Final: {ultimo_final}")
    print(f"Beneficio Diario: {beneficios_diarios[-1][1] if beneficios_diarios else None}")
    print(f"Beneficio Diario %: {beneficios_diarios_pct[-1][1] if beneficios_diarios_pct else None}")
    print(f"Beneficio Acumulado: {ultimo_acum}")
    print(f"Beneficio Acumulado %: {ultimo_acum_pct}")
