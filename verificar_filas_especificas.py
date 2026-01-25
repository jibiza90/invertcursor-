from openpyxl import load_workbook
from datetime import datetime

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb_formulas.sheetnames:
    ws_formulas = wb_formulas['Diario VIP']
    ws_valores = wb_valores['Diario VIP']
    
    print("=" * 80)
    print("VERIFICACION DETALLADA FILA POR FILA")
    print("=" * 80)
    
    # Verificar filas específicas alrededor del 13/01/2026
    print("\nFila 15 (01/01/2026):")
    fecha15 = ws_valores.cell(row=15, column=4).value
    print(f"  Fecha: {fecha15}")
    for col in range(5, 11):
        letra = chr(64 + col)
        tiene_form = ws_formulas.cell(row=15, column=col).value and str(ws_formulas.cell(row=15, column=col).value).startswith('=')
        valor = ws_valores.cell(row=15, column=col).value
        print(f"  Col {letra}: valor={valor}, formula={tiene_form}")
    
    print("\nFila 16 (01/01/2026):")
    fecha16 = ws_valores.cell(row=16, column=4).value
    print(f"  Fecha: {fecha16}")
    for col in range(5, 11):
        letra = chr(64 + col)
        tiene_form = ws_formulas.cell(row=16, column=col).value and str(ws_formulas.cell(row=16, column=col).value).startswith('=')
        valor = ws_valores.cell(row=16, column=col).value
        print(f"  Col {letra}: valor={valor}, formula={tiene_form}")
    
    print("\nFila 17 (01/01/2026):")
    fecha17 = ws_valores.cell(row=17, column=4).value
    print(f"  Fecha: {fecha17}")
    for col in range(5, 11):
        letra = chr(64 + col)
        tiene_form = ws_formulas.cell(row=17, column=col).value and str(ws_formulas.cell(row=17, column=col).value).startswith('=')
        valor = ws_valores.cell(row=17, column=col).value
        print(f"  Col {letra}: valor={valor}, formula={tiene_form}")
    
    print("\nFila 27 (05/01/2026):")
    fecha27 = ws_valores.cell(row=27, column=4).value
    print(f"  Fecha: {fecha27}")
    for col in range(5, 11):
        letra = chr(64 + col)
        tiene_form = ws_formulas.cell(row=27, column=col).value and str(ws_formulas.cell(row=27, column=col).value).startswith('=')
        valor = ws_valores.cell(row=27, column=col).value
        print(f"  Col {letra}: valor={valor}, formula={tiene_form}")
    
    print("\nFila 51 (13/01/2026):")
    fecha51 = ws_valores.cell(row=51, column=4).value
    print(f"  Fecha: {fecha51}")
    for col in range(5, 11):
        letra = chr(64 + col)
        tiene_form = ws_formulas.cell(row=51, column=col).value and str(ws_formulas.cell(row=51, column=col).value).startswith('=')
        valor = ws_valores.cell(row=51, column=col).value
        print(f"  Col {letra}: valor={valor}, formula={tiene_form}")
    
    print("\nFila 54 (14/01/2026):")
    fecha54 = ws_valores.cell(row=54, column=4).value
    print(f"  Fecha: {fecha54}")
    for col in range(5, 11):
        letra = chr(64 + col)
        tiene_form = ws_formulas.cell(row=54, column=col).value and str(ws_formulas.cell(row=54, column=col).value).startswith('=')
        valor = ws_valores.cell(row=54, column=col).value
        print(f"  Col {letra}: valor={valor}, formula={tiene_form}")
