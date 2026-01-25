#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para verificar que los cálculos en Diario STD cuadren 100% con las fórmulas del Excel
"""

import json
import openpyxl

# Cargar Excel
wb_formulas = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

ws_std_formulas = None
ws_std_valores = None
for sheet_name in wb_formulas.sheetnames:
    if 'STD' in sheet_name.upper() and 'VIP' not in sheet_name.upper():
        ws_std_formulas = wb_formulas[sheet_name]
        ws_std_valores = wb_valores[sheet_name]
        print(f"Hoja encontrada: {sheet_name}")
        break

# Cargar JSON
with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos = json.load(f)

hoja_std = datos.get('hojas', {}).get('Diario STD')

print("="*80)
print("VERIFICACIÓN DE CÁLCULOS EN DIARIO STD")
print("="*80)

# Analizar fila 17 (día 1, tercera fila del día)
fila = 17

print(f"\nFILA {fila} (Día 1, tercera fila):")
print("="*80)

# Valores en Excel
imp_final_excel = ws_std_valores.cell(row=fila, column=6).value
imp_inicial_excel = ws_std_valores.cell(row=15, column=5).value  # E15
benef_euro_excel = ws_std_valores.cell(row=fila, column=7).value
benef_porcentaje_excel = ws_std_valores.cell(row=fila, column=8).value

print(f"\nValores en Excel:")
print(f"  E15 (imp_inicial día 1): {imp_inicial_excel}")
print(f"  F17 (imp_final día 1): {imp_final_excel}")
print(f"  G17 (benef_euro): {benef_euro_excel}")
print(f"  H17 (benef_porcentaje): {benef_porcentaje_excel}")

# Fórmulas en Excel
formula_benef_euro = ws_std_formulas.cell(row=fila, column=7).value
formula_benef_porcentaje = ws_std_formulas.cell(row=fila, column=8).value

print(f"\nFórmulas en Excel:")
print(f"  G17: {formula_benef_euro}")
print(f"  H17: {formula_benef_porcentaje}")

# Calcular manualmente según las fórmulas
# benef_euro = IF(F17<>0,F17-E15,0)
if imp_final_excel and imp_final_excel != 0:
    benef_euro_calculado = imp_final_excel - (imp_inicial_excel or 0)
else:
    benef_euro_calculado = 0

# benef_porcentaje = IF(G17<>0,((G17/E15)/1),0)
if benef_euro_calculado and benef_euro_calculado != 0 and imp_inicial_excel and imp_inicial_excel != 0:
    benef_porcentaje_calculado = (benef_euro_calculado / imp_inicial_excel) / 1
else:
    benef_porcentaje_calculado = 0

print(f"\nValores calculados manualmente según fórmulas:")
print(f"  benef_euro = IF(F17<>0,F17-E15,0)")
print(f"    = IF({imp_final_excel}<>0,{imp_final_excel}-{imp_inicial_excel or 0},0)")
print(f"    = {benef_euro_calculado}")
print(f"  benef_porcentaje = IF(G17<>0,((G17/E15)/1),0)")
print(f"    = IF({benef_euro_calculado}<>0,(({benef_euro_calculado}/{imp_inicial_excel or 0})/1),0)")
print(f"    = {benef_porcentaje_calculado}")

print(f"\nComparación:")
print(f"  benef_euro Excel: {benef_euro_excel}, Calculado: {benef_euro_calculado}, ¿Coinciden? {abs((benef_euro_excel or 0) - benef_euro_calculado) < 0.01}")
print(f"  benef_porcentaje Excel: {benef_porcentaje_excel}, Calculado: {benef_porcentaje_calculado}, ¿Coinciden? {abs((benef_porcentaje_excel or 0) - benef_porcentaje_calculado) < 0.0001}")

wb_formulas.close()
wb_valores.close()
