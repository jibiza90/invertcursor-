#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificar en Excel qué es realmente N17 y cómo debería calcularse la base del Cliente 2
"""

import openpyxl

wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)

ws_std = None
for sheet_name in wb.sheetnames:
    if 'STD' in sheet_name.upper() and 'VIP' not in sheet_name.upper():
        ws_std = wb[sheet_name]
        print(f"Hoja encontrada: {sheet_name}")
        break

print("="*80)
print("VERIFICACIÓN DE FÓRMULAS EN EXCEL - CLIENTE 2 FILA 17")
print("="*80)

fila = 17

# Ver fórmula de base del Cliente 2 (U17, columna 21)
print(f"\nFila {fila}:")

# Cliente 2: U=21 (base), V=22 (saldo_diario), W=23 (beneficio_diario), X=24 (beneficio_diario_pct)
celda_base = ws_std.cell(row=fila, column=21)  # U17
if celda_base.data_type == 'f' and celda_base.value:
    formula_base = str(celda_base.value)
    print(f"U17 (base Cliente 2): {formula_base}")
    
    # Extraer referencias
    import re
    referencias = re.findall(r'([A-Z]+\d+)', formula_base)
    print(f"  Referencias: {referencias}")
    
    # Verificar cada referencia
    for ref in referencias:
        col_letra = re.match(r'([A-Z]+)', ref).group(1)
        fila_num = int(re.match(r'\d+', ref).group())
        col_num = openpyxl.utils.column_index_from_string(col_letra)
        
        celda_ref = ws_std.cell(row=fila_num, column=col_num)
        if celda_ref.data_type == 'f' and celda_ref.value:
            print(f"    {ref} (col {col_num}): fórmula = {celda_ref.value}")
        else:
            valor = celda_ref.value
            print(f"    {ref} (col {col_num}): valor = {valor}")

# Verificar qué es N17 realmente
print(f"\nN17 (columna 14) específicamente:")
celda_n17 = ws_std.cell(row=fila, column=14)
if celda_n17.data_type == 'f' and celda_n17.value:
    print(f"  Fórmula: {celda_n17.value}")
else:
    print(f"  Valor: {celda_n17.value}")

# Verificar si hay otras fórmulas que usen S15 o T16
print(f"\nBuscando otras fórmulas que usen S15 o T16:")
for fila_check in range(15, 30):
    for col_check in range(19, 27):  # Columnas S-Z (Cliente 2)
        celda = ws_std.cell(row=fila_check, column=col_check)
        if celda.data_type == 'f' and celda.value:
            formula = str(celda.value)
            if 'S15' in formula or 'T16' in formula:
                col_letra = openpyxl.utils.get_column_letter(col_check)
                print(f"  {col_letra}{fila_check}: {formula}")

wb.close()
