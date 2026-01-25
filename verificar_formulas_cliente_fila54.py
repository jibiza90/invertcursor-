from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("VERIFICACION FORMULAS CLIENTE - FILA 54 Y CERCANAS")
print("=" * 80)

# Verificar filas alrededor de la 54
filas_verificar = [50, 51, 52, 53, 54, 55, 56, 57]

for fila_idx in filas_verificar:
    fecha_celda = ws.cell(row=fila_idx, column=4).value
    if fecha_celda is None:
        continue
    
    print(f"\nFila {fila_idx}:")
    
    # Verificar si tiene fórmulas en columnas generales G, H, I, J
    tiene_formulas_generales = False
    for col_idx in [7, 8, 9, 10]:  # G, H, I, J
        celda = ws.cell(row=fila_idx, column=col_idx)
        if celda.data_type == 'f' and celda.value:
            tiene_formulas_generales = True
            col_nombre = ['G', 'H', 'I', 'J'][col_idx - 7]
            print(f"  {col_nombre}: {str(celda.value)[:80]}")
    
    # Verificar fórmulas del cliente 1 (columnas N, O, P, Q, R)
    tiene_formulas_cliente = False
    columnas_cliente = {
        14: 'N - Saldo Diario',
        15: 'O - Benef. € Diario',
        16: 'P - Benef. % Diario',
        17: 'Q - Benef. € Acum',
        18: 'R - Benef. % Acum'
    }
    
    for col_idx, nombre in columnas_cliente.items():
        celda = ws.cell(row=fila_idx, column=col_idx)
        if celda.data_type == 'f' and celda.value:
            tiene_formulas_cliente = True
            formula = str(celda.value)
            print(f"  {nombre}: {formula[:80]}")
            
            # Verificar si referencia columnas generales
            if any(col in formula for col in ['E', 'F', 'G', 'H', 'I', 'J']):
                print(f"    -> DEPENDE DE DATOS GENERALES")
    
    if not tiene_formulas_generales and not tiene_formulas_cliente:
        print("  (Sin fórmulas)")
