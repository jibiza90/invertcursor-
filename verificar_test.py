from openpyxl import load_workbook
import json

print("Iniciando verificacion...")

try:
    wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
    ws_formulas = wb_formulas['Diario VIP']
    print("Excel cargado OK")
except Exception as e:
    print(f"Error cargando Excel: {e}")
    exit(1)

try:
    with open('datos_completos.json', 'r', encoding='utf-8') as f:
        datos_json = json.load(f)
    hoja_json = datos_json['hojas']['Diario VIP']
    cliente1 = next((c for c in hoja_json.get('clientes', []) if c.get('numero_cliente') == 1), None)
    print("JSON cargado OK")
except Exception as e:
    print(f"Error cargando JSON: {e}")
    exit(1)

if not cliente1:
    print("Cliente 1 no encontrado")
    exit(1)

datos_diarios = cliente1.get('datos_diarios', [])
print(f"Encontradas {len(datos_diarios)} filas del cliente 1")

# Verificar fila 17 (que tiene formulas)
fila_verificar = 17
fila_data = next((d for d in datos_diarios if d['fila'] == fila_verificar), None)

if fila_data:
    print(f"\nVerificando Fila {fila_verificar}:")
    print(f"  Fecha: {fila_data.get('fecha')}")
    
    # Verificar columna M (base) - columna 13
    celda_m = ws_formulas.cell(row=fila_verificar, column=13)
    formula_excel_m = str(celda_m.value) if celda_m.data_type == 'f' else None
    formula_json_m = fila_data.get('formulas', {}).get('base')
    bloqueada_m = fila_data.get('bloqueadas', {}).get('base', False)
    
    print(f"\n  Columna M (base):")
    print(f"    Excel tiene formula: {formula_excel_m is not None}")
    if formula_excel_m:
        print(f"    Excel: {formula_excel_m}")
    print(f"    JSON tiene formula: {formula_json_m is not None}")
    if formula_json_m:
        print(f"    JSON: {formula_json_m}")
    print(f"    Bloqueada en JSON: {bloqueada_m}")
    
    if formula_excel_m and formula_json_m:
        if formula_excel_m == formula_json_m:
            print(f"    ✅ Formulas coinciden")
        else:
            print(f"    ❌ Formulas DIFERENTES")
    elif formula_excel_m and not formula_json_m:
        print(f"    ❌ Excel tiene formula pero JSON NO")
    elif not formula_excel_m and formula_json_m:
        print(f"    ❌ JSON tiene formula pero Excel NO")
    
    if formula_excel_m and not bloqueada_m:
        print(f"    ❌ Tiene formula pero NO esta bloqueada")
    
    # Verificar columna N (saldo_diario) - columna 14
    celda_n = ws_formulas.cell(row=fila_verificar, column=14)
    formula_excel_n = str(celda_n.value) if celda_n.data_type == 'f' else None
    formula_json_n = fila_data.get('formulas', {}).get('saldo_diario')
    bloqueada_n = fila_data.get('bloqueadas', {}).get('saldo_diario', False)
    
    print(f"\n  Columna N (saldo_diario):")
    print(f"    Excel tiene formula: {formula_excel_n is not None}")
    if formula_excel_n:
        print(f"    Excel: {formula_excel_n}")
    print(f"    JSON tiene formula: {formula_json_n is not None}")
    if formula_json_n:
        print(f"    JSON: {formula_json_n}")
    print(f"    Bloqueada en JSON: {bloqueada_n}")
    
    if formula_excel_n and formula_json_n:
        if formula_excel_n == formula_json_n:
            print(f"    ✅ Formulas coinciden")
        else:
            print(f"    ❌ Formulas DIFERENTES")
    elif formula_excel_n and not formula_json_n:
        print(f"    ❌ Excel tiene formula pero JSON NO")
    
    if formula_excel_n and not bloqueada_n:
        print(f"    ❌ Tiene formula pero NO esta bloqueada")

print("\nVerificacion completada")
