"""
Extractor de datos v3 - Consolida las 3 filas por fecha en una sola entrada
El archivo Excel tiene 3 filas por día:
- Fila 1: imp_inicial con fórmula
- Fila 2: auxiliar (vacía)  
- Fila 3: imp_final con valor, beneficios con fórmulas

Esta versión consolida todo en una entrada, manteniendo las referencias a la fila 3 para las fórmulas.
"""
from openpyxl import load_workbook
import json
from datetime import datetime, date

# Cargar dos veces: una para fórmulas, otra para valores
print("Cargando archivo Excel...")
wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
print("Archivo cargado.")

def tiene_formula(ws_formulas, fila, columna):
    """Verifica si una celda tiene fórmula"""
    try:
        celda = ws_formulas.cell(row=fila, column=columna)
        if celda.data_type == 'f':
            return True
        if celda.value and str(celda.value).strip().startswith('='):
            return True
        return False
    except:
        return False

def obtener_formula(ws_formulas, fila, columna):
    """Obtiene la fórmula de una celda"""
    try:
        celda = ws_formulas.cell(row=fila, column=columna)
        if celda.value and str(celda.value).strip().startswith('='):
            return str(celda.value)
        return None
    except:
        return None

def procesar_hoja(nombre_hoja):
    """Procesa una hoja completa"""
    if nombre_hoja not in wb_formulas.sheetnames:
        return None
    
    ws_formulas = wb_formulas[nombre_hoja]
    ws_valores = wb_valores[nombre_hoja]
    
    resultado = {
        'nombre': nombre_hoja,
        'datos_generales': [],
        'datos_diarios_generales': [],
        'clientes': []
    }
    
    print(f"\n{'='*70}")
    print(f"Procesando: {nombre_hoja}")
    print('='*70)
    
    # 1. DATOS GENERALES (filas 3-6)
    print("\n1. Datos generales (filas 3-6)...")
    for fila_idx in [3, 4, 5, 6]:
        fila_data = {
            'fila': fila_idx,
            'fecha': ws_valores.cell(fila_idx, 4).value,
            'imp_inicial': ws_valores.cell(fila_idx, 5).value,
            'imp_final': ws_valores.cell(fila_idx, 6).value,
            'benef_euro': ws_valores.cell(fila_idx, 7).value,
            'benef_porcentaje': ws_valores.cell(fila_idx, 8).value,
            'benef_euro_acum': ws_valores.cell(fila_idx, 9).value,
            'benef_porcentaje_acum': ws_valores.cell(fila_idx, 10).value,
            'bloqueadas': {},
            'formulas': {}
        }
        
        # Verificar fórmulas
        cols_map = [('imp_inicial', 5), ('imp_final', 6), ('benef_euro', 7), 
                    ('benef_porcentaje', 8), ('benef_euro_acum', 9), ('benef_porcentaje_acum', 10)]
        for campo, col in cols_map:
            if tiene_formula(ws_formulas, fila_idx, col):
                fila_data['bloqueadas'][campo] = True
                formula = obtener_formula(ws_formulas, fila_idx, col)
                if formula:
                    fila_data['formulas'][campo] = formula
            else:
                fila_data['bloqueadas'][campo] = False
        
        resultado['datos_generales'].append(fila_data)
    print(f"   {len(resultado['datos_generales'])} filas")
    
    # 2. DATOS DIARIOS GENERALES - CONSOLIDADOS (3 filas -> 1 entrada)
    # Patron: fila1=imp_inicial con formula, fila2=auxiliar, fila3=imp_final y beneficios
    print("\n2. Datos diarios generales (consolidados)...")
    
    fila_idx = 15
    while fila_idx <= ws_valores.max_row - 2:  # -2 porque leemos grupos de 3
        fecha = ws_valores.cell(fila_idx, 4).value
        if fecha is None:
            fila_idx += 1
            continue
        
        # Verificar que es un grupo de 3 filas con la misma fecha
        fecha2 = ws_valores.cell(fila_idx + 1, 4).value
        fecha3 = ws_valores.cell(fila_idx + 2, 4).value
        
        if fecha2 == fecha and fecha3 == fecha:
            # Es un grupo de 3 filas
            fila1 = fila_idx
            fila2 = fila_idx + 1
            fila3 = fila_idx + 2
            
            # La fila de referencia para las fórmulas es la tercera (fila3)
            # Esto es importante porque las fórmulas usan referencias como F{fila3}
            fila_data = {
                'fila': fila3,  # USAR FILA 3 como referencia principal
                'fila_inicial': fila1,  # Guardar referencia a fila 1 para imp_inicial
                'fecha': fecha,
                # imp_inicial viene de fila 1
                'imp_inicial': ws_valores.cell(fila1, 5).value,
                # imp_final viene de fila 3
                'imp_final': ws_valores.cell(fila3, 6).value,
                # beneficios vienen de fila 3
                'benef_euro': ws_valores.cell(fila3, 7).value,
                'benef_porcentaje': ws_valores.cell(fila3, 8).value,
                'benef_euro_acum': ws_valores.cell(fila3, 9).value,
                'benef_porcentaje_acum': ws_valores.cell(fila3, 10).value,
                'bloqueadas': {},
                'formulas': {}
            }
            
            # Bloquear imp_inicial si fila1 tiene fórmula
            if tiene_formula(ws_formulas, fila1, 5):
                fila_data['bloqueadas']['imp_inicial'] = True
                formula = obtener_formula(ws_formulas, fila1, 5)
                if formula:
                    fila_data['formulas']['imp_inicial'] = formula
            else:
                fila_data['bloqueadas']['imp_inicial'] = False
            
            # imp_final en fila3 normalmente NO tiene fórmula (es editable)
            if tiene_formula(ws_formulas, fila3, 6):
                fila_data['bloqueadas']['imp_final'] = True
                formula = obtener_formula(ws_formulas, fila3, 6)
                if formula:
                    fila_data['formulas']['imp_final'] = formula
            else:
                fila_data['bloqueadas']['imp_final'] = False
            
            # Beneficios en fila3 tienen fórmulas
            cols_beneficios = [
                ('benef_euro', 7), 
                ('benef_porcentaje', 8), 
                ('benef_euro_acum', 9), 
                ('benef_porcentaje_acum', 10)
            ]
            for campo, col in cols_beneficios:
                if tiene_formula(ws_formulas, fila3, col):
                    fila_data['bloqueadas'][campo] = True
                    formula = obtener_formula(ws_formulas, fila3, col)
                    if formula:
                        fila_data['formulas'][campo] = formula
                else:
                    fila_data['bloqueadas'][campo] = False
            
            resultado['datos_diarios_generales'].append(fila_data)
            fila_idx += 3  # Saltar al siguiente grupo
        else:
            # No es un grupo de 3, procesar como fila individual
            fila_data = {
                'fila': fila_idx,
                'fecha': fecha,
                'imp_inicial': ws_valores.cell(fila_idx, 5).value,
                'imp_final': ws_valores.cell(fila_idx, 6).value,
                'benef_euro': ws_valores.cell(fila_idx, 7).value,
                'benef_porcentaje': ws_valores.cell(fila_idx, 8).value,
                'benef_euro_acum': ws_valores.cell(fila_idx, 9).value,
                'benef_porcentaje_acum': ws_valores.cell(fila_idx, 10).value,
                'bloqueadas': {},
                'formulas': {}
            }
            
            cols_map = [('imp_inicial', 5), ('imp_final', 6), ('benef_euro', 7), 
                        ('benef_porcentaje', 8), ('benef_euro_acum', 9), ('benef_porcentaje_acum', 10)]
            for campo, col in cols_map:
                if tiene_formula(ws_formulas, fila_idx, col):
                    fila_data['bloqueadas'][campo] = True
                    formula = obtener_formula(ws_formulas, fila_idx, col)
                    if formula:
                        fila_data['formulas'][campo] = formula
                else:
                    fila_data['bloqueadas'][campo] = False
            
            resultado['datos_diarios_generales'].append(fila_data)
            fila_idx += 1
    
    print(f"   {len(resultado['datos_diarios_generales'])} registros consolidados")
    
    # Mostrar ejemplo para verificar
    if len(resultado['datos_diarios_generales']) > 12:
        ejemplo = resultado['datos_diarios_generales'][12]  # Día 13 (enero 13)
        print(f"   Ejemplo fila {ejemplo['fila']}: fecha={ejemplo['fecha']}")
        print(f"      imp_inicial={ejemplo['imp_inicial']}, imp_final={ejemplo['imp_final']}")
        print(f"      benef_euro={ejemplo['benef_euro']}")
        print(f"      formulas={list(ejemplo['formulas'].keys())}")
    
    # 3. IDENTIFICAR CLIENTES
    print("\n3. Identificando clientes...")
    clientes_info = []
    
    # Buscar números de cliente en fila 13
    for col in range(11, ws_valores.max_column + 1):
        valor = ws_valores.cell(13, col).value
        if valor is not None:
            try:
                num = int(valor)
                if num > 0:
                    clientes_info.append({
                        'numero': num,
                        'col_inicio': col,
                        'cols': {
                            'incremento': col,      # K, S, AA...
                            'decremento': col + 1,  # L, T, AB...
                            'base': col + 2,        # M, U, AC...
                            'saldo_diario': col + 3, # N, V, AD...
                            'beneficio_diario': col + 4, # O, W, AE...
                            'beneficio_diario_pct': col + 5, # P, X, AF...
                            'beneficio_acumulado': col + 6, # Q, Y, AG...
                            'beneficio_acumulado_pct': col + 7  # R, Z, AH...
                        }
                    })
            except:
                pass
    
    print(f"   {len(clientes_info)} clientes encontrados")
    
    # 4. EXTRAER DATOS POR CLIENTE - TAMBIÉN CONSOLIDADOS
    print("\n4. Extrayendo datos por cliente (consolidados)...")
    
    for cliente_info in clientes_info:
        cliente_data = {
            'numero_cliente': cliente_info['numero'],
            'columna_inicio': cliente_info['col_inicio'],
            'incrementos_total': 0,
            'decrementos_total': 0,
            'saldo_actual': None,
            'datos_diarios': []
        }
        
        cols = cliente_info['cols']
        incrementos = []
        decrementos = []
        saldos = []
        
        fila_idx = 15
        while fila_idx <= ws_valores.max_row - 2:
            fecha = ws_valores.cell(fila_idx, 4).value
            if fecha is None:
                fila_idx += 1
                continue
            
            # Verificar grupo de 3 filas
            fecha2 = ws_valores.cell(fila_idx + 1, 4).value
            fecha3 = ws_valores.cell(fila_idx + 2, 4).value
            
            if fecha2 == fecha and fecha3 == fecha:
                fila1 = fila_idx
                fila3 = fila_idx + 2
                
                # Consolidar datos del cliente
                # incremento y decremento pueden estar en cualquier fila, pero típicamente en fila1 o fila3
                inc1 = ws_valores.cell(fila1, cols['incremento']).value
                inc3 = ws_valores.cell(fila3, cols['incremento']).value
                dec1 = ws_valores.cell(fila1, cols['decremento']).value
                dec3 = ws_valores.cell(fila3, cols['decremento']).value
                
                incremento = inc1 if inc1 else inc3
                decremento = dec1 if dec1 else dec3
                
                dato = {
                    'fila': fila3,  # Usar fila3 como referencia
                    'fecha': fecha,
                    'incremento': incremento,
                    'decremento': decremento,
                    'base': ws_valores.cell(fila3, cols['base']).value,
                    'saldo_diario': ws_valores.cell(fila3, cols['saldo_diario']).value,
                    'beneficio_diario': ws_valores.cell(fila3, cols['beneficio_diario']).value,
                    'beneficio_diario_pct': ws_valores.cell(fila3, cols['beneficio_diario_pct']).value,
                    'beneficio_acumulado': ws_valores.cell(fila3, cols['beneficio_acumulado']).value,
                    'beneficio_acumulado_pct': ws_valores.cell(fila3, cols['beneficio_acumulado_pct']).value,
                    'bloqueadas': {},
                    'formulas': {}
                }
                
                # Verificar fórmulas en fila3
                campos_orden = [
                    ('base', cols['base']),
                    ('beneficio_diario_pct', cols['beneficio_diario_pct']),
                    ('beneficio_diario', cols['beneficio_diario']),
                    ('saldo_diario', cols['saldo_diario']),
                    ('beneficio_acumulado', cols['beneficio_acumulado']),
                    ('beneficio_acumulado_pct', cols['beneficio_acumulado_pct'])
                ]
                
                for campo, col in campos_orden:
                    if tiene_formula(ws_formulas, fila3, col):
                        dato['bloqueadas'][campo] = True
                        formula = obtener_formula(ws_formulas, fila3, col)
                        if formula:
                            dato['formulas'][campo] = formula
                    else:
                        dato['bloqueadas'][campo] = False
                
                cliente_data['datos_diarios'].append(dato)
                
                # Acumular incrementos/decrementos (solo 2026)
                if fecha:
                    try:
                        fecha_dt = fecha if isinstance(fecha, (datetime, date)) else None
                        if fecha_dt and fecha_dt.year == 2026:
                            if incremento:
                                incrementos.append(float(incremento))
                            if decremento:
                                decrementos.append(float(decremento))
                    except:
                        pass
                
                if dato['saldo_diario'] and dato['saldo_diario'] != 0:
                    saldos.append((fila3, dato['saldo_diario']))
                
                fila_idx += 3
            else:
                # Fila individual
                dato = {
                    'fila': fila_idx,
                    'fecha': fecha,
                    'incremento': ws_valores.cell(fila_idx, cols['incremento']).value,
                    'decremento': ws_valores.cell(fila_idx, cols['decremento']).value,
                    'base': ws_valores.cell(fila_idx, cols['base']).value,
                    'saldo_diario': ws_valores.cell(fila_idx, cols['saldo_diario']).value,
                    'beneficio_diario': ws_valores.cell(fila_idx, cols['beneficio_diario']).value,
                    'beneficio_diario_pct': ws_valores.cell(fila_idx, cols['beneficio_diario_pct']).value,
                    'beneficio_acumulado': ws_valores.cell(fila_idx, cols['beneficio_acumulado']).value,
                    'beneficio_acumulado_pct': ws_valores.cell(fila_idx, cols['beneficio_acumulado_pct']).value,
                    'bloqueadas': {},
                    'formulas': {}
                }
                
                campos_orden = [
                    ('base', cols['base']),
                    ('beneficio_diario_pct', cols['beneficio_diario_pct']),
                    ('beneficio_diario', cols['beneficio_diario']),
                    ('saldo_diario', cols['saldo_diario']),
                    ('beneficio_acumulado', cols['beneficio_acumulado']),
                    ('beneficio_acumulado_pct', cols['beneficio_acumulado_pct'])
                ]
                
                for campo, col in campos_orden:
                    if tiene_formula(ws_formulas, fila_idx, col):
                        dato['bloqueadas'][campo] = True
                        formula = obtener_formula(ws_formulas, fila_idx, col)
                        if formula:
                            dato['formulas'][campo] = formula
                    else:
                        dato['bloqueadas'][campo] = False
                
                cliente_data['datos_diarios'].append(dato)
                
                if fecha:
                    try:
                        fecha_dt = fecha if isinstance(fecha, (datetime, date)) else None
                        if fecha_dt and fecha_dt.year == 2026:
                            if dato['incremento']:
                                incrementos.append(float(dato['incremento']))
                            if dato['decremento']:
                                decrementos.append(float(dato['decremento']))
                    except:
                        pass
                
                if dato['saldo_diario'] and dato['saldo_diario'] != 0:
                    saldos.append((fila_idx, dato['saldo_diario']))
                
                fila_idx += 1
        
        cliente_data['incrementos_total'] = sum(incrementos)
        cliente_data['decrementos_total'] = sum(decrementos)
        if saldos:
            cliente_data['saldo_actual'] = saldos[-1][1]
            cliente_data['saldo_actual_fila'] = saldos[-1][0]
        
        resultado['clientes'].append(cliente_data)
        
        if cliente_info['numero'] <= 3:
            print(f"   Cliente {cliente_info['numero']}: {len(cliente_data['datos_diarios'])} registros, saldo={cliente_data['saldo_actual']}")
    
    # Ordenar clientes
    resultado['clientes'].sort(key=lambda x: x['numero_cliente'])
    
    return resultado

# Procesar hojas
resultado_final = {
    'fecha_extraccion': datetime.now().isoformat(),
    'hojas': {}
}

for hoja in ['Diario STD', 'Diario VIP']:
    if hoja in wb_formulas.sheetnames:
        resultado_final['hojas'][hoja] = procesar_hoja(hoja)

# Guardar
with open('datos_completos.json', 'w', encoding='utf-8') as f:
    json.dump(resultado_final, f, ensure_ascii=False, indent=2, default=str)

print(f"\n{'='*70}")
print("Datos guardados en: datos_completos.json")
print('='*70)

# Verificación
print("\nVerificacion de datos consolidados:")
vip = resultado_final['hojas']['Diario VIP']
print(f"  Total registros diarios generales: {len(vip['datos_diarios_generales'])}")
if len(vip['datos_diarios_generales']) > 12:
    dia13 = vip['datos_diarios_generales'][12]
    print(f"  Dia 13/01 - fila={dia13['fila']}, imp_final={dia13['imp_final']}")
    print(f"  Formulas: {list(dia13['formulas'].keys())}")
    
if len(vip['clientes']) >= 1:
    c1 = vip['clientes'][0]
    print(f"  Cliente 1 - {len(c1['datos_diarios'])} registros")
    if len(c1['datos_diarios']) > 12:
        c1_dia13 = c1['datos_diarios'][12]
        print(f"  Cliente 1 dia 13/01 - fila={c1_dia13['fila']}")
        print(f"  Formulas cliente: {list(c1_dia13['formulas'].keys())}")
