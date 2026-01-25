from openpyxl import load_workbook
import json
from datetime import datetime

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

def procesar_hoja(nombre_hoja):
    """Procesa una hoja completa respetando la estructura de columnas"""
    if nombre_hoja not in wb.sheetnames:
        return None
    
    ws = wb[nombre_hoja]
    
    resultado = {
        'nombre': nombre_hoja,
        'datos_generales': [],
        'datos_clientes': []
    }
    
    # 1. EXTRAER DATOS GENERALES (filas 3-6)
    # Las filas 3-6 contienen los totales/general
    print(f"\nProcesando datos generales de {nombre_hoja}...")
    
    filas_generales = [3, 4, 5, 6]
    for fila_idx in filas_generales:
        fila_data = {
            'fila': fila_idx,
            'columnas': {}
        }
        
        # Leer todas las columnas con datos
        for col_idx in range(1, min(100, ws.max_column + 1)):
            celda = ws.cell(row=fila_idx, column=col_idx)
            valor = celda.value
            
            if valor is not None:
                # Identificar tipo de dato por posición y contenido
                fila_data['columnas'][col_idx] = valor
        
        if len(fila_data['columnas']) > 0:
            resultado['datos_generales'].append(fila_data)
    
    print(f"  Datos generales: {len(resultado['datos_generales'])} filas")
    
    # 2. EXTRAER HEADERS DE CLIENTES (fila 14)
    # La fila 14 contiene los headers de las columnas
    print(f"\nProcesando headers de clientes...")
    
    headers = {}
    fila_header = 14
    
    for col_idx in range(1, ws.max_column + 1):
        celda = ws.cell(row=fila_header, column=col_idx)
        valor = celda.value
        
        if valor is not None and str(valor).strip() != '':
            headers[col_idx] = str(valor).strip()
    
    print(f"  Headers encontrados: {len(headers)} columnas")
    
    # 3. EXTRAER DATOS DE CLIENTES (desde fila 15)
    print(f"\nProcesando datos de clientes...")
    
    # Las columnas 3-6 son comunes a todos los clientes (MES, FECHA, IMP. INICIAL, IMP. FINAL)
    # Las demás columnas pueden variar
    
    cliente_actual = None
    fila_inicio_cliente = None
    
    for fila_idx in range(15, ws.max_row + 1):
        # Leer columna MES (columna 3)
        celda_mes = ws.cell(row=fila_idx, column=3)
        mes_valor = celda_mes.value
        
        # Si hay un MES válido, es el inicio de un nuevo registro de cliente
        if mes_valor is not None:
            try:
                mes_num = int(mes_valor)
                if mes_num > 0:
                    # Si había un cliente anterior, guardarlo
                    if cliente_actual is not None:
                        resultado['datos_clientes'].append(cliente_actual)
                    
                    # Crear nuevo cliente
                    cliente_actual = {
                        'fila': fila_idx,
                        'datos': {}
                    }
                    fila_inicio_cliente = fila_idx
            except:
                pass
        
        # Si tenemos un cliente activo, leer sus datos
        if cliente_actual is not None:
            # Leer todas las columnas con headers
            for col_idx, header in headers.items():
                celda = ws.cell(row=fila_idx, column=col_idx)
                valor = celda.value
                
                # Determinar si está bloqueada
                bloqueada = header in ['MES', 'FECHA', 'IMP. INICIAL']
                
                # Si la columna ya existe, combinar valores (puede haber múltiples filas)
                if header in cliente_actual['datos']:
                    # Si el nuevo valor no es None y es diferente, actualizar
                    if valor is not None and cliente_actual['datos'][header]['valor'] != valor:
                        # Para algunas columnas, tomar el último valor no vacío
                        if valor != '' and valor is not None:
                            cliente_actual['datos'][header] = {
                                'valor': valor,
                                'bloqueada': bloqueada,
                                'columna': col_idx,
                                'fila': fila_idx
                            }
                else:
                    # Nueva columna
                    cliente_actual['datos'][header] = {
                        'valor': valor,
                        'bloqueada': bloqueada,
                        'columna': col_idx,
                        'fila': fila_idx
                    }
            
            # Si la siguiente fila tiene un MES diferente o no tiene MES, finalizar este cliente
            if fila_idx < ws.max_row:
                siguiente_mes = ws.cell(row=fila_idx + 1, column=3).value
                if siguiente_mes is not None:
                    try:
                        siguiente_mes_num = int(siguiente_mes)
                        mes_actual = int(ws.cell(row=fila_inicio_cliente, column=3).value)
                        if siguiente_mes_num != mes_actual:
                            # Nuevo cliente, guardar el actual
                            resultado['datos_clientes'].append(cliente_actual)
                            cliente_actual = None
                    except:
                        pass
    
    # Guardar último cliente
    if cliente_actual is not None:
        resultado['datos_clientes'].append(cliente_actual)
    
    print(f"  Clientes encontrados: {len(resultado['datos_clientes'])}")
    
    # Agregar headers al resultado
    resultado['headers'] = {str(k): v for k, v in headers.items()}
    
    return resultado

# Procesar ambas hojas
resultado_completo = {
    'fecha_extraccion': datetime.now().isoformat(),
    'hojas': {}
}

print("=" * 80)
print("EXTRACCION DE DATOS DEL EXCEL")
print("=" * 80)

# Procesar Diario STD
if 'Diario STD' in wb.sheetnames:
    resultado_completo['hojas']['Diario STD'] = procesar_hoja('Diario STD')

# Procesar Diario VIP
if 'Diario VIP' in wb.sheetnames:
    resultado_completo['hojas']['Diario VIP'] = procesar_hoja('Diario VIP')

# Guardar resultado
with open('datos_completos.json', 'w', encoding='utf-8') as f:
    json.dump(resultado_completo, f, ensure_ascii=False, indent=2, default=str)

print("\n" + "=" * 80)
print(f"Datos guardados en: datos_completos.json")
print(f"Total de hojas procesadas: {len(resultado_completo['hojas'])}")
print("=" * 80)

# Mostrar resumen
for nombre_hoja, datos in resultado_completo['hojas'].items():
    if datos:
        print(f"\n{nombre_hoja}:")
        print(f"  - Datos generales: {len(datos['datos_generales'])} filas")
        print(f"  - Clientes: {len(datos['datos_clientes'])} registros")
        if datos['datos_clientes']:
            primer_cliente = datos['datos_clientes'][0]
            print(f"  - Campos del primer cliente: {len(primer_cliente['datos'])} columnas")
