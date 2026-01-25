from openpyxl import load_workbook
from datetime import datetime

wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)
ws = wb_valores['Diario VIP']

print("=" * 80)
print("BUSCANDO FILA DEL 14 ENERO 2026")
print("=" * 80)

# Buscar la fila del 14 de enero 2026
for fila_idx in range(15, min(200, ws.max_row + 1)):
    fecha_celda = ws.cell(row=fila_idx, column=4).value
    if fecha_celda:
        try:
            fecha_dt = None
            if isinstance(fecha_celda, datetime):
                fecha_dt = fecha_celda
            elif isinstance(fecha_celda, str):
                try:
                    fecha_dt = datetime.strptime(fecha_celda.split()[0], '%Y-%m-%d')
                except:
                    try:
                        fecha_dt = datetime.strptime(fecha_celda.split()[0], '%d/%m/%Y')
                    except:
                        pass
            
            if fecha_dt and fecha_dt.year == 2026 and fecha_dt.month == 1 and fecha_dt.day == 14:
                print(f"\nFila encontrada: {fila_idx}")
                print(f"Fecha: {fecha_dt.strftime('%d/%m/%Y')}")
                
                # Verificar datos generales
                print("\nDatos generales:")
                print(f"  E (Imp. Inicial): {ws.cell(row=fila_idx, column=5).value}")
                print(f"  F (Imp. Final): {ws.cell(row=fila_idx, column=6).value}")
                print(f"  G (Benef. €): {ws.cell(row=fila_idx, column=7).value}")
                print(f"  H (Benef. %): {ws.cell(row=fila_idx, column=8).value}")
                
                # Verificar fórmulas del cliente 1
                wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
                ws_formulas = wb_formulas['Diario VIP']
                
                print("\nCliente 1 - Fórmulas:")
                print(f"  N (Saldo Diario): {ws_formulas.cell(row=fila_idx, column=14).value}")
                print(f"  O (Benef. € Diario): {ws_formulas.cell(row=fila_idx, column=15).value}")
                print(f"  P (Benef. % Diario): {ws_formulas.cell(row=fila_idx, column=16).value}")
                
                break
        except Exception as e:
            pass
