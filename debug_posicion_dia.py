#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de prueba para verificar por qué posicion_dia es None
"""

from openpyxl import load_workbook
from datetime import datetime, date, time

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

ws_formulas = None
ws_valores_obj = None
for sheet_name in wb_formulas.sheetnames:
    if 'VIP' in sheet_name.upper():
        ws_formulas = wb_formulas[sheet_name]
        ws_valores_obj = wb_valores[sheet_name]
        break

# Agrupar por fecha
fechas_agrupadas = {}
for fila_idx in range(15, 25):
    fecha_valor = ws_valores_obj.cell(row=fila_idx, column=4).value
    print(f"Fila {fila_idx}: fecha_valor={fecha_valor}, tipo={type(fecha_valor)}")
    if fecha_valor is not None:
        fecha_str = None
        if isinstance(fecha_valor, datetime):
            fecha_str = fecha_valor.strftime('%Y-%m-%d')
        elif isinstance(fecha_valor, date):
            fecha_str = datetime.combine(fecha_valor, time.min).strftime('%Y-%m-%d')
        elif isinstance(fecha_valor, str):
            if ' ' in fecha_valor:
                fecha_str = fecha_valor.split()[0]
            else:
                fecha_str = fecha_valor
        
        print(f"  fecha_str normalizada: {fecha_str}")
        if fecha_str:
            if fecha_str not in fechas_agrupadas:
                fechas_agrupadas[fecha_str] = []
            fechas_agrupadas[fecha_str].append(fila_idx)

# Ordenar
for fecha_str in fechas_agrupadas:
    fechas_agrupadas[fecha_str].sort()

print("\n" + "="*80)
print("FECHAS AGRUPADAS:")
print("="*80)
for fecha_str, filas in sorted(fechas_agrupadas.items()):
    print(f"{fecha_str}: {filas}")

# Probar obtener_posicion_dia
print("\n" + "="*80)
print("PRUEBA DE obtener_posicion_dia:")
print("="*80)

def obtener_posicion_dia(fila_idx, fecha_valor):
    fecha_str = None
    if isinstance(fecha_valor, datetime):
        fecha_str = fecha_valor.strftime('%Y-%m-%d')
    elif isinstance(fecha_valor, date):
        fecha_str = datetime.combine(fecha_valor, time.min).strftime('%Y-%m-%d')
    elif isinstance(fecha_valor, str):
        if ' ' in fecha_valor:
            fecha_str = fecha_valor.split()[0]
        else:
            fecha_str = fecha_valor
    
    print(f"  obtener_posicion_dia({fila_idx}, {fecha_valor}) -> fecha_str={fecha_str}")
    if fecha_str and fecha_str in fechas_agrupadas:
        filas_del_dia = fechas_agrupadas[fecha_str]
        try:
            pos = filas_del_dia.index(fila_idx) + 1
            print(f"    -> posición: {pos}")
            return pos
        except:
            print(f"    -> ERROR: fila {fila_idx} no encontrada en {filas_del_dia}")
            return None
    else:
        print(f"    -> ERROR: fecha_str={fecha_str} no está en fechas_agrupadas")
        return None

for fila in [15, 16, 17]:
    fecha_valor = ws_valores_obj.cell(row=fila, column=4).value
    pos = obtener_posicion_dia(fila, fecha_valor)
    print(f"Fila {fila}: posicion_dia = {pos}\n")

wb_formulas.close()
wb_valores.close()
