from openpyxl import load_workbook
import json
from datetime import datetime

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

def identificar_clientes_por_columnas(ws):
    """Identifica qué columnas pertenecen a cada cliente"""
    # Fila 13 tiene los números de cliente
    fila_nombres = 13
    # Fila 14 tiene los headers
    fila_header = 14
    
    # Leer números de clientes de la fila 13
    clientes_columnas = {}
    
    for col_idx in range(1, ws.max_column + 1):
        celda_numero = ws.cell(row=fila_nombres, column=col_idx)
        valor_numero = celda_numero.value
        
        if valor_numero is not None:
            try:
                numero_cliente = int(valor_numero)
                if numero_cliente > 0:
                    # Este cliente empieza en esta columna
                    # El grupo de columnas para este cliente es aproximadamente:
                    # INCREMENTO (columna actual), DECREMENTO (+1), INICIAL (+2), FINAL (+3), 
                    # BENEF. € (+4), BENEF. % (+5), BENEF. € acum (+6), BENEF. % acum (+7)
                    
                    grupo_columnas = {
                        'numero': numero_cliente,
                        'columna_inicio': col_idx,
                        'INCREMENTO': col_idx,      # K (11 para cliente 1)
                        'DECREMENTO': col_idx + 1,  # L (12 para cliente 1)
                        'INICIAL': col_idx + 2,      # M (13 para cliente 1)
                        'FINAL': col_idx + 3,        # N (14 para cliente 1)
                        'BENEF_€_DIARIO': col_idx + 4,  # O (15 para cliente 1)
                        'BENEF_%_DIARIO': col_idx + 5,  # P (16 para cliente 1)
                        'BENEF_€_ACUM': col_idx + 6,    # Q (17 para cliente 1)
                        'BENEF_%_ACUM': col_idx + 7     # R (18 para cliente 1)
                    }
                    
                    clientes_columnas[numero_cliente] = grupo_columnas
            except:
                pass
    
    return clientes_columnas

def calcular_valores_cliente(ws, numero_cliente, grupo_columnas, fila_inicio=15):
    """Calcula los valores solicitados para un cliente específico"""
    valores = {
        'numero_cliente': numero_cliente,
        'incrementos': [],
        'decrementos': [],
        'saldos_finales': [],
        'beneficios_diarios': [],
        'beneficios_diarios_pct': [],
        'beneficios_acumulados': [],
        'beneficios_acumulados_pct': []
    }
    
    # Leer todas las filas desde fila_inicio hasta el final
    # No necesitamos verificar MES, solo leer los valores de las columnas del cliente
    for fila_idx in range(fila_inicio, ws.max_row + 1):
        
        # Leer valores de las columnas del cliente
        incremento = ws.cell(row=fila_idx, column=grupo_columnas['INCREMENTO']).value
        decremento = ws.cell(row=fila_idx, column=grupo_columnas['DECREMENTO']).value
        final = ws.cell(row=fila_idx, column=grupo_columnas['FINAL']).value
        benef_diario = ws.cell(row=fila_idx, column=grupo_columnas['BENEF_€_DIARIO']).value
        benef_diario_pct = ws.cell(row=fila_idx, column=grupo_columnas['BENEF_%_DIARIO']).value
        benef_acum = ws.cell(row=fila_idx, column=grupo_columnas['BENEF_€_ACUM']).value
        benef_acum_pct = ws.cell(row=fila_idx, column=grupo_columnas['BENEF_%_ACUM']).value
        
        # Agregar valores no nulos
        if incremento is not None:
            try:
                valores['incrementos'].append(float(incremento))
            except:
                pass
        
        if decremento is not None:
            try:
                valores['decrementos'].append(float(decremento))
            except:
                pass
        
        if final is not None:
            try:
                valores['saldos_finales'].append(float(final))
            except:
                pass
        
        if benef_diario is not None:
            try:
                valores['beneficios_diarios'].append(float(benef_diario))
            except:
                pass
        
        if benef_diario_pct is not None:
            try:
                valores['beneficios_diarios_pct'].append(float(benef_diario_pct))
            except:
                pass
        
        if benef_acum is not None:
            try:
                valores['beneficios_acumulados'].append(float(benef_acum))
            except:
                pass
        
        if benef_acum_pct is not None:
            try:
                valores['beneficios_acumulados_pct'].append(float(benef_acum_pct))
            except:
                pass
    
    # Calcular resumen
    # Para saldo final y beneficios acumulados, tomar el último valor NO CERO
    def ultimo_no_cero(lista):
        # Buscar desde el final hacia atrás el primer valor no cero
        for valor in reversed(lista):
            if valor != 0:
                return valor
        return lista[-1] if lista else None
    
    resumen = {
        'numero_cliente': numero_cliente,
        'incrementos_total': sum(valores['incrementos']),
        'decrementos_total': sum(valores['decrementos']),
        'saldo_final': ultimo_no_cero(valores['saldos_finales']) if valores['saldos_finales'] else None,
        'beneficio_diario': valores['beneficios_diarios'][-1] if valores['beneficios_diarios'] else None,
        'beneficio_diario_pct': valores['beneficios_diarios_pct'][-1] if valores['beneficios_diarios_pct'] else None,
        'beneficio_acumulado': ultimo_no_cero(valores['beneficios_acumulados']) if valores['beneficios_acumulados'] else None,
        'beneficio_acumulado_pct': ultimo_no_cero(valores['beneficios_acumulados_pct']) if valores['beneficios_acumulados_pct'] else None
    }
    
    return resumen

def procesar_hoja(nombre_hoja):
    """Procesa una hoja completa con cálculos por cliente"""
    if nombre_hoja not in wb.sheetnames:
        return None
    
    ws = wb[nombre_hoja]
    
    resultado = {
        'nombre': nombre_hoja,
        'datos_generales': [],
        'clientes': []
    }
    
    # 1. Extraer datos generales (filas 3-6)
    filas_generales = [3, 4, 5, 6]
    for fila_idx in filas_generales:
        fila_data = {
            'fila': fila_idx,
            'columnas': {}
        }
        
        for col_idx in range(1, min(100, ws.max_column + 1)):
            celda = ws.cell(row=fila_idx, column=col_idx)
            valor = celda.value
            if valor is not None:
                fila_data['columnas'][col_idx] = valor
        
        if len(fila_data['columnas']) > 0:
            resultado['datos_generales'].append(fila_data)
    
    # 2. Identificar clientes por columnas
    clientes_columnas = identificar_clientes_por_columnas(ws)
    
    # 3. Calcular valores para cada cliente
    for numero_cliente, grupo_columnas in clientes_columnas.items():
        resumen = calcular_valores_cliente(ws, numero_cliente, grupo_columnas)
        resultado['clientes'].append(resumen)
    
    # Ordenar clientes por número
    resultado['clientes'].sort(key=lambda x: x['numero_cliente'])
    
    return resultado

# Procesar ambas hojas
resultado_completo = {
    'fecha_extraccion': datetime.now().isoformat(),
    'hojas': {}
}

print("=" * 80)
print("EXTRACCION CON CALCULOS POR CLIENTE")
print("=" * 80)

# Procesar Diario STD
if 'Diario STD' in wb.sheetnames:
    print("\nProcesando Diario STD...")
    resultado_completo['hojas']['Diario STD'] = procesar_hoja('Diario STD')
    if resultado_completo['hojas']['Diario STD']:
        print(f"  Clientes encontrados: {len(resultado_completo['hojas']['Diario STD']['clientes'])}")

# Procesar Diario VIP
if 'Diario VIP' in wb.sheetnames:
    print("\nProcesando Diario VIP...")
    resultado_completo['hojas']['Diario VIP'] = procesar_hoja('Diario VIP')
    if resultado_completo['hojas']['Diario VIP']:
        print(f"  Clientes encontrados: {len(resultado_completo['hojas']['Diario VIP']['clientes'])}")
        
        # Mostrar ejemplo del Cliente 1
        if resultado_completo['hojas']['Diario VIP']['clientes']:
            cliente1 = resultado_completo['hojas']['Diario VIP']['clientes'][0]
            print(f"\n  Ejemplo - Cliente {cliente1['numero_cliente']}:")
            print(f"    Incrementos: {cliente1['incrementos_total']:.2f}")
            print(f"    Decrementos: {cliente1['decrementos_total']:.2f}")
            print(f"    Saldo Final: {cliente1['saldo_final']}")
            print(f"    Beneficio Diario: {cliente1['beneficio_diario']}")
            print(f"    Beneficio Diario %: {cliente1['beneficio_diario_pct']}")
            print(f"    Beneficio Acumulado: {cliente1['beneficio_acumulado']}")
            print(f"    Beneficio Acumulado %: {cliente1['beneficio_acumulado_pct']}")

# Guardar resultado
with open('datos_completos.json', 'w', encoding='utf-8') as f:
    json.dump(resultado_completo, f, ensure_ascii=False, indent=2, default=str)

print("\n" + "=" * 80)
print(f"Datos guardados en: datos_completos.json")
print("=" * 80)
