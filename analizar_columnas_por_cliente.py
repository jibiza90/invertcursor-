from openpyxl import load_workbook

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb.sheetnames:
    ws = wb['Diario VIP']
    
    print("=" * 80)
    print("ANALISIS DE COLUMNAS POR CLIENTE - DIARIO VIP")
    print("=" * 80)
    
    # Fila 14 tiene los headers
    fila_header = 14
    
    print("\nHeaders de la fila 14:")
    headers_por_columna = {}
    for col_idx in range(1, min(50, ws.max_column + 1)):
        celda = ws.cell(row=fila_header, column=col_idx)
        valor = celda.value
        if valor is not None and str(valor).strip() != '':
            headers_por_columna[col_idx] = str(valor).strip()
            print(f"  Columna {col_idx}: {valor}")
    
    # Identificar grupos de columnas por cliente
    # Parece que las columnas 3-6 son comunes (MES, FECHA, IMP. INICIAL, IMP. FINAL)
    # Luego hay grupos que se repiten
    
    print("\n" + "=" * 80)
    print("IDENTIFICANDO GRUPOS DE COLUMNAS POR CLIENTE")
    print("=" * 80)
    
    # Buscar patrones en la fila 13 que tiene "Nombre cliente"
    fila_nombres = 13
    print(f"\nFila {fila_nombres} (Nombres de clientes):")
    grupos_clientes = []
    
    for col_idx in range(1, min(100, ws.max_column + 1)):
        celda = ws.cell(row=fila_nombres, column=col_idx)
        valor = celda.value
        if valor is not None:
            valor_str = str(valor).strip()
            if 'cliente' in valor_str.lower() or valor_str.isdigit():
                # Buscar el grupo de columnas asociado
                # Parece que cada cliente tiene un grupo de columnas
                print(f"  Columna {col_idx}: {valor_str}")
                
                # Identificar el rango de columnas para este cliente
                # Buscar hacia atrás y adelante para encontrar el grupo
                inicio_grupo = max(1, col_idx - 5)
                fin_grupo = min(ws.max_column, col_idx + 10)
                
                grupo_columnas = []
                for c in range(inicio_grupo, fin_grupo + 1):
                    if c in headers_por_columna:
                        grupo_columnas.append(c)
                
                if grupo_columnas:
                    grupos_clientes.append({
                        'numero': valor_str,
                        'columna_nombre': col_idx,
                        'columnas': grupo_columnas
                    })
    
    print(f"\nGrupos encontrados: {len(grupos_clientes)}")
    for grupo in grupos_clientes[:5]:  # Mostrar primeros 5
        print(f"  Cliente {grupo['numero']}: columnas {grupo['columnas'][:10]}")
    
    # Analizar estructura de un cliente específico (Cliente 1)
    print("\n" + "=" * 80)
    print("ANALISIS DETALLADO CLIENTE 1")
    print("=" * 80)
    
    # Buscar columna K (columna 11) que es INCREMENTO del primer grupo
    # Buscar columna L (columna 12) que es DECREMENTO del primer grupo
    # Buscar columna N (columna 14) que es FINAL del primer grupo
    # Buscar columna O (columna 15) que es BENEF. € del primer grupo
    # Buscar columna P (columna 16) que es BENEF. % del primer grupo
    # Buscar columna Q (columna 17) que es BENEF. € acumulado del primer grupo
    # Buscar columna R (columna 18) que es BENEF. % acumulado del primer grupo
    
    print("\nMapeo de columnas para Cliente 1:")
    mapeo_cliente1 = {
        'INCREMENTO': 11,  # Columna K
        'DECREMENTO': 12,  # Columna L
        'FINAL': 14,       # Columna N
        'BENEF. € (Diario)': 15,  # Columna O
        'BENEF. % (Diario)': 16,  # Columna P
        'BENEF. € (Acumulado)': 17,  # Columna Q
        'BENEF. % (Acumulado)': 18   # Columna R
    }
    
    for nombre, col in mapeo_cliente1.items():
        header = headers_por_columna.get(col, 'N/A')
        print(f"  {nombre} (Columna {col}): {header}")
    
    # Leer datos del Cliente 1 (filas desde 15 en adelante con MES=1)
    print("\nDatos del Cliente 1 (primeras 10 filas):")
    filas_cliente1 = []
    for fila_idx in range(15, min(25, ws.max_row + 1)):
        mes_celda = ws.cell(row=fila_idx, column=3)
        mes_valor = mes_celda.value
        
        if mes_valor == 1:
            fila_data = {'fila': fila_idx}
            for nombre, col in mapeo_cliente1.items():
                celda = ws.cell(row=fila_idx, column=col)
                fila_data[nombre] = celda.value
            filas_cliente1.append(fila_data)
            
            if len(filas_cliente1) <= 5:
                print(f"  Fila {fila_idx}:")
                for nombre, valor in fila_data.items():
                    if nombre != 'fila':
                        print(f"    {nombre}: {valor}")
    
    # Calcular valores solicitados
    print("\n" + "=" * 80)
    print("CALCULOS PARA CLIENTE 1")
    print("=" * 80)
    
    incrementos = []
    decrementos = []
    saldos_finales = []
    beneficios_diarios = []
    beneficios_diarios_pct = []
    beneficios_acumulados = []
    beneficios_acumulados_pct = []
    
    for fila_data in filas_cliente1:
        if fila_data.get('INCREMENTO') is not None:
            try:
                incrementos.append(float(fila_data['INCREMENTO']))
            except:
                pass
        if fila_data.get('DECREMENTO') is not None:
            try:
                decrementos.append(float(fila_data['DECREMENTO']))
            except:
                pass
        if fila_data.get('FINAL') is not None:
            try:
                saldos_finales.append(float(fila_data['FINAL']))
            except:
                pass
        if fila_data.get('BENEF. € (Diario)') is not None:
            try:
                beneficios_diarios.append(float(fila_data['BENEF. € (Diario)']))
            except:
                pass
        if fila_data.get('BENEF. % (Diario)') is not None:
            try:
                beneficios_diarios_pct.append(float(fila_data['BENEF. % (Diario)']))
            except:
                pass
        if fila_data.get('BENEF. € (Acumulado)') is not None:
            try:
                beneficios_acumulados.append(float(fila_data['BENEF. € (Acumulado)']))
            except:
                pass
        if fila_data.get('BENEF. % (Acumulado)') is not None:
            try:
                beneficios_acumulados_pct.append(float(fila_data['BENEF. % (Acumulado)']))
            except:
                pass
    
    print(f"\nIncrementos (suma columna K): {sum(incrementos):.2f}")
    print(f"Decrementos (suma columna L): {sum(decrementos):.2f}")
    if saldos_finales:
        print(f"Saldo Final (último columna N): {saldos_finales[-1]:.2f}")
    if beneficios_diarios:
        print(f"Beneficio Diario (columna O): {beneficios_diarios[-1]:.2f}")
    if beneficios_diarios_pct:
        print(f"Beneficio Diario % (columna P): {beneficios_diarios_pct[-1]:.2f}")
    if beneficios_acumulados:
        print(f"Beneficio Acumulado (último columna Q): {beneficios_acumulados[-1]:.2f}")
    if beneficios_acumulados_pct:
        print(f"Beneficio Acumulado % (último columna R): {beneficios_acumulados_pct[-1]:.2f}")
