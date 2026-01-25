from openpyxl import load_workbook

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb.sheetnames:
    ws = wb['Diario VIP']
    
    print("=" * 80)
    print("VERIFICANDO IDENTIFICACION DE COLUMNAS POR CLIENTE")
    print("=" * 80)
    
    # Fila 13 tiene los números de cliente
    fila_nombres = 13
    
    print("\nFila 13 - Números de clientes:")
    clientes_encontrados = {}
    
    for col_idx in range(1, min(100, ws.max_column + 1)):
        celda = ws.cell(row=fila_nombres, column=col_idx)
        valor = celda.value
        
        if valor is not None:
            try:
                numero_cliente = int(valor)
                if numero_cliente > 0:
                    print(f"  Columna {col_idx}: Cliente {numero_cliente}")
                    clientes_encontrados[numero_cliente] = col_idx
            except:
                pass
    
    print(f"\nTotal clientes encontrados: {len(clientes_encontrados)}")
    
    # Verificar Cliente 1
    if 1 in clientes_encontrados:
        col_inicio = clientes_encontrados[1]
        print(f"\nCliente 1 empieza en columna: {col_inicio}")
        print(f"  Columna K (INCREMENTO) debería ser: {col_inicio}")
        print(f"  Columna L (DECREMENTO) debería ser: {col_inicio + 1}")
        print(f"  Columna N (FINAL) debería ser: {col_inicio + 3}")
        print(f"  Columna O (BENEF. €) debería ser: {col_inicio + 4}")
        print(f"  Columna P (BENEF. %) debería ser: {col_inicio + 5}")
        print(f"  Columna Q (BENEF. € acum) debería ser: {col_inicio + 6}")
        print(f"  Columna R (BENEF. % acum) debería ser: {col_inicio + 7}")
        
        # Verificar headers en fila 14
        print(f"\nVerificando headers en fila 14:")
        print(f"  Columna {col_inicio}: {ws.cell(row=14, column=col_inicio).value}")
        print(f"  Columna {col_inicio + 1}: {ws.cell(row=14, column=col_inicio + 1).value}")
        print(f"  Columna {col_inicio + 3}: {ws.cell(row=14, column=col_inicio + 3).value}")
        print(f"  Columna {col_inicio + 4}: {ws.cell(row=14, column=col_inicio + 4).value}")
        print(f"  Columna {col_inicio + 5}: {ws.cell(row=14, column=col_inicio + 5).value}")
        print(f"  Columna {col_inicio + 6}: {ws.cell(row=14, column=col_inicio + 6).value}")
        print(f"  Columna {col_inicio + 7}: {ws.cell(row=14, column=col_inicio + 7).value}")
