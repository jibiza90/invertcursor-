from openpyxl import load_workbook
import json
from datetime import datetime

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

resultado_completo = {
    'fecha_extraccion': datetime.now().isoformat(),
    'hojas': {}
}

# Función para extraer datos generales
def extraer_datos_generales(ws, inicio=2, fin=6):
    """Extrae los datos generales/totales de las primeras filas"""
    datos_generales = []
    
    for fila_idx in range(inicio, min(fin + 1, ws.max_row + 1)):
        fila = []
        for col_idx in range(1, min(20, ws.max_column + 1)):  # Primeras 20 columnas
            celda = ws.cell(row=fila_idx, column=col_idx)
            valor = celda.value
            fila.append(valor)
        
        # Solo agregar si tiene valores significativos
        valores_no_vacios = [v for v in fila if v is not None and str(v).strip() != '']
        if len(valores_no_vacios) > 2:
            datos_generales.append({
                'fila': fila_idx,
                'datos': fila
            })
    
    return datos_generales

# Función para extraer datos de clientes
def extraer_datos_clientes(ws, fila_header=14):
    """Extrae los datos de clientes desde la fila de headers"""
    # Leer header
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        celda = ws.cell(row=fila_header, column=col_idx)
        valor = celda.value
        if valor is not None and str(valor).strip() != '':
            headers[col_idx] = str(valor).strip()
    
    # Leer datos de clientes
    datos_clientes = []
    inicio_datos = fila_header + 1
    
    for fila_idx in range(inicio_datos, ws.max_row + 1):
        # Verificar si esta fila tiene un MES válido
        celda_mes = ws.cell(row=fila_idx, column=3)  # Columna C (MES)
        mes_valor = celda_mes.value
        
        if mes_valor is None:
            continue
        
        try:
            mes_num = int(mes_valor)
            if mes_num <= 0:
                continue
        except:
            continue
        
        # Leer toda la fila
        fila_dict = {}
        tiene_datos = False
        
        for col_idx in range(1, ws.max_column + 1):
            celda = ws.cell(row=fila_idx, column=col_idx)
            valor = celda.value
            header = headers.get(col_idx, f'Columna_{col_idx}')
            
            # Determinar si está bloqueada
            bloqueada = False
            if header in ['MES', 'FECHA', 'IMP. INICIAL']:
                bloqueada = True
            elif col_idx < 4:
                bloqueada = True
            elif header.startswith('Columna_'):
                bloqueada = True
            
            if valor is not None or header in ['MES', 'FECHA', 'IMP. INICIAL', 'IMP. FINAL']:
                fila_dict[header] = {
                    'valor': valor,
                    'bloqueada': bloqueada,
                    'columna': col_idx,
                    'fila': fila_idx
                }
                
                if valor is not None and str(valor).strip() != '':
                    tiene_datos = True
        
        if tiene_datos or len(fila_dict) > 0:
            datos_clientes.append(fila_dict)
    
    return {
        'headers': {str(k): v for k, v in headers.items()},
        'datos_clientes': datos_clientes
    }

# Procesar Diario STD
if 'Diario STD' in wb.sheetnames:
    print("Procesando Diario STD...")
    ws_std = wb['Diario STD']
    
    datos_generales_std = extraer_datos_generales(ws_std, 2, 12)
    datos_clientes_std = extraer_datos_clientes(ws_std, 14)
    
    resultado_completo['hojas']['Diario STD'] = {
        'tipo': 'STD',
        'total_filas': ws_std.max_row,
        'total_columnas': ws_std.max_column,
        'datos_generales': datos_generales_std,
        'headers': datos_clientes_std['headers'],
        'datos_clientes': datos_clientes_std['datos_clientes']
    }
    
    print(f"  - Datos generales: {len(datos_generales_std)} filas")
    print(f"  - Clientes: {len(datos_clientes_std['datos_clientes'])} registros")

# Procesar Diario VIP
if 'Diario VIP' in wb.sheetnames:
    print("Procesando Diario VIP...")
    ws_vip = wb['Diario VIP']
    
    datos_generales_vip = extraer_datos_generales(ws_vip, 2, 12)
    datos_clientes_vip = extraer_datos_clientes(ws_vip, 14)
    
    resultado_completo['hojas']['Diario VIP'] = {
        'tipo': 'VIP',
        'total_filas': ws_vip.max_row,
        'total_columnas': ws_vip.max_column,
        'datos_generales': datos_generales_vip,
        'headers': datos_clientes_vip['headers'],
        'datos_clientes': datos_clientes_vip['datos_clientes']
    }
    
    print(f"  - Datos generales: {len(datos_generales_vip)} filas")
    print(f"  - Clientes: {len(datos_clientes_vip['datos_clientes'])} registros")

# Guardar resultado
with open('datos_completos.json', 'w', encoding='utf-8') as f:
    json.dump(resultado_completo, f, ensure_ascii=False, indent=2, default=str)

print(f"\nDatos extraidos y guardados en: datos_completos.json")
print(f"Total de hojas procesadas: {len(resultado_completo['hojas'])}")
