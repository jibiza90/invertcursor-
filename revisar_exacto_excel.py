#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para revisar EXACTAMENTE qué celdas están bloqueadas y cuáles no en el Excel
y comparar con el JSON generado
"""

import json
import openpyxl
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

def obtener_nombre_columna(num_col):
    result = ""
    num = num_col
    while num > 0:
        num -= 1
        result = chr(65 + (num % 26)) + result
        num //= 26
    return result

def tiene_formula(ws, fila, columna):
    try:
        celda = ws.cell(row=fila, column=columna)
        return celda.data_type == 'f' and celda.value is not None
    except:
        return False

print("="*80)
print("REVISIÓN EXACTA DE CELDAS BLOQUEADAS/EDITABLES EN EXCEL")
print("="*80)

# Cargar Excel
print("\n📂 Cargando Excel...")
wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = None
for sheet_name in wb.sheetnames:
    if 'VIP' in sheet_name.upper():
        ws = wb[sheet_name]
        print(f"✅ Hoja encontrada: {sheet_name}")
        break

if not ws:
    ws = wb.active
    print(f"⚠️ Usando hoja activa: {ws.title}")

# Cargar JSON
print("\n📂 Cargando JSON...")
with open('datos_completos.json', 'r', encoding='utf-8') as f:
    datos_json = json.load(f)

hoja_vip = datos_json.get('hojas', {}).get('Diario VIP')
if not hoja_vip:
    print("❌ ERROR: No se encontró hoja Diario VIP")
    sys.exit(1)

datos_diarios = hoja_vip.get('datos_diarios_generales', [])

# Mapeo de columnas
columnas_map = {
    'imp_inicial': 5,      # E
    'imp_final': 6,         # F
    'benef_euro': 7,        # G
    'benef_porcentaje': 8,  # H
    'benef_euro_acum': 9,   # I
    'benef_porcentaje_acum': 10  # J
}

print("\n" + "="*80)
print("ANÁLISIS FILA POR FILA (15-50)")
print("="*80)

errores = []
estadisticas = {
    'total_filas': 0,
    'imp_final_editable_excel': 0,
    'imp_final_bloqueada_excel': 0,
    'imp_final_editable_json': 0,
    'imp_final_bloqueada_json': 0,
    'discrepancias': []
}

# Analizar filas 15-50 como muestra
for fila in range(15, 51):
    # Buscar en Excel
    fecha_excel = ws.cell(row=fila, column=4).value
    if fecha_excel is None:
        continue
    
    # Buscar en JSON
    dato_json = next((d for d in datos_diarios if d.get('fila') == fila), None)
    if not dato_json:
        continue
    
    estadisticas['total_filas'] += 1
    
    print(f"\n{'='*80}")
    print(f"FILA {fila} (Fecha Excel: {fecha_excel})")
    print(f"{'='*80}")
    
    # Analizar cada columna
    for campo, col_num in columnas_map.items():
        col_letra = obtener_nombre_columna(col_num)
        
        # Estado en Excel
        tiene_formula_excel = tiene_formula(ws, fila, col_num)
        estado_excel = "🔒 BLOQUEADA" if tiene_formula_excel else "✏️ EDITABLE"
        
        # Estado en JSON
        bloqueadas_json = dato_json.get('bloqueadas', {})
        formulas_json = dato_json.get('formulas', {})
        esta_bloqueada_json = bloqueadas_json.get(campo, False)
        tiene_formula_json = campo in formulas_json and formulas_json[campo]
        estado_json = "🔒 BLOQUEADA" if esta_bloqueada_json else "✏️ EDITABLE"
        
        # Comparar
        coincide = (tiene_formula_excel == esta_bloqueada_json)
        
        print(f"\n{campo.upper()} (Col {col_letra}):")
        print(f"  Excel: {estado_excel}", end="")
        if tiene_formula_excel:
            celda = ws.cell(row=fila, column=col_num)
            print(f" - Fórmula: {celda.value}")
        else:
            print()
        
        print(f"  JSON:  {estado_json}", end="")
        if tiene_formula_json:
            print(f" - Fórmula: {formulas_json[campo]}")
        else:
            print()
        
        if not coincide:
            error = {
                'fila': fila,
                'campo': campo,
                'columna': col_letra,
                'excel_tiene_formula': tiene_formula_excel,
                'excel_editable': not tiene_formula_excel,
                'json_bloqueada': esta_bloqueada_json,
                'json_editable': not esta_bloqueada_json,
                'problema': f'Excel: {"BLOQUEADA" if tiene_formula_excel else "EDITABLE"}, JSON: {"BLOQUEADA" if esta_bloqueada_json else "EDITABLE"}'
            }
            errores.append(error)
            estadisticas['discrepancias'].append(error)
            print(f"  ❌ DISCREPANCIA: {error['problema']}")
        else:
            print(f"  ✅ COINCIDE")
        
        # Estadísticas para imp_final
        if campo == 'imp_final':
            if tiene_formula_excel:
                estadisticas['imp_final_bloqueada_excel'] += 1
            else:
                estadisticas['imp_final_editable_excel'] += 1
            
            if esta_bloqueada_json:
                estadisticas['imp_final_bloqueada_json'] += 1
            else:
                estadisticas['imp_final_editable_json'] += 1

# Resumen
print("\n" + "="*80)
print("RESUMEN")
print("="*80)
print(f"Total de filas analizadas: {estadisticas['total_filas']}")
print(f"\nIMP_FINAL en Excel:")
print(f"  Editables: {estadisticas['imp_final_editable_excel']}")
print(f"  Bloqueadas: {estadisticas['imp_final_bloqueada_excel']}")
print(f"\nIMP_FINAL en JSON:")
print(f"  Editables: {estadisticas['imp_final_editable_json']}")
print(f"  Bloqueadas: {estadisticas['imp_final_bloqueada_json']}")
print(f"\nTotal de discrepancias encontradas: {len(errores)}")

if errores:
    print("\n" + "="*80)
    print("ERRORES ENCONTRADOS")
    print("="*80)
    for error in errores[:20]:  # Mostrar primeros 20
        print(f"\nFila {error['fila']}, {error['campo']} (Col {error['columna']}):")
        print(f"  {error['problema']}")
    
    # Guardar reporte completo
    with open('reporte_discrepancias_excel_json.json', 'w', encoding='utf-8') as f:
        json.dump({
            'estadisticas': estadisticas,
            'errores': errores
        }, f, indent=2, ensure_ascii=False)
    print(f"\n📄 Reporte completo guardado en 'reporte_discrepancias_excel_json.json'")
else:
    print("\n✅ PERFECTO: Todas las celdas coinciden entre Excel y JSON")

# Análisis específico de imp_final por día
print("\n" + "="*80)
print("ANÁLISIS DE IMP_FINAL POR DÍA")
print("="*80)

fechas_agrupadas = defaultdict(list)
for fila in range(15, 51):
    fecha_excel = ws.cell(row=fila, column=4).value
    if fecha_excel:
        fecha_str = str(fecha_excel)
        tiene_formula_imp_final = tiene_formula(ws, fila, 6)
        fechas_agrupadas[fecha_str].append({
            'fila': fila,
            'imp_final_editable': not tiene_formula_imp_final
        })

for fecha_str, filas_info in sorted(fechas_agrupadas.items()):
    filas_editables = [f for f in filas_info if f['imp_final_editable']]
    print(f"\nFecha: {fecha_str}")
    print(f"  Total filas: {len(filas_info)}")
    print(f"  Filas con imp_final EDITABLE: {len(filas_editables)}")
    if filas_editables:
        print(f"    Filas editables: {[f['fila'] for f in filas_editables]}")
    else:
        print(f"    ⚠️ Ninguna fila tiene imp_final editable")

wb.close()
