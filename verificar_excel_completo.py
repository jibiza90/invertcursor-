from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("VERIFICACION DETALLADA - FILAS 15-50")
print("=" * 80)

# Verificar cada fila individualmente
for fila_idx in range(15, 51):
    fecha_celda = ws.cell(row=fila_idx, column=4)
    if fecha_celda.value is None:
        continue
    
    print(f"\nFila {fila_idx}:")
    
    # Verificar cada columna G, H, I, J
    columnas = {
        7: 'G - Benef. €',
        8: 'H - Benef. %',
        9: 'I - Benef. € Acum',
        10: 'J - Benef. % Acum'
    }
    
    tiene_alguna_formula = False
    for col_idx, nombre_col in columnas.items():
        celda = ws.cell(row=fila_idx, column=col_idx)
        
        tiene_formula = False
        formula_texto = None
        
        # Verificar tipo de dato
        tipo_dato = celda.data_type
        valor_raw = celda.value
        
        if tipo_dato == 'f':
            tiene_formula = True
            formula_texto = str(valor_raw)
            tiene_alguna_formula = True
        elif valor_raw is not None:
            valor_str = str(valor_raw).strip()
            if valor_str.startswith('='):
                tiene_formula = True
                formula_texto = valor_str
                tiene_alguna_formula = True
        
        estado = "[BLOQUEADA]" if tiene_formula else "[EDITABLE]"
        print(f"  {nombre_col}: tipo={tipo_dato}, tiene_formula={tiene_formula} {estado}")
        if formula_texto:
            print(f"    Formula: {formula_texto[:70]}")
