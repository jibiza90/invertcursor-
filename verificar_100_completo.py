from openpyxl import load_workbook
import json
import traceback
from datetime import datetime
import sys

# Forzar escritura inmediata
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Abrir archivo de salida desde el inicio
output_file = open('reporte_verificacion_100porciento.txt', 'w', encoding='utf-8', buffering=1)

def escribir(mensaje):
    """Escribe tanto a consola como a archivo"""
    try:
        print(mensaje, flush=True)
        output_file.write(mensaje + '\n')
        output_file.flush()
    except:
        pass

escribir("=" * 80)
escribir("VERIFICACION COMPLETA: 100% DE LAS CELDAS")
escribir(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
escribir("=" * 80)

try:
    escribir("\n[PASO 1/5] Cargando Excel...")
    wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
    ws_formulas = wb_formulas['Diario VIP']
    max_fila = ws_formulas.max_row
    max_col = ws_formulas.max_column
    escribir(f"   ✓ Excel cargado")
    escribir(f"   - Max fila: {max_fila}")
    escribir(f"   - Max columna: {max_col}")
    
    escribir("\n[PASO 2/5] Cargando JSON...")
    with open('datos_completos.json', 'r', encoding='utf-8') as f:
        datos_json = json.load(f)
    
    hoja_json = datos_json['hojas']['Diario VIP']
    clientes_json = hoja_json.get('clientes', [])
    datos_diarios_json = hoja_json.get('datos_diarios_generales', [])
    datos_generales_json = hoja_json.get('datos_generales', [])
    
    escribir(f"   ✓ JSON cargado")
    escribir(f"   - Datos generales: {len(datos_generales_json)} filas")
    escribir(f"   - Datos diarios: {len(datos_diarios_json)} filas")
    escribir(f"   - Clientes: {len(clientes_json)}")
    
    # Mapeos
    col_general = {5: 'imp_inicial', 6: 'imp_final', 7: 'benef_euro', 
                   8: 'benef_porcentaje', 9: 'benef_euro_acum', 10: 'benef_porcentaje_acum'}
    offset_cliente = {0: 'incremento', 1: 'decremento', 2: 'base', 3: 'saldo_diario',
                      4: 'beneficio_diario', 5: 'beneficio_diario_pct',
                      6: 'beneficio_acumulado', 7: 'beneficio_acumulado_pct'}
    
    def col_a_letra(n):
        """Convierte número de columna a letra Excel"""
        result = ""
        while n > 0:
            n -= 1
            result = chr(65 + (n % 26)) + result
            n //= 26
        return result
    
    errores = []
    advertencias = []
    total_celdas = 0
    
    # 3. Verificar datos generales (FILAS 3-6, COLUMNAS E-J)
    escribir("\n[PASO 3/5] Verificando DATOS GENERALES (Filas 3-6, Columnas E-J)...")
    celdas_generales = 0
    for fila in range(3, 7):
        fila_json = next((d for d in datos_generales_json if d.get('fila') == fila), None)
        if not fila_json:
            errores.append(f"Fila {fila}: No encontrada en JSON (datos_generales)")
            continue
        
        for col, campo in col_general.items():
            total_celdas += 1
            celdas_generales += 1
            celda = ws_formulas.cell(row=fila, column=col)
            tiene_formula = celda.data_type == 'f' and celda.value is not None
            
            if tiene_formula:
                formula_excel = str(celda.value)
                formula_json = fila_json.get('formulas', {}).get(campo)
                bloqueada = fila_json.get('bloqueadas', {}).get(campo, False)
                
                if not formula_json:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Excel tiene formula pero JSON NO")
                elif formula_excel != formula_json:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Formulas diferentes")
                    errores.append(f"  Excel: {formula_excel[:150]}")
                    errores.append(f"  JSON:  {formula_json[:150]}")
                if not bloqueada:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Tiene formula pero NO bloqueada")
    
    escribir(f"   ✓ Verificadas {celdas_generales} celdas de datos generales")
    
    # 4. Verificar datos diarios generales (TODAS LAS FILAS DESDE 15, COLUMNAS E-J)
    escribir("\n[PASO 4/5] Verificando DATOS DIARIOS GENERALES (Todas las filas desde 15, Columnas E-J)...")
    celdas_diarias = 0
    filas_procesadas = 0
    
    # Crear mapa de filas JSON para acceso rápido
    mapa_filas_json = {d.get('fila'): d for d in datos_diarios_json}
    
    # Verificar TODAS las filas desde 15 hasta max_fila
    for fila in range(15, max_fila + 1):
        # Verificar si esta fila tiene fecha (es una fila de datos)
        fecha_excel = ws_formulas.cell(row=fila, column=4).value
        if fecha_excel is None:
            continue
        
        fila_json = mapa_filas_json.get(fila)
        if not fila_json:
            advertencias.append(f"Fila {fila} (datos_diarios): No encontrada en JSON")
            continue
        
        filas_procesadas += 1
        
        # Verificar TODAS las columnas E-J
        for col, campo in col_general.items():
            total_celdas += 1
            celdas_diarias += 1
            celda = ws_formulas.cell(row=fila, column=col)
            tiene_formula = celda.data_type == 'f' and celda.value is not None
            
            if tiene_formula:
                formula_excel = str(celda.value)
                formula_json = fila_json.get('formulas', {}).get(campo)
                bloqueada = fila_json.get('bloqueadas', {}).get(campo, False)
                
                if not formula_json:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Excel tiene formula pero JSON NO")
                elif formula_excel != formula_json:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Formulas diferentes")
                    errores.append(f"  Excel: {formula_excel[:150]}")
                    errores.append(f"  JSON:  {formula_json[:150]}")
                if not bloqueada:
                    errores.append(f"Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Tiene formula pero NO bloqueada")
    
    escribir(f"   ✓ Verificadas {celdas_diarias} celdas de datos diarios generales")
    escribir(f"   ✓ Procesadas {filas_procesadas} filas con datos")
    
    # 5. Verificar TODOS LOS CLIENTES (TODAS LAS FILAS DESDE 15, TODAS LAS COLUMNAS DE CLIENTES)
    escribir("\n[PASO 5/5] Verificando TODOS LOS CLIENTES (Todas las filas desde 15, Columnas K+)...")
    
    # Identificar TODAS las columnas de clientes desde Excel (fila 13 tiene números)
    clientes_cols = {}
    for col in range(11, min(max_col + 1, 500)):  # K en adelante, hasta columna 500
        try:
            num_cliente = ws_formulas.cell(row=13, column=col).value
            if num_cliente is not None:
                try:
                    num = int(float(num_cliente))
                    if num > 0:
                        if num not in clientes_cols:
                            clientes_cols[num] = []
                        clientes_cols[num].append(col)
                except:
                    pass
        except:
            pass
    
    escribir(f"   ✓ Encontrados {len(clientes_cols)} clientes en Excel")
    
    celdas_clientes = 0
    clientes_procesados = 0
    
    # Crear mapa de clientes JSON para acceso rápido
    mapa_clientes_json = {c.get('numero_cliente'): c for c in clientes_json}
    
    # Verificar CADA CLIENTE encontrado en Excel
    for num_cliente, columnas in sorted(clientes_cols.items()):
        cliente_json = mapa_clientes_json.get(num_cliente)
        if not cliente_json:
            advertencias.append(f"Cliente {num_cliente}: Encontrado en Excel pero NO en JSON")
            continue
        
        clientes_procesados += 1
        datos_cliente = cliente_json.get('datos_diarios', [])
        col_inicio = min(columnas)
        
        # Crear mapa de filas del cliente para acceso rápido
        mapa_filas_cliente = {d.get('fila'): d for d in datos_cliente}
        
        escribir(f"   Verificando Cliente {num_cliente} (columnas {col_a_letra(col_inicio)}-{col_a_letra(col_inicio+7)})...")
        
        # Verificar TODAS las filas desde 15 hasta max_fila
        filas_cliente_procesadas = 0
        for fila in range(15, max_fila + 1):
            # Verificar si esta fila tiene fecha
            fecha_excel = ws_formulas.cell(row=fila, column=4).value
            if fecha_excel is None:
                continue
            
            fila_data = mapa_filas_cliente.get(fila)
            if not fila_data:
                continue
            
            filas_cliente_procesadas += 1
            
            # Verificar TODAS las 8 columnas del cliente (incremento, decremento, base, etc.)
            for offset, campo in offset_cliente.items():
                col = col_inicio + offset
                if col > max_col:
                    break
                
                total_celdas += 1
                celdas_clientes += 1
                
                celda = ws_formulas.cell(row=fila, column=col)
                tiene_formula = celda.data_type == 'f' and celda.value is not None
                
                if tiene_formula:
                    formula_excel = str(celda.value)
                    formula_json = fila_data.get('formulas', {}).get(campo)
                    bloqueada = fila_data.get('bloqueadas', {}).get(campo, False)
                    
                    if not formula_json:
                        errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Excel tiene formula pero JSON NO")
                    elif formula_excel != formula_json:
                        errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Formulas diferentes")
                        errores.append(f"  Excel: {formula_excel[:150]}")
                        errores.append(f"  JSON:  {formula_json[:150]}")
                    if not bloqueada:
                        errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): Tiene formula pero NO bloqueada")
                else:
                    # Incremento y decremento NUNCA deben estar bloqueados
                    if campo in ['incremento', 'decremento']:
                        bloqueada = fila_data.get('bloqueadas', {}).get(campo, False)
                        if bloqueada:
                            errores.append(f"Cliente {num_cliente}, Fila {fila}, {col_a_letra(col)}{fila} ({campo}): NO debe estar bloqueada (es editable)")
        
        escribir(f"     ✓ Cliente {num_cliente}: {filas_cliente_procesadas} filas, {filas_cliente_procesadas * 8} celdas")
    
    escribir(f"   ✓ Verificadas {celdas_clientes} celdas de clientes")
    escribir(f"   ✓ Procesados {clientes_procesados} clientes")
    
    # Resumen final
    escribir("\n" + "=" * 80)
    escribir("RESUMEN FINAL - VERIFICACION 100% COMPLETA")
    escribir("=" * 80)
    escribir(f"Total de celdas verificadas: {total_celdas:,}")
    escribir(f"  - Datos generales (Filas 3-6): {celdas_generales:,} celdas")
    escribir(f"  - Datos diarios generales (Filas 15-{max_fila}): {celdas_diarias:,} celdas")
    escribir(f"  - Celdas de clientes (Filas 15-{max_fila}): {celdas_clientes:,} celdas")
    escribir(f"\nErrores críticos encontrados: {len(errores)}")
    escribir(f"Advertencias: {len(advertencias)}")
    
    if errores:
        escribir("\n" + "=" * 80)
        escribir("ERRORES CRÍTICOS")
        escribir("=" * 80)
        for i, error in enumerate(errores, 1):
            escribir(f"{i}. {error}")
    else:
        escribir("\n✅ NO SE ENCONTRARON ERRORES CRÍTICOS")
    
    if advertencias:
        escribir("\n" + "=" * 80)
        escribir("ADVERTENCIAS")
        escribir("=" * 80)
        for i, adv in enumerate(advertencias[:200], 1):
            escribir(f"{i}. {adv}")
        if len(advertencias) > 200:
            escribir(f"\n... y {len(advertencias) - 200} advertencias más")
    
    escribir("\n" + "=" * 80)
    escribir("VERIFICACION COMPLETADA AL 100%")
    escribir("=" * 80)
    escribir(f"Fecha finalización: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
except Exception as e:
    error_msg = f"\nERROR CRÍTICO: {e}\n{traceback.format_exc()}"
    escribir(error_msg)
    print(error_msg)

finally:
    output_file.close()
    print("\n" + "=" * 80)
    print("Reporte guardado en: reporte_verificacion_100porciento.txt")
    print("=" * 80)
