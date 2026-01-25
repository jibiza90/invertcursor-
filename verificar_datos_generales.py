from openpyxl import load_workbook

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb.sheetnames:
    ws = wb['Diario VIP']
    
    print("=" * 80)
    print("VERIFICANDO DATOS GENERALES - DIARIO VIP")
    print("=" * 80)
    
    # Verificar filas 3-6
    filas_generales = [3, 4, 5, 6]
    
    print("\nFila 14 - Headers de columnas:")
    for col_idx in range(1, 15):
        header = ws.cell(row=14, column=col_idx).value
        if header:
            print(f"  Columna {col_idx}: {header}")
    
    print("\n" + "=" * 80)
    print("DATOS GENERALES (Filas 3-6):")
    print("=" * 80)
    
    for fila_idx in filas_generales:
        print(f"\nFila {fila_idx}:")
        
        # Fecha (columna D)
        fecha = ws.cell(row=fila_idx, column=4).value
        print(f"  Fecha (D): {fecha}")
        
        # Columnas E a J
        headers_fila14 = {}
        for col in range(5, 11):
            header = ws.cell(row=14, column=col).value
            if header:
                headers_fila14[col] = header
        
        for col_idx in range(5, 11):
            valor = ws.cell(row=fila_idx, column=col_idx).value
            header = ws.cell(row=14, column=col_idx).value
            if valor is not None or header:
                print(f"  Columna {col_idx} ({header if header else 'Sin header'}): {valor}")
    
    # Verificar también fila 13 que tiene los números de clientes
    print("\n" + "=" * 80)
    print("FILA 13 (Números de clientes):")
    print("=" * 80)
    for col_idx in range(1, 20):
        valor = ws.cell(row=13, column=col_idx).value
        if valor is not None:
            print(f"  Columna {col_idx}: {valor}")
