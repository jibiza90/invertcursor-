import json
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print('Falta openpyxl. Instala con: pip install openpyxl')
    raise

ARCHIVO_XLSX = r"C:\Users\Jordi\Desktop\InvertCursor\JIG 120126 v1 JIG.xlsx"
ARCHIVO_JSON = r"C:\Users\Jordi\Desktop\InvertCursor\datos_mensuales\Diario_STD_2026-01.json"
ARCHIVO_JSON_FULL = r"C:\Users\Jordi\Desktop\InvertCursor\datos_editados.json"
HOJA_XLSX = "Diario STD"
CLIENTE_IDX = 1  # Cliente 2 (0-index)
FECHA_OBJETIVO = "2026-01-05"


def parse_fecha(valor):
    if not valor:
        return None
    if isinstance(valor, str):
        try:
            return datetime.fromisoformat(valor.split(' ')[0])
        except Exception:
            return None
    return None


def col_letter(n):
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def normalize(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    return val


def main():
    with open(ARCHIVO_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    clientes = data.get('clientes', [])
    if len(clientes) <= CLIENTE_IDX:
        print('No existe cliente en JSON')
        return
    cliente = clientes[CLIENTE_IDX]
    col_inicio = cliente.get('columna_inicio')
    if not col_inicio:
        try:
            with open(ARCHIVO_JSON_FULL, 'r', encoding='utf-8') as f:
                data_full = json.load(f)
            cliente_full = data_full.get('hojas', {}).get(HOJA_XLSX, {}).get('clientes', [])[CLIENTE_IDX]
            col_inicio = cliente_full.get('columna_inicio')
        except Exception:
            col_inicio = None
    if not col_inicio:
        print('Cliente sin columna_inicio')
        return

    wb = load_workbook(ARCHIVO_XLSX, data_only=True)
    if HOJA_XLSX not in wb.sheetnames:
        print('Hoja no encontrada en XLSX')
        return
    ws = wb[HOJA_XLSX]

    datos = [d for d in (cliente.get('datos_diarios') or []) if d.get('fecha')]
    filas_dia = []
    for d in datos:
        fecha = parse_fecha(d.get('fecha'))
        if fecha and fecha.strftime('%Y-%m-%d') == FECHA_OBJETIVO:
            filas_dia.append(d)
    filas_dia.sort(key=lambda x: x.get('fila', 0))

    if not filas_dia:
        print('No hay filas para la fecha objetivo')
        return

    print(f'Comparando Cliente {CLIENTE_IDX + 1} fecha {FECHA_OBJETIVO}')
    print(f'Columna inicio: {col_inicio} ({col_letter(col_inicio)})')

    campos = [
        ('incremento', 0),
        ('decremento', 1),
        ('base', 2),
        ('saldo_diario', 3),
        ('beneficio_diario', 4),
        ('beneficio_diario_pct', 5),
        ('beneficio_acumulado', 6),
        ('beneficio_acumulado_pct', 7),
    ]

    for d in filas_dia:
        fila = d.get('fila')
        print(f'--- Fila {fila}')
        for campo, offset in campos:
            col = col_inicio + offset
            celda = f"{col_letter(col)}{fila}"
            xlsx_val = normalize(ws[celda].value)
            json_val = normalize(d.get(campo))
            if xlsx_val is None and json_val is None:
                estado = 'OK'
            else:
                diff = None
                if isinstance(xlsx_val, (int, float)) and isinstance(json_val, (int, float)):
                    diff = abs(xlsx_val - json_val)
                    estado = 'OK' if diff <= 0.01 else 'DIFF'
                else:
                    estado = 'OK' if xlsx_val == json_val else 'DIFF'
            print(f"{campo:24s} {celda:6s} XLSX={xlsx_val} JSON={json_val} => {estado}")


if __name__ == '__main__':
    main()
