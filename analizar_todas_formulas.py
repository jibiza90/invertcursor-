from openpyxl import load_workbook
from datetime import datetime

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
wb_valores = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

print("=" * 80)
print("ANALISIS COMPLETO DE FORMULAS EN DIARIO VIP")
print("=" * 80)

if 'Diario VIP' in wb_formulas.sheetnames:
    ws_formulas = wb_formulas['Diario VIP']
    ws_valores = wb_valores['Diario VIP']
    
    # Columnas a verificar
    columnas = {
        5: 'E - Imp. Inicial',
        6: 'F - Imp. Final',
        7: 'G - Benef. €',
        8: 'H - Benef. %',
        9: 'I - Benef. € Acum',
        10: 'J - Benef. % Acum'
    }
    
    # Analizar desde fila 15 hasta el final
    filas_con_formulas = {}
    
    for fila_idx in range(15, min(400, ws_valores.max_row + 1)):
        fecha_celda = ws_valores.cell(row=fila_idx, column=4)
        if fecha_celda.value is None:
            continue
            
        formulas_en_fila = {}
        tiene_alguna_formula = False
        
        for col_idx, nombre_col in columnas.items():
            celda_formula = ws_formulas.cell(row=fila_idx, column=col_idx)
            celda_valor = ws_valores.cell(row=fila_idx, column=col_idx)
            
            tiene_formula = False
            formula_texto = None
            
            if celda_formula.value is not None:
                valor_str = str(celda_formula.value).strip()
                if valor_str.startswith('='):
                    tiene_formula = True
                    formula_texto = valor_str[:80]  # Primeros 80 caracteres
                    tiene_alguna_formula = True
            
            formulas_en_fila[nombre_col] = {
                'tiene_formula': tiene_formula,
                'formula': formula_texto,
                'valor': celda_valor.value
            }
        
        if tiene_alguna_formula:
            filas_con_formulas[fila_idx] = formulas_en_fila
    
    print(f"\nTotal de filas con fórmulas encontradas: {len(filas_con_formulas)}")
    print("\n" + "=" * 80)
    print("DETALLE POR FILA:")
    print("=" * 80)
    
    # Mostrar primeras 20 filas con fórmulas
    contador = 0
    for fila_idx in sorted(filas_con_formulas.keys()):
        if contador >= 20:
            print(f"\n... y {len(filas_con_formulas) - 20} filas más")
            break
        
        fecha_valor = ws_valores.cell(row=fila_idx, column=4).value
        fecha_str = fecha_valor.strftime('%Y-%m-%d') if isinstance(fecha_valor, datetime) else str(fecha_valor)
        
        print(f"\nFila {fila_idx} (Fecha: {fecha_str}):")
        for nombre_col, info in formulas_en_fila.items():
            estado = "[BLOQUEADA]" if info['tiene_formula'] else "[EDITABLE]"
            print(f"  {nombre_col}: {estado}")
            if info['tiene_formula']:
                print(f"    Formula: {info['formula']}")
        
        contador += 1
    
    # Resumen por columna
    print("\n" + "=" * 80)
    print("RESUMEN POR COLUMNA:")
    print("=" * 80)
    
    for nombre_col in columnas.values():
        total_con_formula = sum(1 for filas in filas_con_formulas.values() 
                               if filas[nombre_col]['tiene_formula'])
        print(f"{nombre_col}: {total_con_formula} filas con fórmula")
