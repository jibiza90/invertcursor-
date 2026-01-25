from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

if 'Diario VIP' in wb.sheetnames:
    ws = wb['Diario VIP']
    
    print("=" * 80)
    print("VERIFICANDO INCREMENTOS CON FILTRO DE FECHAS 2026")
    print("=" * 80)
    
    # Cliente 1 empieza en columna 11 (K)
    col_incremento = 11  # K
    col_decremento = 12  # L
    col_fecha = 4  # D
    
    # Fechas límite
    fecha_inicio = datetime(2026, 1, 1)
    fecha_fin = datetime(2026, 12, 31)
    
    incrementos = []
    decrementos = []
    fechas_encontradas = []
    
    print(f"\nBuscando incrementos/decrementos desde fila 15 hasta 1130")
    print(f"Filtro de fechas: {fecha_inicio.strftime('%d/%m/%Y')} a {fecha_fin.strftime('%d/%m/%Y')}")
    
    for fila_idx in range(15, min(1131, ws.max_row + 1)):
        fecha_valor = ws.cell(row=fila_idx, column=col_fecha).value
        
        if fecha_valor:
            try:
                # Convertir fecha a datetime si es necesario
                if isinstance(fecha_valor, datetime):
                    fecha_dt = fecha_valor
                elif isinstance(fecha_valor, str):
                    fecha_dt = datetime.strptime(fecha_valor.split()[0], '%Y-%m-%d')
                else:
                    fecha_dt = fecha_valor
                
                # Verificar si está en el rango 2026
                if isinstance(fecha_dt, datetime):
                    if fecha_inicio <= fecha_dt <= fecha_fin:
                        fechas_encontradas.append(fecha_dt)
                        
                        incremento = ws.cell(row=fila_idx, column=col_incremento).value
                        decremento = ws.cell(row=fila_idx, column=col_decremento).value
                        
                        if incremento is not None:
                            try:
                                inc_val = float(incremento)
                                if inc_val != 0:
                                    incrementos.append((fila_idx, fecha_dt, inc_val))
                            except:
                                pass
                        
                        if decremento is not None:
                            try:
                                dec_val = float(decremento)
                                if dec_val != 0:
                                    decrementos.append((fila_idx, fecha_dt, dec_val))
                            except:
                                pass
            except Exception as e:
                pass
    
    print(f"\nFechas encontradas en rango 2026: {len(fechas_encontradas)}")
    print(f"Incrementos encontrados: {len(incrementos)}")
    print(f"Decrementos encontrados: {len(decrementos)}")
    
    suma_incrementos = sum([x[2] for x in incrementos])
    suma_decrementos = sum([x[2] for x in decrementos])
    
    print(f"\nSUMA INCREMENTOS (solo 2026): {suma_incrementos:,.2f} €")
    print(f"SUMA DECREMENTOS (solo 2026): {suma_decrementos:,.2f} €")
    
    print(f"\nPrimeros 10 incrementos:")
    for fila, fecha, valor in incrementos[:10]:
        print(f"  Fila {fila}, Fecha {fecha.strftime('%d/%m/%Y')}: {valor:,.2f} €")
