from openpyxl import load_workbook
import json
from datetime import datetime

wb = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=True)

# Buscar Diario VIP
sheet_name = None
for sheet in wb.sheetnames:
    if 'Diario VIP' in sheet or 'Diario_VIP' in sheet or 'diario_vip' in sheet.lower():
        sheet_name = sheet
        break

if sheet_name:
    ws = wb[sheet_name]
    
    # Leer todas las filas
    todas_las_filas = []
    for row in ws.iter_rows(values_only=True):
        todas_las_filas.append(list(row))
    
    # Identificar headers (fila 14, índice 13)
    headers = {}
    fila_headers_idx = 13  # Fila 14 (0-indexed)
    if fila_headers_idx < len(todas_las_filas):
        fila_headers = todas_las_filas[fila_headers_idx]
        for col_idx, valor in enumerate(fila_headers):
            if valor is not None and str(valor).strip() != '':
                headers[col_idx] = str(valor).strip()
    
    # Definir qué columnas están bloqueadas (solo lectura)
    columnas_bloqueadas = ['MES', 'FECHA', 'IMP. INICIAL']
    
    # Leer datos desde la fila 15 en adelante (datos de clientes)
    datos_clientes = []
    inicio_datos = 14  # Fila 15 (0-indexed)
    
    for fila_idx in range(inicio_datos, len(todas_las_filas)):
        fila = todas_las_filas[fila_idx]
        fila_dict = {}
        tiene_datos = False
        
        # Verificar si esta fila tiene un MES (identificador de cliente)
        tiene_mes = False
        for col_idx, valor in enumerate(fila):
            header = headers.get(col_idx, f'Columna_{col_idx+1}')
            if header == 'MES' and valor is not None:
                try:
                    if int(valor) > 0:
                        tiene_mes = True
                        break
                except:
                    pass
        
        # Solo procesar filas que tengan MES válido
        if not tiene_mes:
            continue
        
        for col_idx, valor in enumerate(fila):
            header = headers.get(col_idx, f'Columna_{col_idx+1}')
            
            # Determinar si la celda está bloqueada
            bloqueada = False
            if header in columnas_bloqueadas:
                bloqueada = True
            elif col_idx < 4:  # Primeras columnas generalmente bloqueadas
                bloqueada = True
            elif header.startswith('Columna_'):
                # Columnas sin nombre generalmente bloqueadas
                bloqueada = True
            
            # Solo incluir si tiene valor o es un campo importante
            if valor is not None or header in ['MES', 'FECHA', 'IMP. INICIAL', 'IMP. FINAL']:
                fila_dict[header] = {
                    'valor': valor,
                    'bloqueada': bloqueada,
                    'columna': col_idx + 1,
                    'fila': fila_idx + 1
                }
                
                if valor is not None and str(valor).strip() != '':
                    tiene_datos = True
        
        if tiene_datos or len(fila_dict) > 0:
            datos_clientes.append(fila_dict)
    
    # Crear estructura de datos completa
    resultado = {
        'hoja': sheet_name,
        'total_filas': len(todas_las_filas),
        'total_columnas': len(todas_las_filas[0]) if todas_las_filas else 0,
        'headers': {str(k): v for k, v in headers.items()},
        'datos_clientes': datos_clientes,
        'configuracion': {
            'columnas_bloqueadas': columnas_bloqueadas,
            'columnas_editables': ['BENEF. €', 'BENEF. %', 'INCREMENTO', 'DECREMENTO', 'IMP. FINAL']
        },
        'fecha_extraccion': datetime.now().isoformat()
    }
    
    # Guardar en JSON
    with open('datos_completos.json', 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"Datos extraídos exitosamente")
    print(f"Total de filas de clientes: {len(datos_clientes)}")
    print(f"Archivo guardado: datos_completos.json")
    
    # Mostrar ejemplo del primer cliente
    if datos_clientes:
        print(f"\nEjemplo - Cliente 1:")
        cliente_ejemplo = datos_clientes[0]
        campos_importantes = ['MES', 'FECHA', 'IMP. INICIAL', 'IMP. FINAL', 'BENEF. €']
        for campo in campos_importantes:
            if campo in cliente_ejemplo:
                valor = cliente_ejemplo[campo]['valor']
                bloqueada = cliente_ejemplo[campo]['bloqueada']
                print(f"  {campo}: {valor} ({'Bloqueada' if bloqueada else 'Editable'})")
