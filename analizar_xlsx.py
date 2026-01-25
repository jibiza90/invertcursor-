import openpyxl
from openpyxl.utils import get_column_letter

# Cargar el archivo Excel
wb = openpyxl.load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)

# Analizar la hoja "Diario STD"
if 'Diario STD' in wb.sheetnames:
    ws = wb['Diario STD']
    
    output = []
    output.append("=== Análisis de Diario STD ===\n")
    
    # Ver las primeras filas del cliente 1 (columnas K-R, filas 15-20)
    output.append("Cliente 1 - Columnas K-R, Filas 15-20:")
    output.append("Columna K=Incremento, L=Decremento, M=Base, N=Saldo, O-R=Beneficios\n")
    
    for row in range(15, 21):
        row_info = f"Fila {row}: "
        for col in range(11, 19):  # K=11 hasta R=18
            cell = ws.cell(row=row, column=col)
            col_letter = get_column_letter(col)
            value = cell.value
            formula = cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
            protection = cell.protection.locked if cell.protection else 'N/A'
            
            if formula:
                row_info += f"{col_letter}=[FORMULA:{formula[:20]}...] "
            elif value is not None:
                row_info += f"{col_letter}={value} "
            else:
                row_info += f"{col_letter}=VACIO "
            
            # Indicar si está bloqueada
            if protection:
                row_info += "(BLOQ) "
            else:
                row_info += "(EDIT) "
        output.append(row_info)
    
    # Verificar si la hoja tiene protección
    output.append(f"\n¿Hoja protegida? {ws.protection.sheet}")
    
    with open('xlsx_analysis.txt', 'w', encoding='utf-8') as f:
        f.write('\n'.join(output))
    
    print("Análisis guardado en xlsx_analysis.txt")
else:
    print("Hoja 'Diario STD' no encontrada")
    print(f"Hojas disponibles: {wb.sheetnames}")
