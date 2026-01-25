from openpyxl import load_workbook

wb_formulas = load_workbook('JIG 120126 v1 JIG.xlsx', data_only=False)
ws = wb_formulas['Diario VIP']

print("=" * 80)
print("VERIFICACION DIRECTA DEL EXCEL - FILA POR FILA")
print("=" * 80)

# Verificar filas específicas
filas_verificar = [15, 16, 17, 27, 51, 54]

for fila_idx in filas_verificar:
    print(f"\n{'='*80}")
    print(f"FILA {fila_idx}:")
    print(f"{'='*80}")
    
    fecha_celda = ws.cell(row=fila_idx, column=4)
    print(f"Fecha (col D): {fecha_celda.value}")
    
    columnas = {
        5: 'E - Imp. Inicial',
        6: 'F - Imp. Final',
        7: 'G - Benef. €',
        8: 'H - Benef. %',
        9: 'I - Benef. € Acum',
        10: 'J - Benef. % Acum'
    }
    
    for col_idx, nombre in columnas.items():
        celda = ws.cell(row=fila_idx, column=col_idx)
        valor_raw = celda.value
        
        tiene_formula = False
        formula_texto = None
        
        if valor_raw is not None:
            valor_str = str(valor_raw).strip()
            if valor_str.startswith('='):
                tiene_formula = True
                formula_texto = valor_str
        
        # También verificar el tipo de dato
        tipo_dato = celda.data_type
        
        estado = "[BLOQUEADA]" if tiene_formula else "[EDITABLE]"
        print(f"\n{nombre} (Col {chr(64+col_idx)}):")
        print(f"  Valor raw: {valor_raw}")
        print(f"  Tipo dato: {tipo_dato}")
        print(f"  Tiene formula: {tiene_formula} {estado}")
        if formula_texto:
            print(f"  Formula completa: {formula_texto}")
